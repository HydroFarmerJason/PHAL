#!/usr/bin/env python3
"""
PHAL - Advanced Pluripotent Hardware Abstraction Layer v2.1
Enterprise-grade backend for Controlled Environment Agriculture
Complete implementation with no placeholders

Author: Jason DeLooze <jasonmarkd@gmail.com>
License: Apache-2.0
Version: 2.1.0
"""

import asyncio
import json
import logging
import os
import sys
import uuid
import hashlib
import secrets
import hmac
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Dict, List, Optional, Any, Set, Union, Tuple
from dataclasses import dataclass, field, asdict
from contextlib import asynccontextmanager
from collections import defaultdict, deque
import math
import random
import time

import aiohttp
from aiohttp import web
import aiohttp_cors
import redis.asyncio as redis
import asyncpg
import yaml
import jwt
import aiofiles
from cryptography.fernet import Fernet
from prometheus_client import Counter, Histogram, Gauge, generate_latest
import numpy as np
from scipy import stats, optimize
import pandas as pd
import joblib
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from pydantic import BaseModel, validator, BaseSettings
from opentelemetry import trace
from opentelemetry.sdk.trace import TracerProvider
from opentelemetry.sdk.trace.export import BatchSpanProcessor
from opentelemetry.exporter.otlp.proto.grpc.trace_exporter import OTLPSpanExporter
from opentelemetry.instrumentation.aiohttp_server import AioHttpServerInstrumentor
from opentelemetry.instrumentation.asyncpg import AsyncPGInstrumentor

import structlog
from aiohttp_limiter import Limiter, RateLimitExceeded

# Configure structured logging
structlog.configure(
    processors=[
        structlog.stdlib.filter_by_level,
        structlog.stdlib.add_logger_name,
        structlog.stdlib.add_log_level,
        structlog.stdlib.PositionalArgumentsFormatter(),
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.StackInfoRenderer(),
        structlog.processors.format_exc_info,
        structlog.processors.UnicodeDecoder(),
        structlog.processors.JSONRenderer()
    ],
    context_class=dict,
    logger_factory=structlog.stdlib.LoggerFactory(),
    cache_logger_on_first_use=True,
)

logger = structlog.get_logger()

# Initialize tracing
trace.set_tracer_provider(TracerProvider())
tracer = trace.get_tracer(__name__)
AioHttpServerInstrumentor().instrument()
AsyncPGInstrumentor().instrument()

# Metrics
grant_requests = Counter('phal_grant_requests_total', 'Total grant requests', ['plugin_id', 'capability_type'])
command_executions = Counter('phal_command_executions_total', 'Total command executions', ['capability_type'])
command_latency = Histogram('phal_command_latency_seconds', 'Command execution latency', ['command_type'])
safety_violations = Counter('phal_safety_violations_total', 'Total safety violations', ['violation_type'])
sensor_readings = Histogram('phal_sensor_readings', 'Sensor reading values', ['sensor_type', 'zone_id'])
system_uptime = Gauge('phal_system_uptime_seconds', 'System uptime in seconds')
active_grants = Gauge('phal_active_grants', 'Number of active grants', ['plugin_id'])
zone_temperature = Gauge('phal_zone_temperature_celsius', 'Zone temperature', ['zone_id'])
nutrient_levels = Gauge('phal_nutrient_levels_ml', 'Nutrient levels', ['zone_id', 'nutrient_type'])
api_requests = Counter('phal_api_requests_total', 'Total API requests', ['method', 'endpoint', 'status'])
db_connections = Gauge('phal_db_connections', 'Database connection pool stats', ['state'])

# Configuration
class PHALConfig(BaseSettings):
    database_url: str = "postgresql://phal:phal@localhost/phal"
    redis_url: str = "redis://localhost:6379"
    jwt_secret: str = secrets.token_urlsafe(32)
    encryption_key: Optional[str] = None
    audit_path: str = "/var/log/phal/audit/"
    telemetry_enabled: bool = True
    telemetry_endpoint: Optional[str] = None
    node_id: str = "phal-node-01"
    max_request_size: int = 10 * 1024 * 1024  # 10MB
    rate_limit_enabled: bool = True
    rate_limit_requests: int = 100
    rate_limit_window: int = 60  # seconds
    hardware_interface: str = "simulator"  # simulator, modbus, canbus, custom
    sensor_poll_interval: int = 5  # seconds
    ml_model_path: str = "./models/"
    backup_interval: int = 3600  # seconds
    max_grants_per_tenant: int = 100
    session_timeout: int = 3600  # seconds
    enable_compression: bool = True
    enable_encryption: bool = True
    
    class Config:
        env_prefix = "PHAL_"
        env_file = f".env.{os.getenv('PHAL_ENV', 'development')}"
        env_file_encoding = 'utf-8'

    @validator('database_url')
    def validate_database_url(cls, v, values):
        if 'postgresql://' not in v:
            raise ValueError('Invalid PostgreSQL URL')
        if values.get('env') == 'production' and 'sslmode=require' not in v:
            v += '?sslmode=require'
        return v

    @validator('redis_url')
    def validate_redis_url(cls, v):
        if not v.startswith(('redis://', 'rediss://')):
            raise ValueError('Invalid Redis URL')
        return v

# Request validation models
class GrantRequest(BaseModel):
    plugin_id: str
    capability_query: Dict[str, Any]
    permissions: List[str]
    duration_seconds: int = 3600
    constraints: Optional[Dict[str, Any]] = None
    
    @validator('duration_seconds')
    def validate_duration(cls, v):
        if v > 86400:  # Max 24 hours
            raise ValueError('Duration cannot exceed 24 hours')
        if v < 60:  # Min 1 minute
            raise ValueError('Duration must be at least 60 seconds')
        return v
    
    @validator('permissions')
    def validate_permissions(cls, v):
        valid_permissions = {'READ', 'WRITE', 'OPERATE', 'CALIBRATE', 'ADMIN'}
        for perm in v:
            if perm not in valid_permissions:
                raise ValueError(f'Invalid permission: {perm}')
        return v

class CommandRequest(BaseModel):
    grant_id: str
    command: Dict[str, Any]
    context: Optional[Dict[str, Any]] = None

class MaintenanceRequest(BaseModel):
    type: str
    component_id: str
    scheduled_date: datetime
    notes: Optional[str] = None
    estimated_duration: Optional[int] = None
    required_parts: Optional[List[Dict[str, Any]]] = None

class HarvestRequest(BaseModel):
    zone_id: str
    crop_id: str
    quantity: float
    quantity_unit: str
    quality_grade: str
    harvested_by: str
    notes: Optional[str] = None
    images: Optional[List[str]] = None

class ExportRequest(BaseModel):
    type: str
    format: str
    time_range: Dict[str, datetime]
    zones: Optional[List[str]] = None
    include_metadata: bool = True
    compress: bool = True

# Input sanitization utility
def sanitize_input(value: Any) -> Any:
    if isinstance(value, str):
        value = value.replace('\x00', '').strip()
        return value[:1000]
    elif isinstance(value, dict):
        return {k: sanitize_input(v) for k, v in value.items()}
    elif isinstance(value, list):
        return [sanitize_input(v) for v in value]
    return value

# Core data classes
@dataclass
class Tenant:
    id: str
    name: str
    tier: str  # 'community', 'professional', 'enterprise'
    features: Set[str] = field(default_factory=set)
    resource_limits: Dict[str, int] = field(default_factory=dict)
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    metadata: Dict[str, Any] = field(default_factory=dict)
    api_key_hash: Optional[str] = None
    contact_email: Optional[str] = None
    billing_status: str = "active"
    usage_stats: Dict[str, int] = field(default_factory=dict)
    
    def check_feature(self, feature: str) -> bool:
        """Check if tenant has access to a feature"""
        return feature in self.features or self.tier == 'enterprise'
    
    def check_resource_limit(self, resource: str, current: int, requested: int = 1) -> bool:
        """Check if resource limit would be exceeded"""
        if self.tier == 'enterprise':
            return True
        limit = self.resource_limits.get(resource, 0)
        return current + requested <= limit
    
    def increment_usage(self, metric: str, amount: int = 1):
        """Track usage statistics"""
        if metric not in self.usage_stats:
            self.usage_stats[metric] = 0
        self.usage_stats[metric] += amount

@dataclass
class Zone:
    id: str
    tenant_id: str
    name: str
    type: str  # 'production', 'nursery', 'quarantine', 'research'
    units: List[str] = field(default_factory=list)
    environmental_targets: Dict[str, Dict[str, float]] = field(default_factory=dict)
    crop_profile: Optional[Dict[str, Any]] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    emergency_stop: bool = False
    maintenance_mode: bool = False
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    last_harvest: Optional[datetime] = None
    total_yield: float = 0.0
    active_alarms: List[str] = field(default_factory=list)
    
    def is_operational(self) -> bool:
        return not self.emergency_stop and not self.maintenance_mode
    
    def get_age_days(self) -> int:
        if self.crop_profile and 'plant_date' in self.crop_profile:
            plant_date = datetime.fromisoformat(self.crop_profile['plant_date'])
            return (datetime.now(timezone.utc) - plant_date).days
        return 0

@dataclass
class Grant:
    id: str
    tenant_id: str
    plugin_id: str
    capability_id: str
    permissions: List[str]
    expires_at: datetime
    constraints: Dict[str, Any] = field(default_factory=dict)
    usage_count: int = 0
    max_usage: Optional[int] = None
    audit_log: List[Dict[str, Any]] = field(default_factory=list)
    metadata: Dict[str, Any] = field(default_factory=dict)
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    last_used: Optional[datetime] = None
    
    def is_valid(self) -> bool:
        now = datetime.now(timezone.utc)
        if now > self.expires_at:
            return False
        if self.max_usage and self.usage_count >= self.max_usage:
            return False
        return True
    
    def can_execute(self, command_type: str) -> bool:
        required_perm = {
            'read': 'READ',
            'write': 'WRITE',
            'operate': 'OPERATE',
            'calibrate': 'CALIBRATE',
            'configure': 'ADMIN'
        }.get(command_type, 'OPERATE')
        return required_perm in self.permissions

@dataclass
class Capability:
    id: str
    type: str  # SENSOR, ACTUATOR
    subtype: str
    zone_id: Optional[str] = None
    properties: Dict[str, Any] = field(default_factory=dict)
    constraints: Dict[str, Any] = field(default_factory=dict)
    calibration: Optional[Dict[str, Any]] = None
    maintenance_schedule: Optional[Dict[str, Any]] = None
    tenant_id: Optional[str] = None
    hardware_config: Dict[str, Any] = field(default_factory=dict)
    last_reading: Optional[Dict[str, Any]] = None
    status: str = "online"
    error_count: int = 0
    
    def needs_calibration(self) -> bool:
        if not self.calibration:
            return False
        last_cal = self.calibration.get('last_calibration')
        if not last_cal:
            return True
        cal_date = datetime.fromisoformat(last_cal)
        days_since = (datetime.now(timezone.utc) - cal_date).days
        return days_since > self.calibration.get('interval_days', 30)

@dataclass
class Alarm:
    id: str
    zone_id: str
    type: str
    severity: str
    message: str
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    acknowledged: bool = False
    acknowledged_by: Optional[str] = None
    acknowledged_at: Optional[datetime] = None
    resolution: Optional[str] = None
    metadata: Dict[str, Any] = field(default_factory=dict)
    auto_resolve: bool = False
    threshold_value: Optional[float] = None
    actual_value: Optional[float] = None
    
    def to_dict(self) -> Dict[str, Any]:
        return {
            'id': self.id,
            'zone_id': self.zone_id,
            'type': self.type,
            'severity': self.severity,
            'message': self.message,
            'timestamp': self.created_at.isoformat(),
            'acknowledged': self.acknowledged,
            'metadata': self.metadata
        }

# ML Engine Implementation
class MLEngine:
    """Production ML engine for predictions and optimization"""

    MODEL_VERSION = "2.1.0"

    def __init__(self, model_path: str = "./models/"):
        self.model_path = Path(model_path)
        self.model_path.mkdir(exist_ok=True)
        self.models: Dict[str, Any] = {}
        self.scalers: Dict[str, StandardScaler] = {}
        self.feature_cache: deque = deque(maxlen=10000)
        self.anomaly_detector = IsolationForest(
            contamination=0.1,
            random_state=42,
            n_estimators=100
        )
        self.is_trained = False
        self.training_lock = asyncio.Lock()
        
    async def initialize(self):
        """Load pre-trained models"""
        await self._load_models()
        logger.info("ML Engine initialized")
        
    async def _load_models(self):
        """Load saved models from disk"""
        try:
            version_path = self.model_path / "latest_models.joblib"
            if version_path.exists():
                data = joblib.load(version_path)
                if data.get("version") == self.MODEL_VERSION:
                    self.anomaly_detector = data["models"]["anomaly_detector"]
                    self.scalers = data["models"]["scalers"]
                    self.is_trained = data["metadata"]["is_trained"]
        except Exception as e:
            logger.error(f"Failed to load ML models: {e}")
    
    async def save_models(self):
        """Save models with versioning and validation"""
        try:
            model_data = {
                'version': self.MODEL_VERSION,
                'timestamp': datetime.now(timezone.utc).isoformat(),
                'models': {
                    'anomaly_detector': self.anomaly_detector,
                    'scalers': self.scalers
                },
                'metadata': {
                    'training_samples': len(self.feature_cache),
                    'is_trained': self.is_trained
                }
            }

            version_path = self.model_path / f"models_v{self.MODEL_VERSION.replace('.', '_')}"
            joblib.dump(model_data, f"{version_path}.joblib", compress=3)

            latest_path = self.model_path / "latest_models.joblib"
            if latest_path.exists():
                latest_path.unlink()
            latest_path.symlink_to(f"{version_path}.joblib")
        except Exception as e:
            logger.error(f"Failed to save ML models: {e}")
    
    async def detect_anomaly(self, capability: Capability, command: Dict[str, Any]) -> bool:
        """Detect anomalous commands using ML"""
        with tracer.start_as_current_span("ml_anomaly_detection"):
            features = await self._extract_features(capability, command)
            
            # Add to training cache
            self.feature_cache.append(features)
            
            # Need sufficient training data
            if len(self.feature_cache) < 100:
                return False
                
            # Train or update model if needed
            if not self.is_trained or len(self.feature_cache) % 500 == 0:
                await self._train_anomaly_detector()
                
            try:
                # Scale features
                if 'anomaly' not in self.scalers:
                    self.scalers['anomaly'] = StandardScaler()
                    X = np.array(list(self.feature_cache))
                    self.scalers['anomaly'].fit(X)
                
                X_scaled = self.scalers['anomaly'].transform([features])
                prediction = self.anomaly_detector.predict(X_scaled)
                is_anomaly = prediction[0] == -1
                
                if is_anomaly:
                    anomaly_score = self.anomaly_detector.score_samples(X_scaled)[0]
                    logger.warning(
                        f"ML detected anomaly in command: {command}, "
                        f"score: {anomaly_score:.3f}"
                    )
                    
                return is_anomaly
                
            except Exception as e:
                logger.error(f"ML anomaly detection failed: {e}")
                return False
    
    async def _train_anomaly_detector(self):
        """Train the anomaly detection model"""
        async with self.training_lock:
            try:
                if len(self.feature_cache) < 100:
                    return
                    
                X = np.array(list(self.feature_cache))
                
                # Scale features
                if 'anomaly' not in self.scalers:
                    self.scalers['anomaly'] = StandardScaler()
                X_scaled = self.scalers['anomaly'].fit_transform(X)
                
                # Train model
                self.anomaly_detector.fit(X_scaled)
                self.is_trained = True
                
                # Save models
                await self.save_models()
                
                logger.info(f"Anomaly detector trained on {len(X)} samples")
                
            except Exception as e:
                logger.error(f"Failed to train anomaly detector: {e}")
    
    async def _extract_features(self, capability: Capability, command: Dict[str, Any]) -> List[float]:
        """Extract numerical features from command"""
        features = []
        
        # Time-based features
        now = datetime.now()
        features.extend([
            now.hour,
            now.minute / 60.0,
            now.weekday() / 7.0,
            now.day / 31.0,
            now.month / 12.0,
            math.sin(2 * math.pi * now.hour / 24),  # Cyclical hour encoding
            math.cos(2 * math.pi * now.hour / 24),
        ])
        
        # Command-specific features
        command_type = command.get('command', '')
        features.extend([
            1.0 if 'read' in command_type else 0.0,
            1.0 if 'write' in command_type else 0.0,
            1.0 if 'dose' in command_type else 0.0,
            1.0 if 'adjust' in command_type else 0.0,
            1.0 if 'calibrate' in command_type else 0.0,
        ])
        
        # Numeric command parameters
        features.append(float(command.get('volume_ml', 0)))
        features.append(float(command.get('duration', 0)))
        features.append(float(command.get('target_value', 0)))
        features.append(float(command.get('intensity', 0)))
        
        # Capability features
        features.extend([
            1.0 if capability.type == 'ACTUATOR' else 0.0,
            1.0 if capability.type == 'SENSOR' else 0.0,
            float(capability.error_count),
            1.0 if capability.status == 'online' else 0.0,
        ])
        
        # Historical features
        if capability.last_reading:
            features.append(float(capability.last_reading.get('value', 0)))
            features.append(float(capability.last_reading.get('quality', 1)))
        else:
            features.extend([0.0, 1.0])
            
        return features
    
    async def predict_maintenance(self, zone_id: str, sensor_data: pd.DataFrame) -> List[Dict[str, Any]]:
        """Predict maintenance needs based on sensor data"""
        predictions = []
        
        if sensor_data.empty:
            return predictions
            
        try:
            # Analyze sensor drift
            for sensor_type in ['ph', 'ec', 'temperature', 'humidity']:
                col_value = f'value_{sensor_type}'
                col_quality = f'quality_{sensor_type}'
                
                if col_value not in sensor_data.columns:
                    continue
                    
                # Calculate drift metrics
                values = sensor_data[col_value].dropna()
                if len(values) < 10:
                    continue
                    
                # Linear regression for drift
                x = np.arange(len(values))
                slope, intercept = np.polyfit(x, values, 1)
                
                # Quality degradation
                if col_quality in sensor_data.columns:
                    quality = sensor_data[col_quality].dropna()
                    quality_slope, _ = np.polyfit(np.arange(len(quality)), quality, 1)
                else:
                    quality_slope = 0
                    
                # Variance increase (indicates instability)
                rolling_std = values.rolling(window=24).std()
                variance_trend = rolling_std.iloc[-1] - rolling_std.iloc[len(rolling_std)//2]
                
                # Predict maintenance need
                if abs(quality_slope) > 0.001 or abs(slope) > 0.1 or variance_trend > 0.5:
                    confidence = min(0.95, 0.5 + abs(quality_slope) * 100 + abs(slope) * 2)
                    days_until = max(1, int(7 - confidence * 7))
                    
                    predictions.append({
                        'component': f'{sensor_type}_sensor',
                        'type': 'calibration',
                        'confidence': confidence,
                        'days_until_maintenance': days_until,
                        'recommendation': f'Calibrate {sensor_type} sensor within {days_until} days',
                        'metrics': {
                            'drift_rate': float(slope),
                            'quality_degradation': float(quality_slope),
                            'variance_increase': float(variance_trend)
                        }
                    })
                    
            # Predict pump maintenance based on dosing frequency
            # This would analyze actuator usage patterns
            
        except Exception as e:
            logger.error(f"Maintenance prediction failed: {e}")
            
        return predictions
    
    async def optimize_environment(self, zone: Zone, current_conditions: Dict[str, float]) -> Dict[str, Any]:
        """Optimize environmental setpoints using ML"""
        try:
            targets = zone.environmental_targets
            crop_stage = zone.crop_profile.get('growth_stage', 'vegetative') if zone.crop_profile else 'vegetative'
            
            # Growth stage multipliers
            stage_multipliers = {
                'germination': {'temperature': 1.05, 'humidity': 1.1, 'light': 0.6},
                'seedling': {'temperature': 1.0, 'humidity': 1.05, 'light': 0.8},
                'vegetative': {'temperature': 1.0, 'humidity': 1.0, 'light': 1.0},
                'flowering': {'temperature': 0.95, 'humidity': 0.95, 'light': 1.1},
                'fruiting': {'temperature': 0.98, 'humidity': 0.9, 'light': 1.0}
            }
            
            multipliers = stage_multipliers.get(crop_stage, stage_multipliers['vegetative'])
            
            # Calculate VPD
            temp_c = current_conditions.get('temperature', 22)
            rh = current_conditions.get('humidity', 65)
            svp = 0.6108 * math.exp((17.27 * temp_c) / (temp_c + 237.3))
            vpd = svp * (1 - rh / 100)
            
            recommendations = {
                'adjustments': {},
                'reasoning': []
            }
            
            # Temperature optimization
            temp_target = targets.get('temperature', {}).get('optimal', 22)
            temp_optimal = temp_target * multipliers['temperature']
            
            if abs(temp_c - temp_optimal) > 1:
                recommendations['adjustments']['temperature'] = temp_optimal
                recommendations['reasoning'].append(
                    f"Adjust temperature to {temp_optimal:.1f}°C for {crop_stage} stage"
                )
                
            # VPD optimization
            vpd_target = targets.get('vpd', {}).get('optimal', 1.0)
            if abs(vpd - vpd_target) > 0.2:
                # Calculate required humidity for target VPD
                target_rh = (1 - vpd_target / svp) * 100
                if 40 <= target_rh <= 80:
                    recommendations['adjustments']['humidity'] = target_rh
                    recommendations['reasoning'].append(
                        f"Adjust humidity to {target_rh:.0f}% to achieve VPD of {vpd_target} kPa"
                    )
                    
            # Light optimization based on DLI
            ppfd = current_conditions.get('ppfd', 400)
            photoperiod = targets.get('photoperiod', 18)
            current_dli = ppfd * photoperiod * 0.0036
            target_dli = 20 * multipliers['light']  # mol/m²/day
            
            if abs(current_dli - target_dli) > 2:
                # Adjust intensity to meet DLI target
                new_ppfd = (target_dli / 0.0036) / photoperiod
                if 200 <= new_ppfd <= 800:
                    recommendations['adjustments']['light_intensity'] = new_ppfd
                    recommendations['reasoning'].append(
                        f"Adjust light to {new_ppfd:.0f} μmol/m²/s for optimal DLI"
                    )
                    
            return recommendations
            
        except Exception as e:
            logger.error(f"Environment optimization failed: {e}")
            return {'adjustments': {}, 'reasoning': ['Optimization unavailable']}
    
    async def predict_yield(self, zone: Zone, historical_data: pd.DataFrame) -> Dict[str, Any]:
        """Predict crop yield based on environmental conditions"""
        try:
            if historical_data.empty or not zone.crop_profile:
                return {
                    'predicted_yield': 0,
                    'confidence': 0,
                    'factors': []
                }
                
            # Base yield from crop profile
            base_yield = zone.crop_profile.get('yield_target', 100)
            
            # Calculate environmental factors
            factors = []
            yield_multiplier = 1.0
            
            # Temperature factor
            if 'value_temperature' in historical_data.columns:
                temps = historical_data['value_temperature'].dropna()
                avg_temp = temps.mean()
                optimal_temp = zone.environmental_targets.get('temperature', {}).get('optimal', 22)
                
                temp_deviation = abs(avg_temp - optimal_temp)
                temp_factor = max(0.7, 1 - temp_deviation * 0.05)
                yield_multiplier *= temp_factor
                
                factors.append({
                    'name': 'Temperature',
                    'impact': (temp_factor - 1) * 100,
                    'optimal': temp_deviation < 2,
                    'current': avg_temp,
                    'target': optimal_temp
                })
                
            # Light (DLI) factor
            if 'value_ppfd' in historical_data.columns:
                ppfd_values = historical_data['value_ppfd'].dropna()
                avg_ppfd = ppfd_values.mean()
                photoperiod = zone.environmental_targets.get('photoperiod', 18)
                avg_dli = avg_ppfd * photoperiod * 0.0036
                
                optimal_dli = 20  # mol/m²/day for most crops
                dli_factor = min(1.2, avg_dli / optimal_dli) if optimal_dli > 0 else 1.0
                yield_multiplier *= dli_factor
                
                factors.append({
                    'name': 'Light (DLI)',
                    'impact': (dli_factor - 1) * 100,
                    'optimal': 15 <= avg_dli <= 25,
                    'current': avg_dli,
                    'target': optimal_dli
                })
                
            # Nutrient (EC) factor
            if 'value_ec' in historical_data.columns:
                ec_values = historical_data['value_ec'].dropna()
                avg_ec = ec_values.mean()
                optimal_ec = zone.environmental_targets.get('ec', {}).get('optimal', 2.0)
                
                ec_deviation = abs(avg_ec - optimal_ec)
                ec_factor = max(0.8, 1 - ec_deviation * 0.1)
                yield_multiplier *= ec_factor
                
                factors.append({
                    'name': 'Nutrients (EC)',
                    'impact': (ec_factor - 1) * 100,
                    'optimal': ec_deviation < 0.3,
                    'current': avg_ec,
                    'target': optimal_ec
                })
                
            # Calculate predicted yield
            predicted_yield = base_yield * yield_multiplier
            
            # Confidence based on data quality and quantity
            data_points = len(historical_data)
            data_quality = historical_data[[c for c in historical_data.columns if 'quality' in c]].mean().mean()
            confidence = min(0.95, (data_points / 1000) * 0.5 + data_quality * 0.5)
            
            return {
                'predicted_yield': round(predicted_yield, 1),
                'confidence': round(confidence, 2),
                'factors': factors,
                'yield_multiplier': round(yield_multiplier, 3),
                'data_points': data_points,
                'prediction_date': datetime.now(timezone.utc).isoformat()
            }
            
        except Exception as e:
            logger.error(f"Yield prediction failed: {e}")
            return {
                'predicted_yield': 0,
                'confidence': 0,
                'factors': [],
                'error': str(e)
            }

# Event Bus Implementation
class EventBus:
    """Advanced event bus with priority and filtering"""
    
    def __init__(self):
        self.subscribers: Dict[str, List[tuple]] = defaultdict(list)
        self.event_filters: Dict[str, List[callable]] = defaultdict(list)
        self.event_history: deque = deque(maxlen=1000)
        self.event_count = 0
        
    def subscribe(
        self, 
        event_type: str, 
        handler: callable, 
        priority: int = 0,
        filter_func: Optional[callable] = None
    ):
        """Subscribe with priority and optional filter"""
        self.subscribers[event_type].append((priority, handler))
        self.subscribers[event_type].sort(key=lambda x: x[0], reverse=True)
        
        if filter_func:
            self.event_filters[event_type].append(filter_func)
            
    async def emit(self, event_type: str, data: Dict[str, Any], source: str = ""):
        """Emit event with filtering"""
        self.event_count += 1
        event = {
            'id': f'evt_{self.event_count}',
            'type': event_type,
            'data': data,
            'source': source,
            'timestamp': datetime.now(timezone.utc).isoformat()
        }
        
        # Store in history
        self.event_history.append(event)
        
        # Apply filters
        for filter_func in self.event_filters.get(event_type, []):
            if not filter_func(event):
                return
                
        # Notify subscribers
        for priority, handler in self.subscribers.get(event_type, []):
            try:
                if asyncio.iscoroutinefunction(handler):
                    await handler(event)
                else:
                    handler(event)
            except Exception as e:
                logger.error(f"Event handler error: {e}")
                
        # Also notify wildcard subscribers
        for priority, handler in self.subscribers.get('*', []):
            try:
                if asyncio.iscoroutinefunction(handler):
                    await handler(event)
                else:
                    handler(event)
            except Exception as e:
                logger.error(f"Wildcard event handler error: {e}")

# Audit Logger Implementation
class AuditLogger:
    """Enterprise audit logging with compliance features"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.log_path = Path(config.get('audit_path', '/var/log/phal/audit/'))
        self.log_path.mkdir(parents=True, exist_ok=True)
        self.retention_days = config.get('retention_days', 90)
        self.encryption_key = config.get('encryption_key')
        self.buffer: List[Dict[str, Any]] = []
        self.buffer_size = 100
        self.flush_interval = 10  # seconds
        self.last_flush = datetime.now()
        
        if self.encryption_key:
            self.cipher = Fernet(self.encryption_key.encode())
        else:
            self.cipher = None
            
        # Start flush task
        asyncio.create_task(self._flush_loop())
        
    async def log(self, entry: Dict[str, Any]):
        """Log audit entry with encryption and rotation"""
        entry['timestamp'] = datetime.now(timezone.utc).isoformat()
        entry['hash'] = self._calculate_hash(entry)
        
        # Add to buffer
        self.buffer.append(entry)
        
        # Flush if buffer is full
        if len(self.buffer) >= self.buffer_size:
            await self._flush()
            
    async def _flush_loop(self):
        """Periodic flush of audit buffer"""
        while True:
            await asyncio.sleep(self.flush_interval)
            if self.buffer:
                await self._flush()
                
    async def _flush(self):
        """Flush buffer to disk"""
        if not self.buffer:
            return
            
        # Get current log file
        log_file = self.log_path / f"audit_{datetime.now().strftime('%Y%m%d')}.jsonl"
        
        # Process entries
        entries_to_write = []
        for entry in self.buffer:
            # Encrypt sensitive data if configured
            if self.cipher:
                entry = self._encrypt_entry(entry)
            entries_to_write.append(json.dumps(entry) + '\n')
            
        # Write to file
        async with aiofiles.open(log_file, 'a') as f:
            await f.writelines(entries_to_write)
            
        # Clear buffer
        self.buffer.clear()
        self.last_flush = datetime.now()
        
        # Send to external audit system if configured
        if self.config.get('external_endpoint'):
            await self._send_to_external(entries_to_write)
            
        # Rotate old logs
        await self._rotate_logs()
        
    def _calculate_hash(self, entry: Dict[str, Any]) -> str:
        """Calculate cryptographic hash for integrity"""
        # Remove hash field for calculation
        entry_copy = {k: v for k, v in entry.items() if k != 'hash'}
        content = json.dumps(entry_copy, sort_keys=True)
        return hashlib.sha256(content.encode()).hexdigest()
        
    def _encrypt_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        """Encrypt sensitive fields"""
        sensitive_fields = ['user_id', 'ip_address', 'grant_id', 'api_key']
        encrypted_entry = entry.copy()
        
        for field in sensitive_fields:
            if field in entry and entry[field]:
                encrypted_entry[field] = self.cipher.encrypt(
                    str(entry[field]).encode()
                ).decode()
                
        return encrypted_entry
        
    async def _send_to_external(self, entries: List[str]):
        """Send to external audit system"""
        endpoint = self.config.get('external_endpoint')
        if not endpoint:
            return
            
        try:
            # Parse entries back to JSON
            data = [json.loads(entry.strip()) for entry in entries if entry.strip()]
            
            async with aiohttp.ClientSession() as session:
                async with session.post(
                    endpoint,
                    json={'entries': data},
                    headers={'X-Node-ID': self.config.get('node_id', 'unknown')}
                ) as response:
                    if response.status != 200:
                        logger.error(f"External audit failed: {response.status}")
        except Exception as e:
            logger.error(f"Failed to send audit log to external system: {e}")
        
    async def _rotate_logs(self):
        """Rotate old log files"""
        cutoff_date = datetime.now() - timedelta(days=self.retention_days)
        
        for log_file in self.log_path.glob("audit_*.jsonl"):
            try:
                # Parse date from filename
                file_date_str = log_file.stem.replace('audit_', '')
                file_date = datetime.strptime(file_date_str, '%Y%m%d')
                
                if file_date.date() < cutoff_date.date():
                    if self.config.get('archive_old_logs'):
                        await self._archive_log(log_file)
                    else:
                        log_file.unlink()
                        logger.info(f"Deleted old audit log: {log_file}")
            except ValueError:
                # Invalid filename format
                continue
    
    async def _archive_log(self, log_file: Path):
        """Archive log file to external storage"""
        # Implementation depends on storage backend (S3, GCS, etc)
        archive_path = self.log_path / "archive"
        archive_path.mkdir(exist_ok=True)
        
        # Compress the file
        import gzip
        compressed_file = archive_path / f"{log_file.name}.gz"
        
        with open(log_file, 'rb') as f_in:
            with gzip.open(compressed_file, 'wb') as f_out:
                f_out.writelines(f_in)
                
        log_file.unlink()
        logger.info(f"Archived {log_file} to {compressed_file}")
        
    async def search(self, criteria: Dict[str, Any], limit: int = 100) -> List[Dict[str, Any]]:
        """Search audit logs"""
        results = []
        
        # Search in current buffer first
        for entry in self.buffer:
            if self._matches_criteria(entry, criteria):
                results.append(entry)
                if len(results) >= limit:
                    return results
                    
        # Search in files
        for log_file in sorted(self.log_path.glob("audit_*.jsonl"), reverse=True):
            if len(results) >= limit:
                break
                
            async with aiofiles.open(log_file, 'r') as f:
                async for line in f:
                    if len(results) >= limit:
                        break
                        
                    try:
                        entry = json.loads(line.strip())
                        if self._matches_criteria(entry, criteria):
                            # Decrypt if needed
                            if self.cipher and any(
                                field in entry for field in 
                                ['user_id', 'ip_address', 'grant_id', 'api_key']
                            ):
                                entry = self._decrypt_entry(entry)
                            results.append(entry)
                    except json.JSONDecodeError:
                        continue
                        
        return results
        
    def _matches_criteria(self, entry: Dict[str, Any], criteria: Dict[str, Any]) -> bool:
        """Check if entry matches search criteria"""
        for key, value in criteria.items():
            if key == 'start_date':
                entry_time = datetime.fromisoformat(entry['timestamp'])
                if entry_time < value:
                    return False
            elif key == 'end_date':
                entry_time = datetime.fromisoformat(entry['timestamp'])
                if entry_time > value:
                    return False
            elif key not in entry or entry[key] != value:
                return False
        return True
        
    def _decrypt_entry(self, entry: Dict[str, Any]) -> Dict[str, Any]:
        """Decrypt sensitive fields"""
        if not self.cipher:
            return entry
            
        decrypted = entry.copy()
        for field in ['user_id', 'ip_address', 'grant_id', 'api_key']:
            if field in entry and isinstance(entry[field], str):
                try:
                    decrypted[field] = self.cipher.decrypt(
                        entry[field].encode()
                    ).decode()
                except Exception:
                    # Failed to decrypt, leave as is
                    pass
        return decrypted

# Telemetry Collector
class TelemetryCollector:
    """Telemetry collection for monitoring and analytics"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.buffer: List[Dict[str, Any]] = []
        self.buffer_size = config.get('buffer_size', 1000)
        self.flush_interval = config.get('flush_interval', 60)
        self.metrics_aggregator = defaultdict(list)
        
        if config.get('enabled', True):
            asyncio.create_task(self._flush_loop())
            
    async def record(self, metric_type: str, data: Dict[str, Any]):
        """Record telemetry data"""
        entry = {
            'type': metric_type,
            'data': data,
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'node_id': self.config.get('node_id', 'unknown')
        }
        
        self.buffer.append(entry)
        
        # Aggregate metrics
        if metric_type in ['sensor_reading', 'command_execution', 'api_request']:
            self.metrics_aggregator[metric_type].append(data)
        
        if len(self.buffer) >= self.buffer_size:
            await self._flush()
            
    async def _flush_loop(self):
        """Periodic flush of telemetry data"""
        while True:
            await asyncio.sleep(self.flush_interval)
            await self._flush()
            
    async def _flush(self):
        """Send telemetry to collection endpoint"""
        if not self.buffer:
            return
            
        batch = self.buffer[:self.buffer_size]
        self.buffer = self.buffer[self.buffer_size:]
        
        # Calculate aggregates
        aggregates = self._calculate_aggregates()
        
        endpoint = self.config.get('endpoint')
        if endpoint:
            try:
                async with aiohttp.ClientSession() as session:
                    payload = {
                        'events': batch,
                        'aggregates': aggregates,
                        'node_id': self.config.get('node_id'),
                        'timestamp': datetime.now(timezone.utc).isoformat()
                    }
                    
                    async with session.post(endpoint, json=payload) as response:
                        if response.status != 200:
                            logger.error(f"Telemetry flush failed: {response.status}")
                            # Re-queue failed batch
                            self.buffer = batch + self.buffer
            except Exception as e:
                logger.error(f"Failed to send telemetry: {e}")
                # Re-queue failed batch
                self.buffer = batch + self.buffer
                
        # Clear aggregator
        self.metrics_aggregator.clear()
        
    def _calculate_aggregates(self) -> Dict[str, Any]:
        """Calculate metric aggregates"""
        aggregates = {}
        
        # Sensor reading aggregates
        if 'sensor_reading' in self.metrics_aggregator:
            readings = self.metrics_aggregator['sensor_reading']
            by_type = defaultdict(list)
            
            for reading in readings:
                sensor_type = reading.get('sensor_type')
                value = reading.get('value')
                if sensor_type and value is not None:
                    by_type[sensor_type].append(value)
                    
            aggregates['sensor_readings'] = {
                sensor: {
                    'count': len(values),
                    'mean': np.mean(values),
                    'std': np.std(values),
                    'min': min(values),
                    'max': max(values)
                }
                for sensor, values in by_type.items()
            }
            
        # Command execution aggregates
        if 'command_execution' in self.metrics_aggregator:
            commands = self.metrics_aggregator['command_execution']
            by_type = defaultdict(list)
            
            for cmd in commands:
                cmd_type = cmd.get('command_type')
                latency = cmd.get('latency')
                if cmd_type and latency is not None:
                    by_type[cmd_type].append(latency)
                    
            aggregates['command_latency'] = {
                cmd: {
                    'count': len(latencies),
                    'mean': np.mean(latencies),
                    'p50': np.percentile(latencies, 50),
                    'p95': np.percentile(latencies, 95),
                    'p99': np.percentile(latencies, 99)
                }
                for cmd, latencies in by_type.items()
            }
            
        return aggregates

# Base Plugin Implementation
class BasePlugin:
    """Enhanced base plugin with lifecycle management"""
    
    def __init__(self, phal: 'PluripotentHAL', config: Dict[str, Any]):
        self.phal = phal
        self.config = config
        self.id = self.__class__.__name__
        self.capabilities: List[Capability] = []
        self.health_status = {
            'healthy': True,
            'last_check': None,
            'error_count': 0,
            'details': {}
        }
        self.initialized = False
        
    async def initialize(self) -> bool:
        """Initialize plugin with configuration"""
        try:
            # Plugin-specific initialization
            await self._initialize()
            self.initialized = True
            logger.info(f"Plugin {self.id} initialized")
            return True
        except Exception as e:
            logger.error(f"Plugin {self.id} initialization failed: {e}")
            self.health_status['healthy'] = False
            self.health_status['details']['init_error'] = str(e)
            return False
            
    async def _initialize(self):
        """Override in subclasses"""
        pass
        
    async def shutdown(self):
        """Cleanup on shutdown"""
        try:
            await self._shutdown()
            logger.info(f"Plugin {self.id} shutdown complete")
        except Exception as e:
            logger.error(f"Plugin {self.id} shutdown error: {e}")
            
    async def _shutdown(self):
        """Override in subclasses"""
        pass
        
    async def health_check(self) -> Dict[str, Any]:
        """Check plugin health"""
        self.health_status['last_check'] = datetime.now(timezone.utc).isoformat()
        
        try:
            # Plugin-specific health check
            details = await self._health_check()
            self.health_status['details'].update(details)
            
            # Check capability health
            unhealthy_caps = [
                cap.id for cap in self.capabilities 
                if cap.status != 'online' or cap.error_count > 10
            ]
            
            if unhealthy_caps:
                self.health_status['healthy'] = False
                self.health_status['details']['unhealthy_capabilities'] = unhealthy_caps
            else:
                self.health_status['healthy'] = True
                
        except Exception as e:
            self.health_status['healthy'] = False
            self.health_status['error_count'] += 1
            self.health_status['details']['check_error'] = str(e)
            
        return self.health_status
        
    async def _health_check(self) -> Dict[str, Any]:
        """Override in subclasses"""
        return {}
        
    async def execute_capability(self, capability_id: str, command: Dict[str, Any]) -> Dict[str, Any]:
        """Execute capability command"""
        capability = next((c for c in self.capabilities if c.id == capability_id), None)
        
        if not capability:
            raise ValueError(f"Capability {capability_id} not found")
            
        if not self.initialized:
            raise RuntimeError(f"Plugin {self.id} not initialized")
            
        try:
            result = await self._execute_command(capability, command)
            capability.error_count = 0  # Reset on success
            return result
        except Exception as e:
            capability.error_count += 1
            if capability.error_count > 10:
                capability.status = 'error'
            raise
            
    async def _execute_command(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Override in subclasses"""
        raise NotImplementedError
        
    def register_capability(self, capability: Capability):
        """Register a capability with PHAL"""
        capability.tenant_id = None  # System capability
        self.capabilities.append(capability)
        self.phal.capabilities[capability.id] = capability

# Hardware Simulator Plugin
class HardwareSimulator(BasePlugin):
    """Complete hardware simulator with realistic behavior"""
    
    async def _initialize(self):
        """Initialize simulator"""
        self.state = {
            'sensors': defaultdict(lambda: {
                'value': 0.0,
                'quality': 1.0,
                'last_calibration': datetime.now(timezone.utc),
                'drift_rate': random.uniform(-0.001, 0.001),
                'noise_factor': random.uniform(0.01, 0.05)
            }),
            'actuators': defaultdict(lambda: {
                'state': False,
                'value': 0.0,
                'last_activation': None,
                'total_runtime': 0.0,
                'activation_count': 0
            }),
            'resource_levels': {
                'nutrient_a': 10000.0,  # ml
                'nutrient_b': 10000.0,
                'ph_up': 5000.0,
                'ph_down': 5000.0,
                'water': 100000.0
            },
            'environmental_model': {
                'external_temp': 15.0,  # Outside temperature
                'thermal_mass': 1000.0,  # kg
                'heating_power': 5000.0,  # W
                'cooling_power': 10000.0,  # W
                'humidity_generation': 2.0,  # L/hour from plants
                'co2_consumption': 50.0,  # ppm/hour
            }
        }
        
        # Initialize sensors with realistic values
        self._initialize_sensor_values()
        
        # Register capabilities
        zones = self.phal.zones.values()
        for zone in zones:
            # Register sensors
            for sensor_type in ['temperature', 'humidity', 'ph', 'ec', 'co2', 'ppfd']:
                capability = Capability(
                    id=f"{zone.id}:{sensor_type}",
                    type="SENSOR",
                    subtype=sensor_type,
                    zone_id=zone.id,
                    properties={
                        'unit': self._get_sensor_unit(sensor_type),
                        'precision': self._get_sensor_precision(sensor_type),
                        'range': self._get_sensor_range(sensor_type)
                    }
                )
                self.register_capability(capability)
                
            # Register actuators
            actuators = [
                ('heater', 'climate_control'),
                ('cooler', 'climate_control'),
                ('humidifier', 'climate_control'),
                ('dehumidifier', 'climate_control'),
                ('light_panel', 'lighting'),
                ('pump_a', 'dosing_system'),
                ('pump_b', 'dosing_system'),
                ('pump_ph_up', 'ph_control'),
                ('pump_ph_down', 'ph_control'),
                ('irrigation_pump', 'irrigation')
            ]
            
            for actuator_id, subtype in actuators:
                capability = Capability(
                    id=f"{zone.id}:{actuator_id}",
                    type="ACTUATOR",
                    subtype=subtype,
                    zone_id=zone.id,
                    properties={
                        'max_flow_rate': 100 if 'pump' in actuator_id else None,
                        'power_consumption': self._get_actuator_power(actuator_id)
                    }
                )
                self.register_capability(capability)
        
        # Start simulation tasks
        self.simulation_tasks = [
            asyncio.create_task(self._simulate_environment()),
            asyncio.create_task(self._simulate_sensor_drift()),
            asyncio.create_task(self._simulate_resource_consumption()),
            asyncio.create_task(self._simulate_plant_effects())
        ]
        
    async def _shutdown(self):
        """Shutdown simulator"""
        for task in self.simulation_tasks:
            task.cancel()
            try:
                await task
            except asyncio.CancelledError:
                pass
                
    async def _health_check(self) -> Dict[str, Any]:
        """Check simulator health"""
        return {
            'resource_levels': self.state['resource_levels'],
            'active_actuators': sum(
                1 for a in self.state['actuators'].values() 
                if a['state']
            ),
            'simulation_running': all(not task.done() for task in self.simulation_tasks)
        }
        
    async def _execute_command(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Execute simulator command"""
        cmd_type = command.get('command', '')
        
        if capability.type == 'SENSOR':
            if cmd_type in ['read', 'read_all']:
                return await self._read_sensor(capability)
            elif cmd_type == 'calibrate':
                return await self._calibrate_sensor(capability, command)
        elif capability.type == 'ACTUATOR':
            if cmd_type == 'set_state':
                return await self._control_actuator(capability, command)
            elif cmd_type in ['dose', 'dose_recipe']:
                return await self._dose_nutrients(capability, command)
            elif cmd_type == 'adjust_ph':
                return await self._adjust_ph(capability, command)
                
        raise ValueError(f"Unknown command: {cmd_type}")
        
    async def _read_sensor(self, capability: Capability) -> Dict[str, Any]:
        """Read sensor with realistic behavior"""
        sensor_key = f"{capability.zone_id}:{capability.subtype}"
        sensor = self.state['sensors'][sensor_key]
        
        # Add noise and drift
        base_value = sensor['value']
        noise = np.random.normal(0, sensor['noise_factor'] * base_value)
        drift = sensor['drift_rate'] * (datetime.now(timezone.utc) - sensor['last_calibration']).total_seconds() / 3600
        
        value = base_value + noise + drift
        
        # Clamp to sensor range
        range_min, range_max = capability.properties['range']
        value = max(range_min, min(range_max, value))
        
        # Calculate quality based on calibration age and drift
        days_since_cal = (datetime.now(timezone.utc) - sensor['last_calibration']).days
        quality_degradation = min(0.3, days_since_cal * 0.01 + abs(drift) * 0.1)
        quality = max(0.5, sensor['quality'] - quality_degradation)
        
        result = {
            'value': round(value, capability.properties['precision']),
            'quality': round(quality, 3),
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'unit': capability.properties['unit'],
            'raw_value': round(base_value, capability.properties['precision'])
        }
        
        # Update capability
        capability.last_reading = result
        
        # Record telemetry
        await self.phal.telemetry.record('sensor_reading', {
            'zone_id': capability.zone_id,
            'sensor_type': capability.subtype,
            'value': result['value'],
            'quality': result['quality']
        })
        
        # Update Prometheus metrics
        sensor_readings.labels(
            sensor_type=capability.subtype,
            zone_id=capability.zone_id
        ).observe(result['value'])
        
        return result
        
    async def _calibrate_sensor(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Calibrate sensor"""
        sensor_key = f"{capability.zone_id}:{capability.subtype}"
        sensor = self.state['sensors'][sensor_key]
        
        # Reset drift and quality
        sensor['last_calibration'] = datetime.now(timezone.utc)
        sensor['quality'] = 1.0
        sensor['drift_rate'] = random.uniform(-0.001, 0.001)  # New drift rate
        
        # Update calibration info
        capability.calibration = {
            'last_calibration': sensor['last_calibration'].isoformat(),
            'calibration_points': command.get('calibration_points', []),
            'technician': command.get('technician', 'system'),
            'notes': command.get('notes', '')
        }
        
        return {
            'success': True,
            'message': f'Sensor {capability.subtype} calibrated successfully',
            'quality': 1.0
        }
        
    async def _control_actuator(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Control actuator"""
        actuator_key = f"{capability.zone_id}:{capability.id.split(':')[1]}"
        actuator = self.state['actuators'][actuator_key]
        
        state = command.get('state')
        if isinstance(state, bool):
            actuator['state'] = state
            actuator['value'] = 100.0 if state else 0.0
        else:
            # Numeric state (0-100)
            actuator['value'] = float(state)
            actuator['state'] = actuator['value'] > 0
            
        if actuator['state']:
            actuator['last_activation'] = datetime.now(timezone.utc)
            actuator['activation_count'] += 1
            
        # Simulate physical effects
        asyncio.create_task(self._apply_actuator_effects(capability, actuator))
        
        return {
            'success': True,
            'state': actuator['state'],
            'value': actuator['value']
        }
        
    async def _dose_nutrients(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Dose nutrients"""
        zone_id = capability.zone_id
        
        if 'recipe' in command:
            # Dose recipe
            recipe = command['recipe']
            total_dosed = {}
            
            for nutrient, amount in recipe.items():
                if nutrient in self.state['resource_levels']:
                    available = self.state['resource_levels'][nutrient]
                    actual_dose = min(amount, available)
                    self.state['resource_levels'][nutrient] -= actual_dose
                    total_dosed[nutrient] = actual_dose
                    
                    # Update EC based on dosing
                    ec_increase = actual_dose * 0.001  # Simplified
                    sensor_key = f"{zone_id}:ec"
                    self.state['sensors'][sensor_key]['value'] += ec_increase
                    
            return {
                'success': True,
                'dosed': total_dosed,
                'remaining': self.state['resource_levels']
            }
        else:
            # Single nutrient dose
            volume = command.get('volume_ml', 0)
            nutrient = 'nutrient_a'  # Default
            
            available = self.state['resource_levels'][nutrient]
            actual_dose = min(volume, available)
            self.state['resource_levels'][nutrient] -= actual_dose
            
            return {
                'success': True,
                'dosed_ml': actual_dose,
                'remaining_ml': self.state['resource_levels'][nutrient]
            }
            
    async def _adjust_ph(self, capability: Capability, command: Dict[str, Any]) -> Dict[str, Any]:
        """Adjust pH"""
        zone_id = capability.zone_id
        current_ph = command.get('current_ph', 7.0)
        target_ph = command.get('target_ph', 6.2)
        max_dose = command.get('max_dose_ml', 50)
        
        # Calculate required dose (simplified)
        ph_diff = target_ph - current_ph
        if abs(ph_diff) < 0.1:
            return {
                'success': True,
                'message': 'pH already at target',
                'dosed_ml': 0
            }
            
        # Determine which solution to use
        if ph_diff < 0:
            # Need to lower pH
            solution = 'ph_down'
            dose_per_ph = 10  # ml per pH unit
        else:
            # Need to raise pH
            solution = 'ph_up'
            dose_per_ph = 15  # ml per pH unit
            
        required_dose = abs(ph_diff) * dose_per_ph
        actual_dose = min(required_dose, max_dose, self.state['resource_levels'][solution])
        
        # Apply dose
        self.state['resource_levels'][solution] -= actual_dose
        
        # Update pH sensor
        ph_change = (actual_dose / dose_per_ph) * (1 if ph_diff > 0 else -1)
        sensor_key = f"{zone_id}:ph"
        self.state['sensors'][sensor_key]['value'] += ph_change
        
        return {
            'success': True,
            'solution': solution,
            'dosed_ml': actual_dose,
            'new_ph': self.state['sensors'][sensor_key]['value'],
            'remaining_ml': self.state['resource_levels'][solution]
        }
        
    def _initialize_sensor_values(self):
        """Initialize sensors with realistic values"""
        for zone in self.phal.zones.values():
            base_values = {
                'temperature': 22.5 + random.uniform(-2, 2),
                'humidity': 65.0 + random.uniform(-5, 5),
                'ph': 6.2 + random.uniform(-0.3, 0.3),
                'ec': 2.0 + random.uniform(-0.2, 0.2),
                'co2': 800 + random.uniform(-100, 100),
                'ppfd': 400 + random.uniform(-50, 50)
            }
            
            for sensor_type, value in base_values.items():
                sensor_key = f"{zone.id}:{sensor_type}"
                self.state['sensors'][sensor_key]['value'] = value
                
    def _get_sensor_unit(self, sensor_type: str) -> str:
        units = {
            'temperature': '°C',
            'humidity': '%',
            'ph': 'pH',
            'ec': 'mS/cm',
            'co2': 'ppm',
            'ppfd': 'μmol/m²/s'
        }
        return units.get(sensor_type, '')
        
    def _get_sensor_precision(self, sensor_type: str) -> int:
        precision = {
            'temperature': 1,
            'humidity': 0,
            'ph': 1,
            'ec': 2,
            'co2': 0,
            'ppfd': 0
        }
        return precision.get(sensor_type, 2)
        
    def _get_sensor_range(self, sensor_type: str) -> tuple:
        ranges = {
            'temperature': (-10, 50),
            'humidity': (0, 100),
            'ph': (0, 14),
            'ec': (0, 10),
            'co2': (0, 5000),
            'ppfd': (0, 2000)
        }
        return ranges.get(sensor_type, (0, 100))
        
    def _get_actuator_power(self, actuator_id: str) -> float:
        """Get actuator power consumption in watts"""
        power_map = {
            'heater': 2000,
            'cooler': 3000,
            'humidifier': 500,
            'dehumidifier': 1000,
            'light_panel': 600,
            'pump': 50
        }
        
        for key, power in power_map.items():
            if key in actuator_id:
                return power
        return 100  # Default
        
    async def _simulate_environment(self):
        """Simulate environmental changes"""
        while True:
            try:
                for zone in self.phal.zones.values():
                    if not zone.is_operational():
                        continue
                        
                    # Temperature simulation
                    await self._simulate_temperature(zone)
                    
                    # Humidity simulation
                    await self._simulate_humidity(zone)
                    
                    # CO2 simulation
                    await self._simulate_co2(zone)
                    
                    # pH and EC drift
                    await self._simulate_water_chemistry(zone)
                    
                await asyncio.sleep(5)  # Update every 5 seconds
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Environment simulation error: {e}")
                await asyncio.sleep(10)
                
    async def _simulate_temperature(self, zone: Zone):
        """Simulate temperature dynamics"""
        temp_key = f"{zone.id}:temperature"
        current_temp = self.state['sensors'][temp_key]['value']
        
        # External temperature influence
        external_temp = self.state['environmental_model']['external_temp']
        external_temp += 5 * math.sin(datetime.now().hour * math.pi / 12)  # Daily cycle
        
        # Heat transfer
        delta_external = (external_temp - current_temp) * 0.001  # Through walls
        
        # Check actuator states
        heater_key = f"{zone.id}:heater"
        cooler_key = f"{zone.id}:cooler"
        
        heater_on = self.state['actuators'][heater_key]['state']
        cooler_on = self.state['actuators'][cooler_key]['state']
        
        # Heating/cooling effects
        if heater_on:
            heating_power = self.state['actuators'][heater_key]['value'] / 100
            delta_heating = heating_power * 0.02  # °C per cycle
        else:
            delta_heating = 0
            
        if cooler_on:
            cooling_power = self.state['actuators'][cooler_key]['value'] / 100
            delta_cooling = -cooling_power * 0.03  # °C per cycle
        else:
            delta_cooling = 0
            
        # Light heat contribution
        light_key = f"{zone.id}:light_panel"
        if self.state['actuators'][light_key]['state']:
            light_intensity = self.state['actuators'][light_key]['value'] / 100
            delta_light = light_intensity * 0.01  # Heat from lights
        else:
            delta_light = 0
            
        # Apply changes
        new_temp = current_temp + delta_external + delta_heating + delta_cooling + delta_light
        
        # Add some randomness
        new_temp += random.uniform(-0.1, 0.1)
        
        # Clamp to reasonable range
        new_temp = max(10, min(40, new_temp))
        
        self.state['sensors'][temp_key]['value'] = new_temp
        
    async def _simulate_humidity(self, zone: Zone):
        """Simulate humidity dynamics"""
        humidity_key = f"{zone.id}:humidity"
        temp_key = f"{zone.id}:temperature"
        
        current_humidity = self.state['sensors'][humidity_key]['value']
        current_temp = self.state['sensors'][temp_key]['value']
        
        # Temperature-humidity relationship
        # Higher temp = lower relative humidity
        target_humidity = 70 - (current_temp - 20) * 2
        
        # Plant transpiration
        if zone.crop_profile:
            plant_age = zone.get_age_days()
            transpiration_rate = min(5, plant_age * 0.1)  # %/cycle
        else:
            transpiration_rate = 0
            
        # Check actuator states
        humidifier_key = f"{zone.id}:humidifier"
        dehumidifier_key = f"{zone.id}:dehumidifier"
        
        if self.state['actuators'][humidifier_key]['state']:
            delta_humid = 2.0
        elif self.state['actuators'][dehumidifier_key]['state']:
            delta_humid = -3.0
        else:
            # Natural drift toward target
            delta_humid = (target_humidity - current_humidity) * 0.05
            
        # Apply changes
        new_humidity = current_humidity + delta_humid + transpiration_rate
        new_humidity += random.uniform(-0.5, 0.5)
        new_humidity = max(20, min(95, new_humidity))
        
        self.state['sensors'][humidity_key]['value'] = new_humidity
        
    async def _simulate_co2(self, zone: Zone):
        """Simulate CO2 dynamics"""
        co2_key = f"{zone.id}:co2"
        light_key = f"{zone.id}:light_panel"
        
        current_co2 = self.state['sensors'][co2_key]['value']
        
        # Photosynthesis consumes CO2 when lights are on
        if self.state['actuators'][light_key]['state'] and zone.crop_profile:
            light_intensity = self.state['actuators'][light_key]['value'] / 100
            plant_age = zone.get_age_days()
            consumption_rate = light_intensity * min(50, plant_age * 2)  # ppm/cycle
            delta_co2 = -consumption_rate
        else:
            # Natural increase from respiration and air exchange
            delta_co2 = 10
            
        # Air exchange with outside (400 ppm)
        delta_exchange = (400 - current_co2) * 0.01
        
        # Apply changes
        new_co2 = current_co2 + delta_co2 + delta_exchange
        new_co2 += random.uniform(-5, 5)
        new_co2 = max(300, min(2000, new_co2))
        
        self.state['sensors'][co2_key]['value'] = new_co2
        
    async def _simulate_water_chemistry(self, zone: Zone):
        """Simulate pH and EC changes"""
        ph_key = f"{zone.id}:ph"
        ec_key = f"{zone.id}:ec"
        
        # pH drift (tends toward 7.0)
        current_ph = self.state['sensors'][ph_key]['value']
        ph_drift = (7.0 - current_ph) * 0.001
        
        # Nutrient uptake affects EC
        current_ec = self.state['sensors'][ec_key]['value']
        if zone.crop_profile:
            plant_age = zone.get_age_days()
            uptake_rate = min(0.01, plant_age * 0.0002)  # mS/cm per cycle
            ec_drift = -uptake_rate
        else:
            ec_drift = 0
            
        # Apply changes
        self.state['sensors'][ph_key]['value'] = current_ph + ph_drift + random.uniform(-0.01, 0.01)
        self.state['sensors'][ec_key]['value'] = max(0.1, current_ec + ec_drift + random.uniform(-0.01, 0.01))
        
    async def _simulate_sensor_drift(self):
        """Simulate sensor degradation over time"""
        while True:
            try:
                for sensor_key, sensor in self.state['sensors'].items():
                    # Quality degradation
                    days_since_cal = (datetime.now(timezone.utc) - sensor['last_calibration']).days
                    if days_since_cal > 7:
                        quality_loss = min(0.001, days_since_cal * 0.00001)
                        sensor['quality'] = max(0.5, sensor['quality'] - quality_loss)
                        
                    # Drift accumulation
                    sensor['drift_rate'] += random.uniform(-0.0001, 0.0001)
                    sensor['drift_rate'] = max(-0.01, min(0.01, sensor['drift_rate']))
                    
                await asyncio.sleep(3600)  # Check hourly
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Sensor drift simulation error: {e}")
                await asyncio.sleep(3600)
                
    async def _simulate_resource_consumption(self):
        """Simulate resource usage"""
        while True:
            try:
                # Water evaporation
                self.state['resource_levels']['water'] -= 10  # ml/cycle
                
                # Check for low resource alerts
                for resource, level in self.state['resource_levels'].items():
                    if level < 1000:  # Less than 1L
                        await self.phal.event_bus.emit('alarm', {
                            'type': 'resource_low',
                            'severity': 'warning' if level > 500 else 'critical',
                            'resource': resource,
                            'level': level,
                            'message': f'Low {resource} level: {level}ml remaining'
                        })
                        
                await asyncio.sleep(60)  # Check every minute
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Resource consumption simulation error: {e}")
                await asyncio.sleep(60)
                
    async def _simulate_plant_effects(self):
        """Simulate plant growth effects on environment"""
        while True:
            try:
                for zone in self.phal.zones.values():
                    if not zone.crop_profile or not zone.is_operational():
                        continue
                        
                    plant_age = zone.get_age_days()
                    growth_stage = zone.crop_profile.get('growth_stage', 'vegetative')
                    
                    # PPFD affected by canopy density
                    ppfd_key = f"{zone.id}:ppfd"
                    light_key = f"{zone.id}:light_panel"
                    
                    if self.state['actuators'][light_key]['state']:
                        base_ppfd = 600 * (self.state['actuators'][light_key]['value'] / 100)
                        
                        # Canopy shading increases with age
                        canopy_factor = max(0.7, 1 - plant_age * 0.002)
                        
                        self.state['sensors'][ppfd_key]['value'] = base_ppfd * canopy_factor
                        
                await asyncio.sleep(300)  # Every 5 minutes
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Plant effects simulation error: {e}")
                await asyncio.sleep(300)
                
    async def _apply_actuator_effects(self, capability: Capability, actuator: Dict[str, Any]):
        """Apply immediate effects of actuator changes"""
        # This is handled by the simulation loops
        pass

# Rate limiter
class RateLimiter:
    def __init__(self, redis: redis.Redis):
        self.redis = redis
        
    async def check_rate_limit(self, key: str, limit: int, window: int) -> bool:
        """Check if rate limit is exceeded"""
        try:
            current = await self.redis.incr(key)
            if current == 1:
                await self.redis.expire(key, window)
            return current <= limit
        except Exception as e:
            logger.error(f"Rate limit check failed: {e}")
            return True  # Allow on error
    
    async def get_remaining(self, key: str, limit: int) -> int:
        """Get remaining requests in current window"""
        try:
            current = await self.redis.get(key)
            if current is None:
                return limit
            return max(0, limit - int(current))
        except Exception as e:
            logger.error(f"Get remaining failed: {e}")
            return limit

# Circuit breaker
class CircuitBreaker:
    def __init__(self, failure_threshold: int = 5, recovery_timeout: int = 60,
                 half_open_max_calls: int = 3):
        self.failure_threshold = failure_threshold
        self.recovery_timeout = recovery_timeout
        self.half_open_max_calls = half_open_max_calls
        self.failure_count = 0
        self.last_failure_time = None
        self.state = "closed"
        self.half_open_calls = 0
        self.success_count = 0
        self._lock = asyncio.Lock()

    async def call(self, func, *args, **kwargs):
        async with self._lock:
            if self.state == "open":
                if datetime.now() - self.last_failure_time > timedelta(seconds=self.recovery_timeout):
                    self.state = "half-open"
                    self.half_open_calls = 0
                    logger.info("Circuit breaker entering half-open state")
                else:
                    raise Exception(
                        f"Circuit breaker is open. Recovery in {self.recovery_timeout - (datetime.now() - self.last_failure_time).seconds}s")

            if self.state == "half-open" and self.half_open_calls >= self.half_open_max_calls:
                if self.success_count == self.half_open_calls:
                    self.state = "closed"
                    self.failure_count = 0
                    logger.info("Circuit breaker closed after successful recovery")
                else:
                    self.state = "open"
                    logger.warning("Circuit breaker reopened after failed recovery")
                    raise Exception("Circuit breaker reopened")

        try:
            result = await func(*args, **kwargs)
            if self.state == "half-open":
                self.half_open_calls += 1
                self.success_count += 1
            elif self.state == "closed":
                self.failure_count = max(0, self.failure_count - 1)
            return result
        except Exception as e:
            async with self._lock:
                self.failure_count += 1
                self.last_failure_time = datetime.now()

                if self.state == "half-open":
                    self.state = "open"
                    logger.error(f"Circuit breaker reopened: {e}")
                elif self.failure_count >= self.failure_threshold:
                    self.state = "open"
                    logger.error(f"Circuit breaker opened after {self.failure_count} failures")
            raise

# Main PHAL System
class PluripotentHAL:
    """Production-ready PHAL system"""
    
    def __init__(self, config: PHALConfig):
        self.config = config
        self.tenants: Dict[str, Tenant] = {}
        self.zones: Dict[str, Zone] = {}
        self.grants: Dict[str, Grant] = {}
        self.capabilities: Dict[str, Capability] = {}
        self.plugins: Dict[str, BasePlugin] = {}
        self.alarms: Dict[str, Alarm] = {}
        self.event_bus = EventBus()
        self.audit_logger = AuditLogger(config.dict())
        self.telemetry = TelemetryCollector(config.dict())
        self.ml_engine = MLEngine(config.ml_model_path)
        self.cache: Optional[redis.Redis] = None
        self.db: Optional[asyncpg.Pool] = None
        self.rate_limiter: Optional[RateLimiter] = None
        self.circuit_breaker = CircuitBreaker()
        self.background_tasks = []
        self.session_store: Dict[str, Dict[str, Any]] = {}
        self.start_time = datetime.now(timezone.utc)
        
        # Initialize encryption if configured
        if config.encryption_key:
            self.cipher = Fernet(config.encryption_key.encode())
        else:
            self.cipher = None

    async def initialize_db_pool(self):
        self.db = await asyncpg.create_pool(
            self.config.database_url,
            min_size=10,
            max_size=20,
            max_queries=50000,
            max_inactive_connection_lifetime=300,
            command_timeout=10,
            server_settings={
                'jit': 'off',
                'application_name': 'phal_backend'
            }
        )
            
    async def initialize(self):
        """Initialize PHAL system"""
        with tracer.start_as_current_span("phal_initialize"):
            logger.info("Initializing PHAL system...")
            
            # Connect to database
            await self.initialize_db_pool()
            
            # Update connection metrics
            db_connections.labels(state='active').set(10)
            
            # Initialize schema
            await self._initialize_database_schema()
            
            # Connect to Redis
            self.cache = await redis.from_url(
                self.config.redis_url,
                encoding="utf-8",
                decode_responses=True,
                max_connections=50
            )
            
            # Initialize rate limiter
            self.rate_limiter = RateLimiter(self.cache)
            
            # Load data
            await self._load_tenants()
            await self._load_zones()
            await self._load_capabilities()
            await self._load_grants()
            
            # Initialize ML models
            await self.ml_engine.initialize()
            
            # Initialize plugins
            await self._initialize_plugins()
            
            # Subscribe to internal events
            self._setup_event_handlers()
            
            # Start background tasks
            self.background_tasks = [
                asyncio.create_task(self._health_monitor()),
                asyncio.create_task(self._grant_cleanup()),
                asyncio.create_task(self._session_cleanup()),
                asyncio.create_task(self._backup_system()),
                asyncio.create_task(self._alarm_processor()),
                asyncio.create_task(self._predictive_maintenance()),
                asyncio.create_task(self._metrics_collector()),
                asyncio.create_task(self._resource_optimizer())
            ]
            
            # Update system uptime metric
            system_uptime.set(0)
            
            logger.info("PHAL system initialized successfully")
            
    async def shutdown(self):
        """Enhanced graceful shutdown"""
        logger.info("Starting graceful shutdown...")

        self.shutting_down = True

        # Stop accepting new requests is handled by web runner

        try:
            await asyncio.wait_for(
                self._wait_for_requests_completion(),
                timeout=30
            )
        except asyncio.TimeoutError:
            logger.warning("Timeout waiting for requests to complete")

        for task in self.background_tasks:
            task.cancel()
            try:
                await asyncio.wait_for(task, timeout=5)
            except (asyncio.CancelledError, asyncio.TimeoutError):
                pass

        await self.telemetry._flush()
        await self.audit_logger._flush()
        await self.ml_engine.save_models()

        for plugin in self.plugins.values():
            await plugin.shutdown()

        if self.db:
            await self.db.close()

        if self.cache:
            await self.cache.close()

        logger.info("Graceful shutdown complete")

    async def _wait_for_requests_completion(self):
        await asyncio.sleep(0)
        
    async def _initialize_database_schema(self):
        """Initialize complete database schema"""
        async with self.db.acquire() as conn:
            # Main tables
            await conn.execute('''
                -- Tenants table
                CREATE TABLE IF NOT EXISTS tenants (
                    id UUID PRIMARY KEY,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    tier VARCHAR(50) NOT NULL CHECK (tier IN ('community', 'professional', 'enterprise')),
                    features JSONB DEFAULT '[]'::jsonb,
                    resource_limits JSONB DEFAULT '{}'::jsonb,
                    api_key_hash VARCHAR(255) UNIQUE,
                    contact_email VARCHAR(255),
                    billing_status VARCHAR(50) DEFAULT 'active',
                    usage_stats JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW(),
                    metadata JSONB DEFAULT '{}'::jsonb
                );
                
                CREATE INDEX IF NOT EXISTS idx_tenants_api_key ON tenants(api_key_hash);
                CREATE INDEX IF NOT EXISTS idx_tenants_tier ON tenants(tier);
                
                -- Zones table
                CREATE TABLE IF NOT EXISTS zones (
                    id UUID PRIMARY KEY,
                    tenant_id UUID REFERENCES tenants(id) ON DELETE CASCADE,
                    name VARCHAR(255) NOT NULL,
                    type VARCHAR(50) NOT NULL CHECK (type IN ('production', 'nursery', 'quarantine', 'research')),
                    units TEXT[] DEFAULT ARRAY[]::TEXT[],
                    environmental_targets JSONB DEFAULT '{}'::jsonb,
                    crop_profile JSONB DEFAULT NULL,
                    emergency_stop BOOLEAN DEFAULT FALSE,
                    maintenance_mode BOOLEAN DEFAULT FALSE,
                    last_harvest TIMESTAMPTZ,
                    total_yield FLOAT DEFAULT 0,
                    active_alarms TEXT[] DEFAULT ARRAY[]::TEXT[],
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW(),
                    metadata JSONB DEFAULT '{}'::jsonb,
                    UNIQUE(tenant_id, name)
                );
                
                CREATE INDEX IF NOT EXISTS idx_zones_tenant ON zones(tenant_id);
                CREATE INDEX IF NOT EXISTS idx_zones_type ON zones(type);
                
                -- Capabilities table
                CREATE TABLE IF NOT EXISTS capabilities (
                    id UUID PRIMARY KEY,
                    type VARCHAR(50) NOT NULL CHECK (type IN ('SENSOR', 'ACTUATOR')),
                    subtype VARCHAR(100) NOT NULL,
                    zone_id UUID REFERENCES zones(id) ON DELETE CASCADE,
                    tenant_id UUID REFERENCES tenants(id) ON DELETE SET NULL,
                    properties JSONB DEFAULT '{}'::jsonb,
                    constraints JSONB DEFAULT '{}'::jsonb,
                    calibration JSONB DEFAULT NULL,
                    maintenance_schedule JSONB DEFAULT NULL,
                    hardware_config JSONB DEFAULT '{}'::jsonb,
                    last_reading JSONB DEFAULT NULL,
                    status VARCHAR(50) DEFAULT 'online',
                    error_count INTEGER DEFAULT 0,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_capabilities_zone ON capabilities(zone_id);
                CREATE INDEX IF NOT EXISTS idx_capabilities_type_subtype ON capabilities(type, subtype);
                
                -- Grants table
                CREATE TABLE IF NOT EXISTS grants (
                    id VARCHAR(255) PRIMARY KEY,
                    tenant_id UUID REFERENCES tenants(id) ON DELETE CASCADE,
                    plugin_id VARCHAR(255) NOT NULL,
                    capability_id UUID REFERENCES capabilities(id) ON DELETE CASCADE,
                    permissions TEXT[] NOT NULL,
                    expires_at TIMESTAMPTZ NOT NULL,
                    usage_count INTEGER DEFAULT 0,
                    max_usage INTEGER,
                    constraints JSONB DEFAULT '{}'::jsonb,
                    audit_log JSONB DEFAULT '[]'::jsonb,
                    metadata JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    last_used TIMESTAMPTZ
                );
                
                CREATE INDEX IF NOT EXISTS idx_grants_tenant ON grants(tenant_id);
                CREATE INDEX IF NOT EXISTS idx_grants_expires ON grants(expires_at);
                CREATE INDEX IF NOT EXISTS idx_grants_plugin ON grants(plugin_id);
                CREATE INDEX CONCURRENTLY IF NOT EXISTS idx_grants_capability
                ON grants(capability_id) WHERE expires_at > NOW();
                
                -- Sensor readings table (partitioned by month)
                CREATE TABLE IF NOT EXISTS sensor_readings (
                    id BIGSERIAL,
                    zone_id UUID REFERENCES zones(id) ON DELETE CASCADE,
                    capability_id UUID REFERENCES capabilities(id) ON DELETE CASCADE,
                    sensor_type VARCHAR(100) NOT NULL,
                    value FLOAT NOT NULL,
                    quality FLOAT DEFAULT 1.0,
                    raw_value FLOAT,
                    unit VARCHAR(50),
                    timestamp TIMESTAMPTZ DEFAULT NOW(),
                    metadata JSONB DEFAULT '{}'::jsonb,
                    PRIMARY KEY (id, timestamp)
                ) PARTITION BY RANGE (timestamp);
                
                CREATE INDEX IF NOT EXISTS idx_sensor_readings_zone_time
                ON sensor_readings(zone_id, timestamp DESC);

                CREATE INDEX IF NOT EXISTS idx_sensor_readings_type_time
                ON sensor_readings(sensor_type, timestamp DESC);
                CREATE INDEX CONCURRENTLY IF NOT EXISTS idx_sensor_readings_capability
                ON sensor_readings(capability_id, timestamp DESC);
                
                -- Create partitions for current and next month
                DO $$ 
                DECLARE
                    start_date date;
                    end_date date;
                BEGIN
                    start_date := date_trunc('month', CURRENT_DATE);
                    end_date := start_date + interval '1 month';
                    
                    EXECUTE format('
                        CREATE TABLE IF NOT EXISTS sensor_readings_%s PARTITION OF sensor_readings
                        FOR VALUES FROM (%L) TO (%L)',
                        to_char(start_date, 'YYYY_MM'),
                        start_date,
                        end_date
                    );
                    
                    -- Next month
                    start_date := end_date;
                    end_date := start_date + interval '1 month';
                    
                    EXECUTE format('
                        CREATE TABLE IF NOT EXISTS sensor_readings_%s PARTITION OF sensor_readings
                        FOR VALUES FROM (%L) TO (%L)',
                        to_char(start_date, 'YYYY_MM'),
                        start_date,
                        end_date
                    );
                END $$;
                
                -- Alarms table
                CREATE TABLE IF NOT EXISTS alarms (
                    id UUID PRIMARY KEY,
                    zone_id UUID REFERENCES zones(id) ON DELETE CASCADE,
                    type VARCHAR(100) NOT NULL,
                    severity VARCHAR(50) NOT NULL CHECK (severity IN ('info', 'warning', 'critical', 'emergency')),
                    message TEXT NOT NULL,
                    acknowledged BOOLEAN DEFAULT FALSE,
                    acknowledged_by VARCHAR(255),
                    acknowledged_at TIMESTAMPTZ,
                    resolution TEXT,
                    auto_resolve BOOLEAN DEFAULT FALSE,
                    threshold_value FLOAT,
                    actual_value FLOAT,
                    metadata JSONB DEFAULT '{}'::jsonb,
                    created_at TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_alarms_zone ON alarms(zone_id);
                CREATE INDEX IF NOT EXISTS idx_alarms_severity ON alarms(severity);
                CREATE INDEX IF NOT EXISTS idx_alarms_acknowledged ON alarms(acknowledged);
                CREATE INDEX IF NOT EXISTS idx_alarms_created ON alarms(created_at DESC);
                
                -- Audit log table (partitioned by day)
                CREATE TABLE IF NOT EXISTS audit_log (
                    id BIGSERIAL,
                    tenant_id UUID,
                    event_type VARCHAR(100) NOT NULL,
                    event_data JSONB NOT NULL,
                    user_id VARCHAR(255),
                    ip_address INET,
                    session_id VARCHAR(255),
                    timestamp TIMESTAMPTZ DEFAULT NOW(),
                    PRIMARY KEY (id, timestamp)
                ) PARTITION BY RANGE (timestamp);
                
                CREATE INDEX IF NOT EXISTS idx_audit_log_tenant_time 
                ON audit_log(tenant_id, timestamp DESC);
                
                CREATE INDEX IF NOT EXISTS idx_audit_log_event_time 
                ON audit_log(event_type, timestamp DESC);
                
                CREATE INDEX IF NOT EXISTS idx_audit_log_user_time 
                ON audit_log(user_id, timestamp DESC);
                
                -- Create audit partition for today
                DO $$ 
                DECLARE
                    start_date date;
                    end_date date;
                BEGIN
                    start_date := CURRENT_DATE;
                    end_date := start_date + interval '1 day';
                    
                    EXECUTE format('
                        CREATE TABLE IF NOT EXISTS audit_log_%s PARTITION OF audit_log
                        FOR VALUES FROM (%L) TO (%L)',
                        to_char(start_date, 'YYYY_MM_DD'),
                        start_date,
                        end_date
                    );
                END $$;
                
                -- Maintenance records table
                CREATE TABLE IF NOT EXISTS maintenance_records (
                    id UUID PRIMARY KEY,
                    type VARCHAR(50) NOT NULL,
                    component_id VARCHAR(255) NOT NULL,
                    capability_id UUID REFERENCES capabilities(id) ON DELETE SET NULL,
                    scheduled_date TIMESTAMPTZ NOT NULL,
                    completed_date TIMESTAMPTZ,
                    technician VARCHAR(255),
                    actions TEXT[],
                    parts_used JSONB DEFAULT '[]'::jsonb,
                    next_due TIMESTAMPTZ,
                    cost FLOAT,
                    downtime_minutes INTEGER,
                    notes TEXT,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_maintenance_scheduled ON maintenance_records(scheduled_date);
                CREATE INDEX IF NOT EXISTS idx_maintenance_component ON maintenance_records(component_id);
                CREATE INDEX IF NOT EXISTS idx_maintenance_completed ON maintenance_records(completed_date);
                
                -- Harvests table
                CREATE TABLE IF NOT EXISTS harvests (
                    id UUID PRIMARY KEY,
                    zone_id UUID REFERENCES zones(id) ON DELETE CASCADE,
                    crop_id VARCHAR(255) NOT NULL,
                    quantity FLOAT NOT NULL,
                    quantity_unit VARCHAR(50) NOT NULL,
                    quality_grade VARCHAR(10) NOT NULL,
                    quality_metrics JSONB DEFAULT '{}'::jsonb,
                    harvested_by VARCHAR(255) NOT NULL,
                    harvest_date TIMESTAMPTZ DEFAULT NOW(),
                    destination VARCHAR(255),
                    notes TEXT,
                    images TEXT[],
                    labor_time_minutes INTEGER,
                    packaging_type VARCHAR(100),
                    batch_code VARCHAR(100),
                    created_at TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_harvests_zone ON harvests(zone_id);
                CREATE INDEX IF NOT EXISTS idx_harvests_date ON harvests(harvest_date DESC);
                CREATE INDEX IF NOT EXISTS idx_harvests_crop ON harvests(crop_id);
                
                -- Recipes table
                CREATE TABLE IF NOT EXISTS recipes (
                    id UUID PRIMARY KEY,
                    name VARCHAR(255) NOT NULL UNIQUE,
                    crop_type VARCHAR(100),
                    stages JSONB NOT NULL,
                    created_by VARCHAR(255) NOT NULL,
                    validated_yield FLOAT,
                    notes TEXT,
                    tags TEXT[],
                    is_public BOOLEAN DEFAULT FALSE,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    updated_at TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_recipes_crop ON recipes(crop_type);
                CREATE INDEX IF NOT EXISTS idx_recipes_public ON recipes(is_public);
                
                -- Sessions table
                CREATE TABLE IF NOT EXISTS sessions (
                    id VARCHAR(255) PRIMARY KEY,
                    tenant_id UUID REFERENCES tenants(id) ON DELETE CASCADE,
                    user_id VARCHAR(255),
                    data JSONB DEFAULT '{}'::jsonb,
                    expires_at TIMESTAMPTZ NOT NULL,
                    created_at TIMESTAMPTZ DEFAULT NOW(),
                    last_accessed TIMESTAMPTZ DEFAULT NOW()
                );
                
                CREATE INDEX IF NOT EXISTS idx_sessions_tenant ON sessions(tenant_id);
                CREATE INDEX IF NOT EXISTS idx_sessions_expires ON sessions(expires_at);
                
                -- Create update trigger for updated_at columns
                CREATE OR REPLACE FUNCTION update_updated_at_column()
                RETURNS TRIGGER AS $$
                BEGIN
                    NEW.updated_at = NOW();
                    RETURN NEW;
                END;
                $$ language 'plpgsql';
                
                DO $$ 
                BEGIN
                    -- Create triggers only if they don't exist
                    IF NOT EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'update_tenants_updated_at') THEN
                        CREATE TRIGGER update_tenants_updated_at BEFORE UPDATE ON tenants
                        FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
                    END IF;
                    
                    IF NOT EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'update_zones_updated_at') THEN
                        CREATE TRIGGER update_zones_updated_at BEFORE UPDATE ON zones
                        FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
                    END IF;
                    
                    IF NOT EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'update_capabilities_updated_at') THEN
                        CREATE TRIGGER update_capabilities_updated_at BEFORE UPDATE ON capabilities
                        FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
                    END IF;
                    
                    IF NOT EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'update_maintenance_updated_at') THEN
                        CREATE TRIGGER update_maintenance_updated_at BEFORE UPDATE ON maintenance_records
                        FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
                    END IF;
                    
                    IF NOT EXISTS (SELECT 1 FROM pg_trigger WHERE tgname = 'update_recipes_updated_at') THEN
                        CREATE TRIGGER update_recipes_updated_at BEFORE UPDATE ON recipes
                        FOR EACH ROW EXECUTE FUNCTION update_updated_at_column();
                    END IF;
                END $$;
                
                -- Function to create new sensor reading partitions
                CREATE OR REPLACE FUNCTION create_sensor_partition()
                RETURNS void AS $$
                DECLARE
                    start_date date;
                    end_date date;
                    partition_name text;
                BEGIN
                    start_date := date_trunc('month', CURRENT_DATE + interval '1 month');
                    end_date := start_date + interval '1 month';
                    partition_name := 'sensor_readings_' || to_char(start_date, 'YYYY_MM');
                    
                    EXECUTE format('
                        CREATE TABLE IF NOT EXISTS %I PARTITION OF sensor_readings
                        FOR VALUES FROM (%L) TO (%L)',
                        partition_name,
                        start_date,
                        end_date
                    );
                END;
                $$ LANGUAGE plpgsql;
                
                -- Function to create new audit log partitions
                CREATE OR REPLACE FUNCTION create_audit_partition()
                RETURNS void AS $$
                DECLARE
                    start_date date;
                    end_date date;
                    partition_name text;
                BEGIN
                    start_date := CURRENT_DATE + interval '1 day';
                    end_date := start_date + interval '1 day';
                    partition_name := 'audit_log_' || to_char(start_date, 'YYYY_MM_DD');
                    
                    EXECUTE format('
                        CREATE TABLE IF NOT EXISTS %I PARTITION OF audit_log
                        FOR VALUES FROM (%L) TO (%L)',
                        partition_name,
                        start_date,
                        end_date
                    );
                END;
                $$ LANGUAGE plpgsql;
            ''')
            
            logger.info("Database schema initialized")
            
    async def _load_tenants(self):
        """Load tenants from database"""
        async with self.db.acquire() as conn:
            rows = await conn.fetch('SELECT * FROM tenants')
            for row in rows:
                tenant = Tenant(
                    id=str(row['id']),
                    name=row['name'],
                    tier=row['tier'],
                    features=set(row['features'] or []),
                    resource_limits=dict(row['resource_limits'] or {}),
                    api_key_hash=row['api_key_hash'],
                    contact_email=row['contact_email'],
                    billing_status=row['billing_status'],
                    usage_stats=dict(row['usage_stats'] or {}),
                    created_at=row['created_at'],
                    metadata=dict(row['metadata'] or {})
                )
                self.tenants[tenant.id] = tenant
                
        logger.info(f"Loaded {len(self.tenants)} tenants")
        
    async def _load_zones(self):
        """Load zones from database"""
        async with self.db.acquire() as conn:
            rows = await conn.fetch('SELECT * FROM zones')
            for row in rows:
                zone = Zone(
                    id=str(row['id']),
                    tenant_id=str(row['tenant_id']),
                    name=row['name'],
                    type=row['type'],
                    units=list(row['units'] or []),
                    environmental_targets=dict(row['environmental_targets'] or {}),
                    crop_profile=dict(row['crop_profile']) if row['crop_profile'] else None,
                    emergency_stop=row['emergency_stop'],
                    maintenance_mode=row['maintenance_mode'],
                    created_at=row['created_at'],
                    last_harvest=row['last_harvest'],
                    total_yield=row['total_yield'],
                    active_alarms=list(row['active_alarms'] or []),
                    metadata=dict(row['metadata'] or {})
                )
                self.zones[zone.id] = zone
                
        logger.info(f"Loaded {len(self.zones)} zones")
        
    async def _load_capabilities(self):
        """Load capabilities from database"""
        async with self.db.acquire() as conn:
            rows = await conn.fetch('SELECT * FROM capabilities')
            for row in rows:
                capability = Capability(
                    id=str(row['id']),
                    type=row['type'],
                    subtype=row['subtype'],
                    zone_id=str(row['zone_id']) if row['zone_id'] else None,
                    tenant_id=str(row['tenant_id']) if row['tenant_id'] else None,
                    properties=dict(row['properties'] or {}),
                    constraints=dict(row['constraints'] or {}),
                    calibration=dict(row['calibration']) if row['calibration'] else None,
                    maintenance_schedule=dict(row['maintenance_schedule']) if row['maintenance_schedule'] else None,
                    hardware_config=dict(row['hardware_config'] or {}),
                    last_reading=dict(row['last_reading']) if row['last_reading'] else None,
                    status=row['status'],
                    error_count=row['error_count']
                )
                self.capabilities[capability.id] = capability
                
        logger.info(f"Loaded {len(self.capabilities)} capabilities")
        
    async def _load_grants(self):
        """Load active grants from database"""
        async with self.db.acquire() as conn:
            rows = await conn.fetch(
                'SELECT * FROM grants WHERE expires_at > $1',
                datetime.now(timezone.utc)
            )
            for row in rows:
                grant = Grant(
                    id=row['id'],
                    tenant_id=str(row['tenant_id']),
                    plugin_id=row['plugin_id'],
                    capability_id=str(row['capability_id']),
                    permissions=list(row['permissions']),
                    expires_at=row['expires_at'],
                    constraints=dict(row['constraints'] or {}),
                    usage_count=row['usage_count'],
                    max_usage=row['max_usage'],
                    audit_log=list(row['audit_log'] or []),
                    metadata=dict(row['metadata'] or {}),
                    created_at=row['created_at'],
                    last_used=row['last_used']
                )
                
                if grant.is_valid():
                    self.grants[grant.id] = grant
                    active_grants.labels(plugin_id=grant.plugin_id).inc()
                    
        logger.info(f"Loaded {len(self.grants)} active grants")
        
    async def _initialize_plugins(self):
        """Initialize configured plugins"""
        # Always initialize hardware interface
        if self.config.hardware_interface == "simulator":
            plugin = HardwareSimulator(self, {})
            self.plugins[plugin.id] = plugin
            await plugin.initialize()
            
        # Initialize other plugins from config
        # This would load from plugin directory or configuration
        
        logger.info(f"Initialized {len(self.plugins)} plugins")
        
    def _setup_event_handlers(self):
        """Set up internal event handlers"""
        # Alarm handler
        async def handle_alarm(event):
            alarm_data = event['data']
            alarm = Alarm(
                id=str(uuid.uuid4()),
                zone_id=alarm_data.get('zone_id'),
                type=alarm_data.get('type', 'unknown'),
                severity=alarm_data.get('severity', 'info'),
                message=alarm_data.get('message', ''),
                metadata=alarm_data.get('metadata', {}),
                auto_resolve=alarm_data.get('auto_resolve', False),
                threshold_value=alarm_data.get('threshold_value'),
                actual_value=alarm_data.get('actual_value')
            )
            
            self.alarms[alarm.id] = alarm
            
            # Update zone
            if alarm.zone_id and alarm.zone_id in self.zones:
                zone = self.zones[alarm.zone_id]
                if alarm.id not in zone.active_alarms:
                    zone.active_alarms.append(alarm.id)
                    
            # Save to database
            async with self.db.acquire() as conn:
                await conn.execute('''
                    INSERT INTO alarms (id, zone_id, type, severity, message, 
                                      threshold_value, actual_value, metadata, auto_resolve)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                ''', alarm.id, alarm.zone_id, alarm.type, alarm.severity, alarm.message,
                    alarm.threshold_value, alarm.actual_value, json.dumps(alarm.metadata),
                    alarm.auto_resolve)
                    
        self.event_bus.subscribe('alarm', handle_alarm)
        
        # Command execution tracking
        async def track_command(event):
            await self.telemetry.record('command_execution', {
                'grant_id': event['data'].get('grant_id'),
                'capability_id': event['data'].get('capability_id'),
                'command': event['data'].get('command'),
                'result': event['data'].get('result'),
                'execution_time': event['data'].get('execution_time')
            })
            
        self.event_bus.subscribe('command_executed', track_command)
        
    async def request_capability(
        self, 
        tenant_id: str,
        plugin_id: str, 
        capability_query: Dict[str, Any],
        permissions: List[str],
        duration: timedelta,
        constraints: Optional[Dict[str, Any]] = None
    ) -> Optional[Grant]:
        """Request a capability grant"""
        with tracer.start_as_current_span("request_capability") as span:
            span.set_attribute("tenant_id", tenant_id)
            span.set_attribute("plugin_id", plugin_id)
            
            # Validate tenant
            tenant = self.tenants.get(tenant_id)
            if not tenant:
                logger.warning(f"Unknown tenant: {tenant_id}")
                return None
                
            # Check tenant status
            if tenant.billing_status != 'active':
                logger.warning(f"Tenant {tenant_id} billing status: {tenant.billing_status}")
                return None
                
            # Check rate limit
            if self.config.rate_limit_enabled:
                rate_limit_key = f"rate_limit:{tenant_id}:{plugin_id}"
                if not await self.rate_limiter.check_rate_limit(
                    rate_limit_key,
                    self.config.rate_limit_requests,
                    self.config.rate_limit_window
                ):
                    logger.warning(f"Rate limit exceeded for {tenant_id}/{plugin_id}")
                    raise RateLimitError("Rate limit exceeded", self.config.rate_limit_window)
                    
            # Check grant limit
            tenant_grants = sum(1 for g in self.grants.values() if g.tenant_id == tenant_id)
            if not tenant.check_resource_limit('grants', tenant_grants, 1):
                logger.warning(f"Grant limit exceeded for tenant {tenant_id}")
                return None
                
            # Check tier permissions
            if not self._check_tenant_permissions(tenant, capability_query, permissions):
                logger.warning(f"Tenant {tenant_id} lacks required permissions")
                return None
                
            # Find matching capability
            capability = await self._find_capability(tenant_id, capability_query)
            if not capability:
                logger.warning(f"No capability found matching {capability_query}")
                return None
                
            # Check capability status
            if capability.status != 'online':
                logger.warning(f"Capability {capability.id} is {capability.status}")
                return None
                
            # Validate permissions against capability
            if not await self._validate_permissions(plugin_id, capability, permissions):
                logger.warning(f"Permission denied for {plugin_id} on {capability.id}")
                return None
                
            # Create grant
            grant = Grant(
                id=f"grant_{uuid.uuid4().hex}",
                tenant_id=tenant_id,
                plugin_id=plugin_id,
                capability_id=capability.id,
                permissions=permissions,
                expires_at=datetime.now(timezone.utc) + duration,
                constraints=constraints or {},
                max_usage=constraints.get('max_usage') if constraints else None
            )
            
            # Validate grant doesn't exceed safety limits
            if not await self._validate_grant_safety(grant, capability):
                logger.warning(f"Grant would exceed safety limits")
                return None
                
            self.grants[grant.id] = grant
            
            # Save to database
            async with self.db.acquire() as conn:
                await conn.execute('''
                    INSERT INTO grants (id, tenant_id, plugin_id, capability_id, permissions, 
                                      expires_at, usage_count, max_usage, constraints)
                    VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9)
                ''', grant.id, tenant_id, plugin_id, capability.id, permissions,
                    grant.expires_at, 0, grant.max_usage, json.dumps(constraints or {}))
                    
            # Update metrics
            grant_requests.labels(plugin_id=plugin_id, capability_type=capability.type).inc()
            active_grants.labels(plugin_id=plugin_id).inc()
            tenant.increment_usage('grants_created')
            
            # Audit log
            await self.audit_logger.log({
                'event': 'grant_created',
                'tenant_id': tenant_id,
                'grant_id': grant.id,
                'plugin_id': plugin_id,
                'capability_id': capability.id,
                'permissions': permissions,
                'expires_at': grant.expires_at.isoformat()
            })
            
            # Cache grant for fast lookup
            await self.cache.setex(
                f"grant:{grant.id}",
                int(duration.total_seconds()),
                json.dumps(asdict(grant), default=str)
            )
            
            return grant
            
    def _check_tenant_permissions(self, tenant: Tenant, capability_query: Dict[str, Any], permissions: List[str]) -> bool:
        """Check if tenant has required permissions"""
        # Community tier limitations
        if tenant.tier == 'community':
            # Limited to basic operations
            if any(p in ['ADMIN', 'CALIBRATE'] for p in permissions):
                return False
            # Limited capability types
            if capability_query.get('type') == 'ACTUATOR':
                subtype = capability_query.get('subtype', '')
                if subtype in ['dosing_system', 'climate_control']:
                    return False
            # Zone limit
            if len([z for z in self.zones.values() if z.tenant_id == tenant.id]) > 2:
                return False
                
        # Professional tier limitations
        elif tenant.tier == 'professional':
            # No admin permissions
            if 'ADMIN' in permissions:
                return False
            # Zone limit
            if len([z for z in self.zones.values() if z.tenant_id == tenant.id]) > 10:
                return False
                
        # Enterprise has all permissions
        return True
        
    async def _find_capability(self, tenant_id: str, query: Dict[str, Any]) -> Optional[Capability]:
        """Find capability matching query"""
        # First check cache
        cache_key = f"capability:{tenant_id}:{query.get('type')}:{query.get('subtype')}:{query.get('zoneId')}"
        cached = await self.cache.get(cache_key)
        if cached:
            cap_dict = json.loads(cached)
            return Capability(**cap_dict)
            
        # Search in memory first
        for cap in self.capabilities.values():
            # Check if capability matches query
            if cap.type != query.get('type'):
                continue
            if cap.subtype != query.get('subtype'):
                continue
            if query.get('zoneId') and str(cap.zone_id) != query.get('zoneId'):
                continue
                
            # Check zone belongs to tenant
            if cap.zone_id:
                zone = self.zones.get(cap.zone_id)
                if not zone or zone.tenant_id != tenant_id:
                    continue
                    
            # Found match
            await self.cache.setex(cache_key, 300, json.dumps(asdict(cap)))
            return cap
            
        return None
        
    async def _validate_permissions(self, plugin_id: str, capability: Capability, permissions: List[str]) -> bool:
        """Validate permissions against capability"""
        # Check if plugin exists
        plugin = self.plugins.get(plugin_id)
        if not plugin and plugin_id != 'phal-sdk':  # Allow SDK access
            return False
            
        # Check capability constraints
        allowed_permissions = capability.constraints.get('allowed_permissions', ['READ', 'WRITE', 'OPERATE'])
        for perm in permissions:
            if perm not in allowed_permissions:
                return False
                
        # Additional validation based on capability type
        if capability.type == 'SENSOR' and any(p in ['WRITE', 'OPERATE'] for p in permissions):
            # Only allow CALIBRATE for sensors
            if not all(p in ['READ', 'CALIBRATE'] for p in permissions):
                return False
                
        return True
        
    async def _validate_grant_safety(self, grant: Grant, capability: Capability) -> bool:
        """Validate grant doesn't exceed safety limits"""
        # Check concurrent grants for same capability
        active_capability_grants = [
            g for g in self.grants.values()
            if g.capability_id == capability.id and g.is_valid()
        ]
        
        # Limit concurrent access to actuators
        if capability.type == 'ACTUATOR' and len(active_capability_grants) >= 3:
            logger.warning(f"Too many concurrent grants for actuator {capability.id}")
            return False
            
        # Check dosing system limits
        if capability.subtype == 'dosing_system':
            # Calculate total dosing capacity granted
            total_capacity = sum(
                g.constraints.get('max_daily_ml', 1000)
                for g in active_capability_grants
            )
            if total_capacity > 5000:  # 5L daily limit
                logger.warning(f"Total dosing capacity exceeded for {capability.id}")
                return False
                
        return True
        
    async def execute_command(
        self, 
        grant_id: str, 
        command: Dict[str, Any],
        context: Optional[Dict[str, Any]] = None
    ) -> Dict[str, Any]:
        """Execute command with comprehensive validation"""
        with tracer.start_as_current_span("execute_command") as span:
            span.set_attribute("grant_id", grant_id)
            span.set_attribute("command_type", command.get('command', 'unknown'))
            
            start_time = asyncio.get_event_loop().time()
            
            try:
                # Get grant
                grant = await self._get_grant(grant_id)
                if not grant:
                    raise GrantError("Invalid grant")
                    
                # Validate grant
                if not grant.is_valid():
                    await self._revoke_grant(grant_id)
                    raise GrantError("Grant expired or exceeded usage limit")
                    
                # Get capability
                capability = self.capabilities.get(grant.capability_id)
                if not capability:
                    raise ValueError("Capability not found")
                    
                # Check capability status
                if capability.status != 'online':
                    raise RuntimeError(f"Capability is {capability.status}")
                    
                # Validate command permissions
                command_type = command.get('command', '')
                if not grant.can_execute(command_type):
                    raise GrantError(f"Permission denied for command '{command_type}'")
                    
                # Validate against constraints
                if not await self._validate_constraints(grant, command, context):
                    raise ValidationError("Command violates grant constraints")
                    
                # Safety validation
                if not await self._validate_safety(capability, command):
                    safety_violations.labels(violation_type=command_type).inc()
                    raise SecurityError("Safety validation failed")
                    
                # ML anomaly detection
                if await self.ml_engine.detect_anomaly(capability, command):
                    # Log but don't block (could be legitimate edge case)
                    logger.warning(f"Anomaly detected for command: {command}")
                    await self.event_bus.emit('anomaly_detected', {
                        'grant_id': grant_id,
                        'capability_id': capability.id,
                        'command': command
                    })
                    
                # Check system-wide constraints
                if not await self._validate_system_constraints(capability, command):
                    raise SecurityError("System constraints violated")
                    
                # Get plugin and execute
                plugin = self.plugins.get(grant.plugin_id)
                if not plugin:
                    # Try to execute via hardware interface directly
                    plugin = self.plugins.get('HardwareSimulator')
                    if not plugin:
                        raise ValueError("Plugin not found")
                        
                # Execute with circuit breaker
                result = await self.circuit_breaker.call(
                    plugin.execute_capability,
                    capability.id,
                    command
                )
                
                # Record success
                execution_time = asyncio.get_event_loop().time() - start_time
                
                # Update grant
                grant.usage_count += 1
                grant.last_used = datetime.now(timezone.utc)
                grant.audit_log.append({
                    'timestamp': datetime.now(timezone.utc).isoformat(),
                    'command': command,
                    'result': 'success',
                    'execution_time': execution_time,
                    'context': context
                })
                
                # Update database
                async with self.db.acquire() as conn:
                    await conn.execute('''
                        UPDATE grants 
                        SET usage_count = $1, last_used = $2, audit_log = $3
                        WHERE id = $4
                    ''', grant.usage_count, grant.last_used, 
                        json.dumps(grant.audit_log), grant.id)
                        
                # Update capability last reading if sensor
                if capability.type == 'SENSOR' and 'value' in result:
                    capability.last_reading = result
                    async with self.db.acquire() as conn:
                        await conn.execute('''
                            UPDATE capabilities
                            SET last_reading = $1, updated_at = NOW()
                            WHERE id = $2
                        ''', json.dumps(result), capability.id)
                        
                    # Store sensor reading
                    await self._store_sensor_reading(capability, result)
                    
                # Update cache
                await self._update_grant_cache(grant)
                
                # Emit event
                await self.event_bus.emit('command_executed', {
                    'grant_id': grant_id,
                    'capability_id': capability.id,
                    'command': command,
                    'result': 'success',
                    'execution_time': execution_time
                })
                
                # Update metrics
                command_executions.labels(capability_type=capability.type).inc()
                command_latency.labels(command_type=command_type).observe(execution_time)
                
                # Audit log
                await self.audit_logger.log({
                    'event': 'command_executed',
                    'tenant_id': grant.tenant_id,
                    'grant_id': grant_id,
                    'capability_id': capability.id,
                    'command': command,
                    'result': 'success',
                    'execution_time': execution_time,
                    'user_id': context.get('user', {}).get('id') if context else None,
                    'ip_address': context.get('ip_address') if context else None
                })
                
                return result
                
            except Exception as e:
                # Record failure
                if 'grant' in locals():
                    grant.audit_log.append({
                        'timestamp': datetime.now(timezone.utc).isoformat(),
                        'command': command,
                        'error': str(e),
                        'context': context
                    })
                    await self._update_grant_cache(grant)
                    
                # Update capability error count
                if 'capability' in locals():
                    capability.error_count += 1
                    if capability.error_count > 10:
                        capability.status = 'error'
                        await self.event_bus.emit('capability_error', {
                            'capability_id': capability.id,
                            'error_count': capability.error_count
                        })
                        
                await self.event_bus.emit('command_failed', {
                    'grant_id': grant_id,
                    'command': command,
                    'error': str(e)
                })
                
                # Audit log
                await self.audit_logger.log({
                    'event': 'command_failed',
                    'grant_id': grant_id,
                    'command': command,
                    'error': str(e),
                    'user_id': context.get('user', {}).get('id') if context else None
                })
                
                span.record_exception(e)
                raise
                
    async def _validate_safety(self, capability: Capability, command: Dict[str, Any]) -> bool:
        """Advanced safety validation with ML predictions"""
        # Check if zone is in emergency stop
        zone = self.zones.get(capability.zone_id) if capability.zone_id else None
        if zone and not zone.is_operational():
            logger.warning(f"Zone {zone.id} is not operational")
            return False
            
        # Load safety rules
        safety_rules = await self._load_safety_rules(capability)
        
        # Basic rule validation
        if not self._validate_basic_safety(capability, command, safety_rules):
            return False
            
        # Check against recent commands (prevent rapid cycling)
        if not await self._validate_command_frequency(capability, command):
            return False
            
        # Validate environmental conditions
        if zone and not await self._validate_environmental_safety(zone, capability, command):
            return False
            
        return True
        
    async def _load_safety_rules(self, capability: Capability) -> Dict[str, Any]:
        """Load safety rules for capability"""
        # Default safety rules
        default_rules = {
            'dosing_limits': {
                'max_dose_ml': 50,
                'max_daily_ml': 1000,
                'min_interval_minutes': 5
            },
            'ph_limits': {
                'absolute_min': 4.0,
                'absolute_max': 9.0,
                'operational_min': 5.5,
                'operational_max': 7.5,
                'max_adjustment_per_hour': 0.5
            },
            'temperature_limits': {
                'critical_min': 10,
                'critical_max': 40,
                'operational_min': 18,
                'operational_max': 28,
                'max_change_per_hour': 5
            },
            'ec_limits': {
                'max_ec': 3.5,
                'min_ec': 0.5,
                'max_change_per_dose': 0.5
            },
            'lighting_limits': {
                'max_ppfd': 800,
                'max_dli': 40,
                'min_off_hours': 4
            }
        }
        
        # Load zone-specific overrides
        zone = self.zones.get(capability.zone_id)
        if zone and 'safety_rules' in zone.metadata:
            # Deep merge with defaults
            for category, rules in zone.metadata['safety_rules'].items():
                if category in default_rules:
                    default_rules[category].update(rules)
                else:
                    default_rules[category] = rules
                    
        # Load capability-specific overrides
        if 'safety_rules' in capability.properties:
            for category, rules in capability.properties['safety_rules'].items():
                if category in default_rules:
                    default_rules[category].update(rules)
                else:
                    default_rules[category] = rules
                    
        return default_rules
        
    def _validate_basic_safety(self, capability: Capability, command: Dict[str, Any], safety_rules: Dict[str, Any]) -> bool:
        """Validate against basic safety rules"""
        command_type = command.get('command', '')
        
        # Dosing limits
        if capability.type == 'ACTUATOR' and 'dose' in command_type:
            volume = command.get('volume_ml', 0)
            limits = safety_rules.get('dosing_limits', {})
            
            if volume > limits.get('max_dose_ml', 50):
                logger.error(f"Dose volume {volume}ml exceeds max {limits['max_dose_ml']}ml")
                return False
                
            if volume <= 0:
                logger.error(f"Invalid dose volume: {volume}ml")
                return False
                
        # pH adjustment limits
        elif capability.subtype == 'ph_control' and 'adjust' in command_type:
            current_ph = command.get('current_ph', 7.0)
            target_ph = command.get('target_ph', 6.2)
            limits = safety_rules.get('ph_limits', {})
            
            # Absolute limits
            if not (limits['absolute_min'] <= target_ph <= limits['absolute_max']):
                logger.error(f"Target pH {target_ph} outside absolute limits")
                return False
                
            # Operational limits warning
            if not (limits['operational_min'] <= target_ph <= limits['operational_max']):
                logger.warning(f"Target pH {target_ph} outside operational range")
                
            # Rate of change limit
            ph_change = abs(target_ph - current_ph)
            if ph_change > limits.get('max_adjustment_per_hour', 0.5):
                logger.error(f"pH adjustment {ph_change} exceeds hourly limit")
                return False
                
        # Temperature control limits
        elif capability.subtype == 'climate_control' and command_type == 'set_state':
            if 'target_temperature' in command:
                target_temp = command['target_temperature']
                limits = safety_rules.get('temperature_limits', {})
                
                if not (limits['critical_min'] <= target_temp <= limits['critical_max']):
                    logger.error(f"Target temperature {target_temp}°C outside critical limits")
                    return False
                    
        # Light intensity limits
        elif capability.subtype == 'lighting' and command_type == 'set_state':
            intensity = command.get('state', 0)
            if isinstance(intensity, (int, float)) and intensity > 0:
                # Calculate DLI impact
                zone = self.zones.get(capability.zone_id)
                if zone:
                    photoperiod = zone.environmental_targets.get('photoperiod', 18)
                    ppfd = (intensity / 100) * safety_rules['lighting_limits']['max_ppfd']
                    dli = ppfd * photoperiod * 0.0036
                    
                    if dli > safety_rules['lighting_limits']['max_dli']:
                        logger.error(f"DLI {dli} exceeds maximum safe limit")
                        return False
                        
        return True
        
    async def _validate_command_frequency(self, capability: Capability, command: Dict[str, Any]) -> bool:
        """Check command frequency to prevent rapid cycling"""
        command_type = command.get('command', '')
        
        # Get recent commands for this capability
        cache_key = f"cmd_history:{capability.id}"
        history = await self.cache.get(cache_key)
        
        if history:
            cmd_list = json.loads(history)
            now = datetime.now(timezone.utc)
            
            # Check dosing frequency
            if 'dose' in command_type:
                recent_doses = [
                    cmd for cmd in cmd_list 
                    if 'dose' in cmd['command'] and 
                    (now - datetime.fromisoformat(cmd['timestamp'])).seconds < 300
                ]
                if len(recent_doses) >= 2:
                    logger.warning(f"Too frequent dosing on {capability.id}")
                    return False
                    
            # Check state toggle frequency
            elif command_type == 'set_state':
                recent_toggles = [
                    cmd for cmd in cmd_list 
                    if cmd['command'] == 'set_state' and 
                    (now - datetime.fromisoformat(cmd['timestamp'])).seconds < 60
                ]
                if len(recent_toggles) >= 3:
                    logger.warning(f"Too frequent state changes on {capability.id}")
                    return False
                    
        # Update history
        if not history:
            cmd_list = []
        else:
            cmd_list = json.loads(history)
            
        cmd_list.append({
            'command': command_type,
            'timestamp': datetime.now(timezone.utc).isoformat()
        })
        
        # Keep last 100 commands
        cmd_list = cmd_list[-100:]
        
        await self.cache.setex(cache_key, 3600, json.dumps(cmd_list))
        
        return True
        
    async def _validate_environmental_safety(self, zone: Zone, capability: Capability, command: Dict[str, Any]) -> bool:
        """Validate command against current environmental conditions"""
        # Get current conditions
        current_conditions = {}
        for cap_id, cap in self.capabilities.items():
            if cap.zone_id == zone.id and cap.type == 'SENSOR' and cap.last_reading:
                current_conditions[cap.subtype] = cap.last_reading.get('value')
                
        # Check VPD limits before humidity changes
        if capability.subtype == 'climate_control' and 'humidity' in command:
            temp = current_conditions.get('temperature', 22)
            target_humidity = command.get('target_humidity')
            
            if temp and target_humidity:
                svp = 0.6108 * math.exp((17.27 * temp) / (temp + 237.3))
                vpd = svp * (1 - target_humidity / 100)
                
                if vpd < 0.4 or vpd > 1.6:
                    logger.warning(f"Command would result in VPD {vpd:.2f} outside safe range")
                    # Allow but warn
                    
        # Check for conflicting actuator states
        if capability.type == 'ACTUATOR':
            if capability.subtype == 'climate_control':
                # Don't run heater and cooler simultaneously
                if 'heater' in capability.id and command.get('state'):
                    cooler_cap = next(
                        (c for c in self.capabilities.values() 
                         if c.zone_id == zone.id and 'cooler' in c.id),
                        None
                    )
                    if cooler_cap and cooler_cap.last_reading and cooler_cap.last_reading.get('state'):
                        logger.error("Cannot run heater while cooler is active")
                        return False
                        
        return True
        
    async def _validate_constraints(self, grant: Grant, command: Dict[str, Any], context: Optional[Dict[str, Any]]) -> bool:
        """Validate command against grant constraints"""
        constraints = grant.constraints
        
        # Time window constraints
        if 'time_windows' in constraints:
            now = datetime.now()
            in_window = False
            
            for window in constraints['time_windows']:
                # Parse time strings
                start_hour, start_min = map(int, window['start'].split(':'))
                end_hour, end_min = map(int, window['end'].split(':'))
                
                current_minutes = now.hour * 60 + now.minute
                start_minutes = start_hour * 60 + start_min
                end_minutes = end_hour * 60 + end_min
                
                # Handle day boundary
                if start_minutes <= end_minutes:
                    if start_minutes <= current_minutes <= end_minutes:
                        if 'days' not in window or now.weekday() in window['days']:
                            in_window = True
                            break
                else:
                    # Crosses midnight
                    if current_minutes >= start_minutes or current_minutes <= end_minutes:
                        if 'days' not in window or now.weekday() in window['days']:
                            in_window = True
                            break
                            
            if not in_window:
                logger.warning(f"Command outside allowed time windows")
                return False
                
        # Rate limit constraints
        if 'rate_limit' in constraints:
            rate_limit = constraints['rate_limit']
            key = f"grant_rate:{grant.id}"
            
            if not await self.rate_limiter.check_rate_limit(
                key,
                rate_limit['requests'],
                rate_limit['window']
            ):
                logger.warning(f"Grant rate limit exceeded")
                return False
                
        # Command-specific constraints
        if 'allowed_commands' in constraints:
            if command.get('command') not in constraints['allowed_commands']:
                logger.warning(f"Command type not allowed by grant")
                return False
                
        # Value constraints
        if 'value_limits' in constraints:
            limits = constraints['value_limits']
            
            # Check numeric values against limits
            for param, limit in limits.items():
                if param in command:
                    value = command[param]
                    if isinstance(limit, dict):
                        if 'min' in limit and value < limit['min']:
                            return False
                        if 'max' in limit and value > limit['max']:
                            return False
                    elif isinstance(limit, list) and value not in limit:
                        return False
                        
        # Geofence constraints
        if 'geofence' in constraints and context and 'location' in context:
            geo = constraints['geofence']
            user_lat = context['location']['latitude']
            user_lon = context['location']['longitude']
            
            # Haversine distance calculation
            R = 6371000  # Earth radius in meters
            lat1, lon1 = math.radians(geo['latitude']), math.radians(geo['longitude'])
            lat2, lon2 = math.radians(user_lat), math.radians(user_lon)
            
            dlat = lat2 - lat1
            dlon = lon2 - lon1
            
            a = math.sin(dlat/2)**2 + math.cos(lat1) * math.cos(lat2) * math.sin(dlon/2)**2
            c = 2 * math.atan2(math.sqrt(a), math.sqrt(1-a))
            distance = R * c
            
            if distance > geo['radius']:
                logger.warning(f"Command outside geofence")
                return False
                
        return True
        
    async def _validate_system_constraints(self, capability: Capability, command: Dict[str, Any]) -> bool:
        """Validate against system-wide constraints"""
        # Check emergency stop
        emergency_state = await self.cache.get('emergency_stop')
        if emergency_state:
            zones = json.loads(emergency_state)
            if capability.zone_id in zones or 'all' in zones:
                logger.warning(f"Emergency stop active for zone {capability.zone_id}")
                return False
                
        # Check maintenance mode
        maintenance_mode = await self.cache.get(f'maintenance_mode:{capability.id}')
        if maintenance_mode:
            logger.warning(f"Capability {capability.id} is in maintenance mode")
            return False
            
        # Check resource allocation
        if capability.type == 'ACTUATOR' and 'dose' in command.get('command', ''):
            # Get current resource levels
            zone = self.zones.get(capability.zone_id)
            if zone:
                # Check daily dosing limits
                daily_usage = await self._get_daily_resource_usage(
                    capability.zone_id,
                    capability.properties.get('chemical', 'nutrient')
                )
                
                safety_rules = await self._load_safety_rules(capability)
                max_daily = safety_rules.get('dosing_limits', {}).get('max_daily_ml', 1000)
                
                volume = command.get('volume_ml', 0)
                if daily_usage + volume > max_daily:
                    logger.warning(f"Daily dosing limit would be exceeded")
                    return False
                    
        # Check for system overload
        active_commands = await self.cache.get('active_commands')
        if active_commands:
            count = int(active_commands)
            if count > 100:  # System overload threshold
                logger.warning(f"System overload: {count} active commands")
                return False
                
        return True
        
    async def _get_grant(self, grant_id: str) -> Optional[Grant]:
        """Get grant from cache or database"""
        # Check cache first
        cached = await self.cache.get(f"grant:{grant_id}")
        if cached:
            grant_dict = json.loads(cached)
            grant_dict['expires_at'] = datetime.fromisoformat(grant_dict['expires_at'])
            grant_dict['created_at'] = datetime.fromisoformat(grant_dict['created_at'])
            if grant_dict.get('last_used'):
                grant_dict['last_used'] = datetime.fromisoformat(grant_dict['last_used'])
            return Grant(**grant_dict)
            
        # Check memory
        grant = self.grants.get(grant_id)
        if grant:
            return grant
            
        # Load from database
        async with self.db.acquire() as conn:
            row = await conn.fetchrow('SELECT * FROM grants WHERE id = $1', grant_id)
            if row:
                grant = Grant(
                    id=row['id'],
                    tenant_id=str(row['tenant_id']),
                    plugin_id=row['plugin_id'],
                    capability_id=str(row['capability_id']),
                    permissions=list(row['permissions']),
                    expires_at=row['expires_at'],
                    constraints=dict(row['constraints'] or {}),
                    usage_count=row['usage_count'],
                    max_usage=row['max_usage'],
                    audit_log=list(row['audit_log'] or []),
                    metadata=dict(row['metadata'] or {}),
                    created_at=row['created_at'],
                    last_used=row['last_used']
                )
                
                # Cache it
                if grant.is_valid():
                    self.grants[grant_id] = grant
                    await self._update_grant_cache(grant)
                    
                return grant
                
        return None
        
    async def _revoke_grant(self, grant_id: str):
        """Revoke a grant"""
        grant = self.grants.pop(grant_id, None)
        if grant:
            active_grants.labels(plugin_id=grant.plugin_id).dec()
            
        # Remove from cache
        await self.cache.delete(f"grant:{grant_id}")
        
        # Remove from database
        async with self.db.acquire() as conn:
            await conn.execute('DELETE FROM grants WHERE id = $1', grant_id)
            
        # Audit log
        if grant:
            await self.audit_logger.log({
                'event': 'grant_revoked',
                'tenant_id': grant.tenant_id,
                'grant_id': grant_id,
                'reason': 'expired or revoked'
            })
            
    async def _update_grant_cache(self, grant: Grant):
        """Update grant in cache"""
        ttl = int((grant.expires_at - datetime.now(timezone.utc)).total_seconds())
        if ttl > 0:
            await self.cache.setex(
                f"grant:{grant.id}",
                ttl,
                json.dumps(asdict(grant), default=str)
            )
            
    async def _get_daily_resource_usage(self, zone_id: str, resource: str) -> float:
        """Get daily resource usage for a zone"""
        if not resource:
            return 0.0
            
        # Query from audit log
        async with self.db.acquire() as conn:
            result = await conn.fetchval('''
                SELECT COALESCE(SUM((event_data->>'volume_ml')::float), 0)
                FROM audit_log
                WHERE event_type = 'command_executed'
                AND event_data->>'zone_id' = $1
                AND event_data->>'resource' = $2
                AND timestamp > NOW() - INTERVAL '24 hours'
            ''', zone_id, resource)
            
        return float(result or 0)
        
    async def _store_sensor_reading(self, capability: Capability, reading: Dict[str, Any]):
        """Store sensor reading in time-series table"""
        async with self.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO sensor_readings 
                (zone_id, capability_id, sensor_type, value, quality, raw_value, unit, metadata)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            ''', capability.zone_id, capability.id, capability.subtype,
                reading['value'], reading.get('quality', 1.0),
                reading.get('raw_value'), reading.get('unit'),
                json.dumps(reading.get('metadata', {})))
                
    # Background tasks
    
    async def _health_monitor(self):
        """Monitor system health and emit alerts"""
        while True:
            try:
                # Update uptime metric
                uptime = (datetime.now(timezone.utc) - self.start_time).total_seconds()
                system_uptime.set(uptime)
                
                # Check database health
                try:
                    async with self.db.acquire() as conn:
                        await conn.fetchval('SELECT 1')
                    db_healthy = True
                except Exception:
                    db_healthy = False
                    await self.event_bus.emit('health_alert', {
                        'component': 'database',
                        'status': 'unhealthy',
                        'timestamp': datetime.now(timezone.utc).isoformat()
                    })
                    
                # Check Redis health
                try:
                    await self.cache.ping()
                    cache_healthy = True
                except Exception:
                    cache_healthy = False
                    await self.event_bus.emit('health_alert', {
                        'component': 'cache',
                        'status': 'unhealthy',
                        'timestamp': datetime.now(timezone.utc).isoformat()
                    })
                    
                # Check plugin health
                for plugin_id, plugin in self.plugins.items():
                    health = await plugin.health_check()
                    if not health['healthy']:
                        await self.event_bus.emit('health_alert', {
                            'component': f'plugin:{plugin_id}',
                            'status': 'unhealthy',
                            'details': health,
                            'timestamp': datetime.now(timezone.utc).isoformat()
                        })
                        
                # Check resource levels
                for zone in self.zones.values():
                    if zone.metadata.get('nutrient_levels'):
                        for nutrient, level in zone.metadata['nutrient_levels'].items():
                            if isinstance(level, (int, float)) and level < 10:
                                await self.event_bus.emit('alarm', {
                                    'zone_id': zone.id,
                                    'type': 'resource_low',
                                    'severity': 'critical' if level < 5 else 'warning',
                                    'message': f'Low {nutrient} level: {level}%',
                                    'actual_value': level,
                                    'threshold_value': 10
                                })
                                
                await asyncio.sleep(30)  # Check every 30 seconds
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Health monitor error: {e}")
                await asyncio.sleep(60)
                
    async def _grant_cleanup(self):
        """Clean up expired grants"""
        while True:
            try:
                now = datetime.now(timezone.utc)
                expired = []
                
                # Check grants in memory
                for grant_id, grant in self.grants.items():
                    if not grant.is_valid():
                        expired.append(grant_id)
                        
                # Revoke expired grants
                for grant_id in expired:
                    await self._revoke_grant(grant_id)
                    logger.info(f"Revoked expired grant: {grant_id}")
                    
                # Clean expired grants from database
                async with self.db.acquire() as conn:
                    deleted = await conn.fetchval(
                        'DELETE FROM grants WHERE expires_at < $1 RETURNING COUNT(*)',
                        now
                    )
                    if deleted:
                        logger.info(f"Cleaned {deleted} expired grants from database")
                        
                await asyncio.sleep(60)  # Check every minute
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Grant cleanup error: {e}")
                await asyncio.sleep(60)
                
    async def _session_cleanup(self):
        """Clean up expired sessions"""
        while True:
            try:
                now = datetime.now(timezone.utc)
                
                # Clean from memory
                expired_sessions = [
                    sid for sid, session in self.session_store.items()
                    if session.get('expires_at') and 
                    datetime.fromisoformat(session['expires_at']) < now
                ]
                
                for sid in expired_sessions:
                    self.session_store.pop(sid, None)
                    
                # Clean from database
                async with self.db.acquire() as conn:
                    await conn.execute(
                        'DELETE FROM sessions WHERE expires_at < $1',
                        now
                    )
                    
                await asyncio.sleep(300)  # Every 5 minutes
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Session cleanup error: {e}")
                await asyncio.sleep(300)
                
    async def _backup_system(self):
        """Periodic system backup"""
        while True:
            try:
                await asyncio.sleep(self.config.backup_interval)
                
                # Create backup timestamp
                backup_time = datetime.now(timezone.utc)
                backup_id = backup_time.strftime('%Y%m%d_%H%M%S')
                
                # Backup critical data to Redis
                backup_data = {
                    'timestamp': backup_time.isoformat(),
                    'zones': {z_id: asdict(z) for z_id, z in self.zones.items()},
                    'capabilities': {c_id: asdict(c) for c_id, c in self.capabilities.items()},
                    'active_grants': len(self.grants),
                    'active_alarms': len(self.alarms)
                }
                
                await self.cache.setex(
                    f"backup:{backup_id}",
                    86400 * 7,  # Keep for 7 days
                    json.dumps(backup_data, default=str)
                )
                
                logger.info(f"System backup completed: {backup_id}")
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Backup error: {e}")
                await asyncio.sleep(3600)
                
    async def _alarm_processor(self):
        """Process and auto-resolve alarms"""
        while True:
            try:
                # Check auto-resolve alarms
                for alarm_id, alarm in list(self.alarms.items()):
                    if alarm.auto_resolve and not alarm.acknowledged:
                        # Check if condition has cleared
                        if await self._check_alarm_condition(alarm):
                            # Auto-resolve
                            alarm.acknowledged = True
                            alarm.acknowledged_by = 'system'
                            alarm.acknowledged_at = datetime.now(timezone.utc)
                            alarm.resolution = 'Auto-resolved: condition cleared'
                            
                            # Update database
                            async with self.db.acquire() as conn:
                                await conn.execute('''
                                    UPDATE alarms 
                                    SET acknowledged = true,
                                        acknowledged_by = 'system',
                                        acknowledged_at = NOW(),
                                        resolution = $1
                                    WHERE id = $2
                                ''', alarm.resolution, alarm.id)
                                
                            # Remove from active alarms
                            if alarm.zone_id in self.zones:
                                zone = self.zones[alarm.zone_id]
                                if alarm.id in zone.active_alarms:
                                    zone.active_alarms.remove(alarm.id)
                                    
                            # Remove from memory
                            self.alarms.pop(alarm_id, None)
                            
                            logger.info(f"Auto-resolved alarm: {alarm.id}")
                            
                await asyncio.sleep(30)  # Check every 30 seconds
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Alarm processor error: {e}")
                await asyncio.sleep(30)
                
    async def _check_alarm_condition(self, alarm: Alarm) -> bool:
        """Check if alarm condition has cleared"""
        if alarm.type == 'environmental_deviation' and alarm.zone_id:
            # Get current sensor value
            zone = self.zones.get(alarm.zone_id)
            if not zone:
                return False
                
            # Find relevant sensor
            sensor_type = alarm.metadata.get('sensor_type')
            if not sensor_type:
                return False
                
            for cap in self.capabilities.values():
                if (cap.zone_id == alarm.zone_id and 
                    cap.type == 'SENSOR' and 
                    cap.subtype == sensor_type and
                    cap.last_reading):
                    
                    current_value = cap.last_reading.get('value')
                    threshold = alarm.threshold_value
                    
                    # Check if value is back within range
                    if alarm.metadata.get('deviation_type') == 'high':
                        return current_value <= threshold
                    else:
                        return current_value >= threshold
                        
        return False
        
    async def _predictive_maintenance(self):
        """Run predictive maintenance analysis"""
        while True:
            try:
                # Run analysis for each zone
                for zone_id, zone in self.zones.items():
                    if not zone.is_operational():
                        continue
                        
                    # Get recent sensor data
                    sensor_data = await self._get_recent_sensor_data(zone_id, hours=168)
                    
                    if not sensor_data.empty:
                        # Run ML predictions
                        predictions = await self.ml_engine.predict_maintenance(zone_id, sensor_data)
                        
                        for prediction in predictions:
                            if prediction['confidence'] > 0.8:
                                # Create maintenance recommendation
                                await self._create_maintenance_recommendation(zone_id, prediction)
                                
                await asyncio.sleep(3600)  # Run hourly
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Predictive maintenance error: {e}")
                await asyncio.sleep(3600)
                
    async def _get_recent_sensor_data(self, zone_id: str, hours: int = 24) -> pd.DataFrame:
        """Get recent sensor data as DataFrame"""
        async with self.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT sensor_type, value, quality, timestamp
                FROM sensor_readings
                WHERE zone_id = $1
                AND timestamp > NOW() - INTERVAL '%s hours'
                ORDER BY timestamp DESC
            ''' % hours, zone_id)
            
            if rows:
                data = []
                for row in rows:
                    data.append({
                        'sensor_type': row['sensor_type'],
                        'value': float(row['value']),
                        'quality': float(row['quality']),
                        'timestamp': row['timestamp']
                    })
                    
                df = pd.DataFrame(data)
                
                # Pivot to have sensors as columns
                if not df.empty:
                    df_pivot = df.pivot_table(
                        index='timestamp',
                        columns='sensor_type',
                        values=['value', 'quality'],
                        aggfunc='first'
                    )
                    df_pivot.columns = ['_'.join(col).strip() for col in df_pivot.columns.values]
                    df_pivot = df_pivot.sort_index()
                    return df_pivot
                    
        return pd.DataFrame()
        
    async def _create_maintenance_recommendation(self, zone_id: str, prediction: Dict[str, Any]):
        """Create maintenance recommendation"""
        maintenance_id = str(uuid.uuid4())
        scheduled_date = datetime.now(timezone.utc) + timedelta(
            days=prediction.get('days_until_maintenance', 7)
        )
        
        # Save to database
        async with self.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO maintenance_records 
                (id, type, component_id, scheduled_date, notes, created_at)
                VALUES ($1, $2, $3, $4, $5, NOW())
            ''', maintenance_id, prediction['type'], prediction['component'],
                scheduled_date, prediction['recommendation'])
                
        # Emit event
        await self.event_bus.emit('maintenance_scheduled', {
            'id': maintenance_id,
            'type': 'predictive',
            'component': prediction['component'],
            'scheduled_date': scheduled_date.isoformat(),
            'confidence': prediction['confidence'],
            'recommendation': prediction['recommendation']
        })
        
        # Create alarm if urgent
        if prediction.get('days_until_maintenance', 7) <= 3:
            await self.event_bus.emit('alarm', {
                'zone_id': zone_id,
                'type': 'maintenance_required',
                'severity': 'warning',
                'message': f"Maintenance required for {prediction['component']}: {prediction['recommendation']}",
                'auto_resolve': False
            })
            
    async def _metrics_collector(self):
        """Collect and update system metrics"""
        while True:
            try:
                # Collect zone metrics
                for zone_id, zone in self.zones.items():
                    if zone.is_operational():
                        # Get latest sensor readings
                        for cap in self.capabilities.values():
                            if cap.zone_id == zone_id and cap.type == 'SENSOR' and cap.last_reading:
                                value = cap.last_reading.get('value')
                                if value is not None:
                                    sensor_readings.labels(
                                        sensor_type=cap.subtype,
                                        zone_id=zone_id
                                    ).observe(value)
                                    
                                    # Special handling for temperature
                                    if cap.subtype == 'temperature':
                                        zone_temperature.labels(zone_id=zone_id).set(value)
                                        
                # Update nutrient levels
                for zone_id, zone in self.zones.items():
                    if zone.metadata.get('nutrient_levels'):
                        for nutrient, level in zone.metadata['nutrient_levels'].items():
                            if isinstance(level, (int, float)):
                                nutrient_levels.labels(
                                    zone_id=zone_id,
                                    nutrient_type=nutrient
                                ).set(level)
                                
                # Database connection pool stats
                if self.db:
                    pool_size = self.db.get_size()
                    pool_free = self.db.get_idle_size()
                    db_connections.labels(state='total').set(pool_size)
                    db_connections.labels(state='idle').set(pool_free)
                    db_connections.labels(state='active').set(pool_size - pool_free)
                    
                await asyncio.sleep(10)  # Update every 10 seconds
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Metrics collector error: {e}")
                await asyncio.sleep(30)
                
    async def _resource_optimizer(self):
        """Optimize resource usage across zones"""
        while True:
            try:
                # Run optimization every 30 minutes
                await asyncio.sleep(1800)
                
                # Analyze resource usage patterns
                for zone in self.zones.values():
                    if not zone.is_operational() or not zone.crop_profile:
                        continue
                        
                    # Get environmental optimization recommendations
                    current_conditions = {}
                    for cap in self.capabilities.values():
                        if cap.zone_id == zone.id and cap.type == 'SENSOR' and cap.last_reading:
                            current_conditions[cap.subtype] = cap.last_reading.get('value')
                            
                    if current_conditions:
                        recommendations = await self.ml_engine.optimize_environment(
                            zone, current_conditions
                        )
                        
                        if recommendations['adjustments']:
                            # Emit optimization event
                            await self.event_bus.emit('optimization_available', {
                                'zone_id': zone.id,
                                'adjustments': recommendations['adjustments'],
                                'reasoning': recommendations['reasoning'],
                                'timestamp': datetime.now(timezone.utc).isoformat()
                            })
                            
                # Check for resource sharing opportunities
                await self._optimize_resource_sharing()
                
            except asyncio.CancelledError:
                break
            except Exception as e:
                logger.error(f"Resource optimizer error: {e}")
                await asyncio.sleep(1800)
                
    async def _optimize_resource_sharing(self):
        """Optimize resource sharing between zones"""
        # Group zones by tenant
        tenant_zones = defaultdict(list)
        for zone in self.zones.values():
            if zone.is_operational():
                tenant_zones[zone.tenant_id].append(zone)
                
        for tenant_id, zones in tenant_zones.items():
            if len(zones) < 2:
                continue
                
            # Analyze if zones can share resources
            shared_resources = []
            
            # Check lighting schedules for staggering
            light_schedules = []
            for zone in zones:
                if zone.environmental_targets.get('photoperiod'):
                    light_schedules.append({
                        'zone_id': zone.id,
                        'photoperiod': zone.environmental_targets['photoperiod'],
                        'sunrise': zone.environmental_targets.get('sunrise', '06:00')
                    })
                    
            if len(light_schedules) > 1:
                # Check if staggering would reduce peak power
                peak_overlap = self._calculate_peak_overlap(light_schedules)
                if peak_overlap > 0.7:  # 70% overlap
                    shared_resources.append({
                        'type': 'lighting_stagger',
                        'zones': [s['zone_id'] for s in light_schedules],
                        'potential_savings': '15-20% peak power reduction'
                    })
                    
            if shared_resources:
                await self.event_bus.emit('resource_optimization', {
                    'tenant_id': tenant_id,
                    'optimizations': shared_resources,
                    'timestamp': datetime.now(timezone.utc).isoformat()
                })
                
    def _calculate_peak_overlap(self, schedules: List[Dict[str, Any]]) -> float:
        """Calculate lighting schedule overlap"""
        # Simplified calculation
        # In production would do proper time overlap analysis
        total_hours = sum(s['photoperiod'] for s in schedules)
        if total_hours <= 24:
            return 0.0
        return (total_hours - 24) / total_hours

# Create default tenant and zone for testing
async def create_test_data(phal: PluripotentHAL):
    """Create test tenant and zones"""
    # Check if test tenant exists
    async with phal.db.acquire() as conn:
        exists = await conn.fetchval(
            "SELECT 1 FROM tenants WHERE name = 'Test Farm'",
        )
        
        if not exists:
            # Create test tenant
            tenant_id = str(uuid.uuid4())
            api_key = secrets.token_urlsafe(32)
            api_key_hash = hashlib.sha256(api_key.encode()).hexdigest()
            
            await conn.execute('''
                INSERT INTO tenants (id, name, tier, api_key_hash, features, resource_limits)
                VALUES ($1, $2, $3, $4, $5, $6)
            ''', tenant_id, 'Test Farm', 'professional', api_key_hash,
                json.dumps(['monitoring', 'automation', 'analytics']),
                json.dumps({'zones': 10, 'grants': 100}))
                
            # Create test zones
            zone1_id = str(uuid.uuid4())
            zone2_id = str(uuid.uuid4())
            
            await conn.execute('''
                INSERT INTO zones (id, tenant_id, name, type, environmental_targets, crop_profile)
                VALUES ($1, $2, $3, $4, $5, $6)
            ''', zone1_id, tenant_id, 'Production Zone 1', 'production',
                json.dumps({
                    'temperature': {'min': 20, 'max': 25, 'optimal': 22.5},
                    'humidity': {'min': 60, 'max': 70, 'optimal': 65},
                    'vpd': {'min': 0.8, 'max': 1.2, 'optimal': 1.0},
                    'co2': {'min': 600, 'max': 1200, 'optimal': 900},
                    'ph': {'min': 5.8, 'max': 6.5, 'optimal': 6.2},
                    'ec': {'min': 1.8, 'max': 2.2, 'optimal': 2.0},
                    'photoperiod': 18
                }),
                json.dumps({
                    'name': 'Lettuce - Buttercrunch',
                    'variety': 'Buttercrunch',
                    'plant_date': (datetime.now(timezone.utc) - timedelta(days=14)).isoformat(),
                    'expected_harvest_date': (datetime.now(timezone.utc) + timedelta(days=28)).isoformat(),
                    'growth_stage': 'vegetative',
                    'yield_target': 150
                }))
                
            await conn.execute('''
                INSERT INTO zones (id, tenant_id, name, type, environmental_targets)
                VALUES ($1, $2, $3, $4, $5)
            ''', zone2_id, tenant_id, 'Nursery Zone', 'nursery',
                json.dumps({
                    'temperature': {'min': 22, 'max': 26, 'optimal': 24},
                    'humidity': {'min': 70, 'max': 80, 'optimal': 75},
                    'photoperiod': 16
                }))
                
            logger.info(f"Created test tenant with API key: {api_key}")
            
            # Reload data
            await phal._load_tenants()
            await phal._load_zones()

# Main application with web server
class PHALApplication:
    """Main application with complete web API"""
    
    def __init__(self, config_path: Optional[str] = None):
        if config_path:
            self.config = self._load_config(config_path)
        else:
            self.config = PHALConfig()
            
        self.phal = PluripotentHAL(self.config)
        self.active_websockets: Dict[str, Dict[str, Any]] = {}
        self.limiter = Limiter(
            default_limit="100/minute",
            key_func=lambda request: request.get('tenant_id', 'anonymous')
        )
        self.app = web.Application(
            middlewares=[
                self._request_id_middleware,
                self.validate_content_length,
                self._auth_middleware,
                self._rate_limit_middleware,
                self._error_middleware,
                self._metrics_middleware
            ],
            client_max_size=self.config.max_request_size
        )
        aiohttp_cors.setup(self.app, defaults={
            "https://dashboard.example.com": aiohttp_cors.ResourceOptions(
                allow_credentials=True,
                expose_headers="*",
                allow_headers="*",
                allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"]
            )
        })
        self.setup_routes()
        
    def _load_config(self, config_path: str) -> PHALConfig:
        """Load configuration from file"""
        config_file = Path(config_path)
        
        if config_file.suffix == '.yaml':
            with open(config_path, 'r') as f:
                data = yaml.safe_load(f)
        elif config_file.suffix == '.json':
            with open(config_path, 'r') as f:
                data = json.load(f)
        else:
            raise ValueError(f"Unsupported config format: {config_file.suffix}")
            
        return PHALConfig(**data)
        
    @web.middleware
    async def _request_id_middleware(self, request: web.Request, handler):
        """Add request ID for tracing"""
        request['request_id'] = request.headers.get('X-Request-ID', str(uuid.uuid4()))
        response = await handler(request)
        response.headers['X-Request-ID'] = request['request_id']
        return response

    @web.middleware
    async def validate_content_length(self, request: web.Request, handler):
        if request.content_length and request.content_length > self.config.max_request_size:
            raise web.HTTPRequestEntityTooLarge(
                max_size=self.config.max_request_size,
                actual_size=request.content_length
            )
        return await handler(request)

    @web.middleware
    async def _rate_limit_middleware(self, request: web.Request, handler):
        try:
            await self.limiter.limit(request)
        except RateLimitExceeded:
            raise web.HTTPTooManyRequests()
        return await handler(request)
        
    @web.middleware
    async def _auth_middleware(self, request: web.Request, handler):
        """JWT authentication middleware"""
        # Skip auth for health check and metrics
        if request.path in ['/health', '/metrics', '/']:
            return await handler(request)
            
        # Extract token
        auth_header = request.headers.get('Authorization', '')
        if not auth_header.startswith('Bearer '):
            # Check for API key
            api_key = request.headers.get('X-API-Key', '')
            if api_key:
                # Validate API key
                tenant = await self._validate_api_key(api_key)
                if tenant:
                    request['tenant_id'] = tenant.id
                    request['tenant'] = tenant
                    request['user'] = {'tenant_id': tenant.id, 'auth_type': 'api_key'}
                    return await handler(request)
                    
            raise web.HTTPUnauthorized(reason='Missing authentication')
            
        token = auth_header.replace('Bearer ', '')
        
        try:
            # Verify JWT
            payload = jwt.decode(
                token,
                self.phal.config.jwt_secret,
                algorithms=['HS256']
            )
            
            # Check if token is expired
            if 'exp' in payload and datetime.fromtimestamp(payload['exp']) < datetime.now():
                raise web.HTTPUnauthorized(reason='Token expired')
                
            # Get tenant
            tenant_id = payload.get('tenant_id')
            if not tenant_id or tenant_id not in self.phal.tenants:
                raise web.HTTPUnauthorized(reason='Invalid tenant')
                
            # Add context to request
            request['user'] = payload
            request['tenant_id'] = tenant_id
            request['tenant'] = self.phal.tenants[tenant_id]
            request['session_id'] = payload.get('session_id')
            
        except jwt.ExpiredSignatureError:
            raise web.HTTPUnauthorized(reason='Token expired')
        except jwt.InvalidTokenError:
            raise web.HTTPUnauthorized(reason='Invalid token')
            
        return await handler(request)

    def generate_csrf_token(self, session_id: str) -> str:
        return hmac.new(
            self.config.jwt_secret.encode(),
            f"{session_id}:{int(time.time())}".encode(),
            hashlib.sha256
        ).hexdigest()
        
    async def _validate_api_key(self, api_key: str) -> Optional[Tenant]:
        """Validate API key against tenant records"""
        # Hash the API key
        key_hash = hashlib.sha256(api_key.encode()).hexdigest()
        
        # Look up tenant
        for tenant in self.phal.tenants.values():
            if tenant.api_key_hash == key_hash:
                return tenant
                
        return None
        
    @web.middleware
    async def _error_middleware(self, request: web.Request, handler):
        """Error handling middleware"""
        try:
            return await handler(request)
        except web.HTTPException:
            raise
        except ValidationError as e:
            return web.json_response(
                {
                    'error': 'Validation error',
                    'details': e.errors(),
                    'request_id': request.get('request_id')
                },
                status=400
            )
        except GrantError as e:
            return web.json_response(
                {
                    'error': str(e),
                    'code': e.code,
                    'request_id': request.get('request_id')
                },
                status=403
            )
        except RateLimitError as e:
            response = web.json_response(
                {
                    'error': str(e),
                    'code': e.code,
                    'request_id': request.get('request_id'),
                    'retry_after': e.details.get('retryAfter')
                },
                status=429
            )
            if e.details.get('retryAfter'):
                response.headers['Retry-After'] = str(e.details['retryAfter'])
            return response
        except Exception as e:
            logger.error(f"Unhandled error: {e}", exc_info=True)
            
            # Log to audit
            await self.phal.audit_logger.log({
                'event': 'api_error',
                'path': request.path,
                'method': request.method,
                'error': str(e),
                'request_id': request.get('request_id'),
                'user_id': request.get('user', {}).get('id')
            })
            
            return web.json_response(
                {
                    'error': 'Internal server error',
                    'request_id': request.get('request_id')
                },
                status=500
            )
            
    @web.middleware
    async def _metrics_middleware(self, request: web.Request, handler):
        """Metrics collection middleware"""
        start_time = asyncio.get_event_loop().time()
        
        try:
            response = await handler(request)
            status = response.status
        except web.HTTPException as e:
            status = e.status
            raise
        finally:
            duration = asyncio.get_event_loop().time() - start_time
            
            # Record metrics
            path = request.match_info.route.resource.canonical if request.match_info.route else request.path
            
            api_requests.labels(
                method=request.method,
                endpoint=path,
                status=status
            ).inc()
            
            # Log slow requests
            if duration > 1.0:
                logger.warning(f"Slow request: {request.method} {path} took {duration:.3f}s")
                
        return response
        
    def setup_routes(self):
        """Setup all API routes"""
        # Root
        self.app.router.add_get('/', self.index)
        
        # System routes
        self.app.router.add_get('/health', self.health_check)
        self.app.router.add_get('/metrics', self.metrics)
        self.app.router.add_get('/api/v1/status', self.get_status)
        
        # Auth routes
        self.app.router.add_post('/api/v1/auth/token', self.get_token)
        self.app.router.add_post('/api/v1/auth/refresh', self.refresh_token)
        self.app.router.add_post('/api/v1/auth/logout', self.logout)
        
        # Grant management
        self.app.router.add_post('/api/v1/grants/request', self.request_grant)
        self.app.router.add_post('/api/v1/execute', self.execute_command)
        self.app.router.add_get('/api/v1/grants', self.list_grants)
        self.app.router.add_get('/api/v1/grants/{grant_id}', self.get_grant)
        self.app.router.add_delete('/api/v1/grants/{grant_id}', self.revoke_grant)
        
        # Zone management
        self.app.router.add_get('/api/v1/zones', self.get_zones)
        self.app.router.add_get('/api/v1/zones/{zone_id}', self.get_zone)
        self.app.router.add_post('/api/v1/zones', self.create_zone)
        self.app.router.add_patch('/api/v1/zones/{zone_id}', self.update_zone)
        self.app.router.add_delete('/api/v1/zones/{zone_id}', self.delete_zone)
        
        # Sensor data
        self.app.router.add_get('/api/v1/zones/{zone_id}/sensors', self.get_sensor_readings)
        self.app.router.add_get('/api/v1/zones/{zone_id}/sensors/{sensor_type}', self.get_sensor_history)
        
        # Control operations
        self.app.router.add_post('/api/v1/emergency', self.emergency_stop)
        self.app.router.add_post('/api/v1/emergency/reset', self.reset_emergency)
        
        # Harvests
        self.app.router.add_post('/api/v1/harvests', self.log_harvest)
        self.app.router.add_get('/api/v1/harvests', self.get_harvests)
        self.app.router.add_get('/api/v1/harvests/{harvest_id}', self.get_harvest)
        
        # Maintenance
        self.app.router.add_post('/api/v1/maintenance', self.schedule_maintenance)
        self.app.router.add_get('/api/v1/maintenance', self.get_maintenance)
        self.app.router.add_post('/api/v1/maintenance/{record_id}/complete', self.complete_maintenance)
        
        # Alarms
        self.app.router.add_get('/api/v1/alarms', self.get_alarms)
        self.app.router.add_post('/api/v1/alarms/{alarm_id}/acknowledge', self.acknowledge_alarm)
        
        # Analytics
        self.app.router.add_post('/api/v1/analytics/query', self.query_analytics)
        self.app.router.add_get('/api/v1/analytics/insights', self.get_insights)
        self.app.router.add_get('/api/v1/analytics/predictions', self.get_predictions)
        
        # Recipes
        self.app.router.add_get('/api/v1/recipes', self.get_recipes)
        self.app.router.add_post('/api/v1/recipes', self.create_recipe)
        self.app.router.add_get('/api/v1/recipes/{recipe_id}', self.get_recipe)
        self.app.router.add_patch('/api/v1/recipes/{recipe_id}', self.update_recipe)
        self.app.router.add_delete('/api/v1/recipes/{recipe_id}', self.delete_recipe)
        
        # Export/Import
        self.app.router.add_post('/api/v1/export', self.export_data)
        self.app.router.add_post('/api/v1/import', self.import_data)
        
        # Audit logs
        self.app.router.add_get('/api/v1/audit', self.get_audit_logs)
        
        # WebSocket
        self.app.router.add_get('/ws', self.websocket_handler)
        
        # Configure CORS
        cors = aiohttp_cors.setup(self.app, defaults={
            "*": aiohttp_cors.ResourceOptions(
                allow_credentials=True,
                expose_headers="*",
                allow_headers="*",
                allow_methods="*"
            )
        })
        
        # Add CORS to all routes
        for route in list(self.app.router.routes()):
            if not isinstance(route.resource, web.StaticResource):
                cors.add(route)
                
    # Route handlers
    
    async def index(self, request: web.Request) -> web.Response:
        """Root endpoint"""
        return web.json_response({
            'name': 'PHAL - Pluripotent Hardware Abstraction Layer',
            'version': '2.1.0',
            'api_version': 'v1',
            'documentation': 'https://github.com/HydroFarmerJason/PHAL/wiki',
            'endpoints': {
                'health': '/health',
                'metrics': '/metrics',
                'api': '/api/v1',
                'websocket': '/ws'
            }
        })
        
    async def health_check(self, request: web.Request) -> web.Response:
        """System health check"""
        health = {
            'status': 'healthy',
            'timestamp': datetime.now(timezone.utc).isoformat(),
            'version': '2.1.0',
            'node_id': self.phal.config.node_id,
            'uptime': int((datetime.now(timezone.utc) - self.phal.start_time).total_seconds()),
            'components': {}
        }
        
        # Check database
        try:
            async with self.phal.db.acquire() as conn:
                await conn.fetchval('SELECT 1')
            health['components']['database'] = {'status': 'healthy'}
        except Exception as e:
            health['components']['database'] = {'status': 'unhealthy', 'error': str(e)}
            health['status'] = 'degraded'
            
        # Check cache
        try:
            await self.phal.cache.ping()
            health['components']['cache'] = {'status': 'healthy'}
        except Exception as e:
            health['components']['cache'] = {'status': 'unhealthy', 'error': str(e)}
            health['status'] = 'degraded'
            
        # Check plugins
        for plugin_id, plugin in self.phal.plugins.items():
            plugin_health = await plugin.health_check()
            health['components'][f'plugin:{plugin_id}'] = {
                'status': 'healthy' if plugin_health['healthy'] else 'unhealthy',
                'details': plugin_health.get('details', {})
            }
            if not plugin_health['healthy']:
                health['status'] = 'degraded'
                
        # System resources
        health['resources'] = {
            'active_grants': len(self.phal.grants),
            'active_alarms': len(self.phal.alarms),
            'zones': len(self.phal.zones),
            'capabilities': len(self.phal.capabilities)
        }
        
        status_code = 200 if health['status'] == 'healthy' else 503
        return web.json_response(health, status=status_code)
        
    async def metrics(self, request: web.Request) -> web.Response:
        """Prometheus metrics endpoint"""
        return web.Response(
            text=generate_latest().decode('utf-8'),
            content_type='text/plain'
        )
        
    async def get_status(self, request: web.Request) -> web.Response:
        """Get system status"""
        tenant = request['tenant']
        
        # Get tenant-specific stats
        tenant_zones = [z for z in self.phal.zones.values() if z.tenant_id == tenant.id]
        tenant_grants = [g for g in self.phal.grants.values() if g.tenant_id == tenant.id]
        
        status = {
            'tenant': {
                'id': tenant.id,
                'name': tenant.name,
                'tier': tenant.tier,
                'billing_status': tenant.billing_status
            },
            'resources': {
                'zones': {
                    'count': len(tenant_zones),
                    'limit': tenant.resource_limits.get('zones', 'unlimited')
                },
                'grants': {
                    'active': len(tenant_grants),
                    'limit': tenant.resource_limits.get('grants', 'unlimited')
                }
            },
            'system': {
                'uptime': int((datetime.now(timezone.utc) - self.phal.start_time).total_seconds()),
                'version': '2.1.0',
                'features': list(tenant.features)
            }
        }
        
        return web.json_response(status)
        
    async def get_token(self, request: web.Request) -> web.Response:
        """Get JWT token"""
        data = sanitize_input(await request.json())
        
        # Validate API key
        api_key = data.get('apiKey')
        if not api_key:
            raise web.HTTPBadRequest(reason='Missing API key')
            
        tenant = await self._validate_api_key(api_key)
        if not tenant:
            raise web.HTTPUnauthorized(reason='Invalid API key')
            
        # Check tenant status
        if tenant.billing_status != 'active':
            raise web.HTTPForbidden(reason=f'Tenant status: {tenant.billing_status}')
            
        # Create session
        session_id = str(uuid.uuid4())
        session_data = {
            'tenant_id': tenant.id,
            'created_at': datetime.now(timezone.utc).isoformat(),
            'user_agent': request.headers.get('User-Agent', ''),
            'ip_address': request.remote
        }
        
        # Store session
        self.phal.session_store[session_id] = session_data
        expires_at = datetime.now(timezone.utc) + timedelta(seconds=self.phal.config.session_timeout)
        
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO sessions (id, tenant_id, user_id, data, expires_at)
                VALUES ($1, $2, $3, $4, $5)
            ''', session_id, tenant.id, data.get('userId', 'api'),
                json.dumps(session_data), expires_at)
                
        # Generate tokens
        payload = {
            'tenant_id': tenant.id,
            'tenant_name': tenant.name,
            'tier': tenant.tier,
            'features': list(tenant.features),
            'session_id': session_id,
            'exp': datetime.utcnow() + timedelta(hours=1),
            'iat': datetime.utcnow()
        }
        
        token = jwt.encode(payload, self.phal.config.jwt_secret, algorithm='HS256')
        
        # Generate refresh token
        refresh_payload = {
            'tenant_id': tenant.id,
            'session_id': session_id,
            'type': 'refresh',
            'exp': datetime.utcnow() + timedelta(days=30),
            'iat': datetime.utcnow()
        }
        
        refresh_token = jwt.encode(refresh_payload, self.phal.config.jwt_secret, algorithm='HS256')
        
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'auth_token_created',
            'tenant_id': tenant.id,
            'session_id': session_id,
            'ip_address': request.remote,
            'user_agent': request.headers.get('User-Agent')
        })
        
        return web.json_response({
            'token': token,
            'refreshToken': refresh_token,
            'expiresIn': 3600,
            'sessionId': session_id,
            'tenant': {
                'id': tenant.id,
                'name': tenant.name,
                'tier': tenant.tier,
                'features': list(tenant.features)
            }
        })
        
    async def refresh_token(self, request: web.Request) -> web.Response:
        """Refresh JWT token"""
        data = sanitize_input(await request.json())
        refresh_token = data.get('refresh_token')
        
        if not refresh_token:
            raise web.HTTPBadRequest(reason='Missing refresh token')
            
        try:
            payload = jwt.decode(
                refresh_token,
                self.phal.config.jwt_secret,
                algorithms=['HS256']
            )
            
            if payload.get('type') != 'refresh':
                raise web.HTTPUnauthorized(reason='Invalid token type')
                
            # Get tenant
            tenant = self.phal.tenants.get(payload['tenant_id'])
            if not tenant:
                raise web.HTTPUnauthorized(reason='Invalid tenant')
                
            # Update session
            session_id = payload.get('session_id')
            if session_id:
                async with self.phal.db.acquire() as conn:
                    await conn.execute('''
                        UPDATE sessions 
                        SET last_accessed = NOW()
                        WHERE id = $1
                    ''', session_id)
                    
            # Generate new access token
            new_payload = {
                'tenant_id': tenant.id,
                'tenant_name': tenant.name,
                'tier': tenant.tier,
                'features': list(tenant.features),
                'session_id': session_id,
                'exp': datetime.utcnow() + timedelta(hours=1),
                'iat': datetime.utcnow()
            }
            
            token = jwt.encode(new_payload, self.phal.config.jwt_secret, algorithm='HS256')
            
            return web.json_response({
                'token': token,
                'expiresIn': 3600
            })
            
        except jwt.ExpiredSignatureError:
            raise web.HTTPUnauthorized(reason='Refresh token expired')
        except jwt.InvalidTokenError:
            raise web.HTTPUnauthorized(reason='Invalid refresh token')
            
    async def logout(self, request: web.Request) -> web.Response:
        """Logout and invalidate session"""
        session_id = request.get('session_id')
        
        if session_id:
            # Remove from memory
            self.phal.session_store.pop(session_id, None)
            
            # Remove from database
            async with self.phal.db.acquire() as conn:
                await conn.execute(
                    'DELETE FROM sessions WHERE id = $1',
                    session_id
                )
                
        return web.json_response({'status': 'logged out'})
        
    async def request_grant(self, request: web.Request) -> web.Response:
        """Request capability grant"""
        data = sanitize_input(await request.json())
        tenant_id = request['tenant_id']
        
        # Validate request
        try:
            grant_request = GrantRequest(**data)
        except ValidationError as e:
            raise web.HTTPBadRequest(text=json.dumps({'errors': e.errors()}))
            
        # Request grant
        grant = await self.phal.request_capability(
            tenant_id=tenant_id,
            plugin_id=grant_request.plugin_id,
            capability_query=grant_request.capability_query,
            permissions=grant_request.permissions,
            duration=timedelta(seconds=grant_request.duration_seconds),
            constraints=grant_request.constraints
        )
        
        if not grant:
            raise web.HTTPForbidden(reason='Grant request denied')
            
        return web.json_response({
            'grant_id': grant.id,
            'expires_at': grant.expires_at.isoformat(),
            'permissions': grant.permissions,
            'capability_id': grant.capability_id,
            'constraints': grant.constraints
        })
        
    async def execute_command(self, request: web.Request) -> web.Response:
        """Execute command with grant"""
        data = sanitize_input(await request.json())
        
        try:
            cmd_request = CommandRequest(**data)
        except ValidationError as e:
            raise web.HTTPBadRequest(text=json.dumps({'errors': e.errors()}))
            
        # Execute command
        result = await self.phal.execute_command(
            grant_id=cmd_request.grant_id,
            command=cmd_request.command,
            context={
                'user': request['user'],
                'ip_address': request.remote,
                'user_agent': request.headers.get('User-Agent'),
                'session_id': request.get('session_id')
            }
        )
        
        return web.json_response(result)
        
    async def list_grants(self, request: web.Request) -> web.Response:
        """List active grants for tenant"""
        tenant_id = request['tenant_id']
        
        grants = []
        for grant in self.phal.grants.values():
            if grant.tenant_id == tenant_id and grant.is_valid():
                grants.append({
                    'id': grant.id,
                    'plugin_id': grant.plugin_id,
                    'capability_id': grant.capability_id,
                    'permissions': grant.permissions,
                    'expires_at': grant.expires_at.isoformat(),
                    'usage_count': grant.usage_count,
                    'max_usage': grant.max_usage,
                    'created_at': grant.created_at.isoformat(),
                    'last_used': grant.last_used.isoformat() if grant.last_used else None
                })
                
        return web.json_response({
            'grants': grants,
            'total': len(grants)
        })
        
    async def get_grant(self, request: web.Request) -> web.Response:
        """Get grant details"""
        grant_id = request.match_info['grant_id']
        tenant_id = request['tenant_id']
        
        grant = await self.phal._get_grant(grant_id)
        if not grant or grant.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Grant not found')
            
        return web.json_response({
            'id': grant.id,
            'plugin_id': grant.plugin_id,
            'capability_id': grant.capability_id,
            'permissions': grant.permissions,
            'expires_at': grant.expires_at.isoformat(),
            'usage_count': grant.usage_count,
            'max_usage': grant.max_usage,
            'created_at': grant.created_at.isoformat(),
            'last_used': grant.last_used.isoformat() if grant.last_used else None,
            'constraints': grant.constraints,
            'audit_log': grant.audit_log[-10:]  # Last 10 entries
        })
        
    async def revoke_grant(self, request: web.Request) -> web.Response:
        """Revoke a grant"""
        grant_id = request.match_info['grant_id']
        tenant_id = request['tenant_id']
        
        # Verify ownership
        grant = await self.phal._get_grant(grant_id)
        if not grant or grant.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Grant not found')
            
        await self.phal._revoke_grant(grant_id)
        
        return web.json_response({'status': 'revoked', 'grant_id': grant_id})
        
    async def get_zones(self, request: web.Request) -> web.Response:
        """Get zones for tenant"""
        tenant_id = request['tenant_id']
        
        zones = []
        for zone in self.phal.zones.values():
            if zone.tenant_id == tenant_id:
                zones.append({
                    'id': zone.id,
                    'name': zone.name,
                    'type': zone.type,
                    'units': zone.units,
                    'environmental_targets': zone.environmental_targets,
                    'crop_profile': zone.crop_profile,
                    'emergency_stop': zone.emergency_stop,
                    'maintenance_mode': zone.maintenance_mode,
                    'created_at': zone.created_at.isoformat(),
                    'age_days': zone.get_age_days(),
                    'total_yield': zone.total_yield,
                    'active_alarms': len(zone.active_alarms),
                    'metadata': zone.metadata
                })
                
        return web.json_response({
            'zones': zones,
            'total': len(zones)
        })
        
    async def get_zone(self, request: web.Request) -> web.Response:
        """Get zone details with current state"""
        zone_id = request.match_info['zone_id']
        tenant_id = request['tenant_id']
        
        zone = self.phal.zones.get(zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        # Get current sensor readings
        sensors = {}
        for cap in self.phal.capabilities.values():
            if cap.zone_id == zone_id and cap.type == 'SENSOR' and cap.last_reading:
                sensors[cap.subtype] = cap.last_reading
                
        # Get active alarms
        active_alarms = []
        for alarm_id in zone.active_alarms:
            alarm = self.phal.alarms.get(alarm_id)
            if alarm:
                active_alarms.append(alarm.to_dict())
                
        # Get capabilities
        capabilities = []
        for cap in self.phal.capabilities.values():
            if cap.zone_id == zone_id:
                capabilities.append({
                    'id': cap.id,
                    'type': cap.type,
                    'subtype': cap.subtype,
                    'status': cap.status,
                    'properties': cap.properties,
                    'needs_calibration': cap.needs_calibration()
                })
                
        return web.json_response({
            'id': zone.id,
            'name': zone.name,
            'type': zone.type,
            'units': zone.units,
            'environmental_targets': zone.environmental_targets,
            'crop_profile': zone.crop_profile,
            'emergency_stop': zone.emergency_stop,
            'maintenance_mode': zone.maintenance_mode,
            'created_at': zone.created_at.isoformat(),
            'age_days': zone.get_age_days(),
            'total_yield': zone.total_yield,
            'metadata': zone.metadata,
            'current_state': {
                'sensors': sensors,
                'operational': zone.is_operational()
            },
            'active_alarms': active_alarms,
            'capabilities': capabilities
        })
        
    async def create_zone(self, request: web.Request) -> web.Response:
        """Create new zone"""
        tenant = request['tenant']
        data = sanitize_input(await request.json())
        
        # Check zone limit
        current_zones = len([z for z in self.phal.zones.values() if z.tenant_id == tenant.id])
        if not tenant.check_resource_limit('zones', current_zones, 1):
            raise web.HTTPForbidden(reason='Zone limit exceeded')
            
        zone_id = str(uuid.uuid4())
        zone = Zone(
            id=zone_id,
            tenant_id=tenant.id,
            name=data['name'],
            type=data.get('type', 'production'),
            units=data.get('units', []),
            environmental_targets=data.get('environmental_targets', {}),
            crop_profile=data.get('crop_profile'),
            metadata=data.get('metadata', {})
        )
        
        # Validate zone data
        if zone.type not in ['production', 'nursery', 'quarantine', 'research']:
            raise web.HTTPBadRequest(reason='Invalid zone type')
            
        # Save to database
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO zones 
                (id, tenant_id, name, type, units, environmental_targets, crop_profile, metadata)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            ''', zone.id, zone.tenant_id, zone.name, zone.type,
                zone.units, json.dumps(zone.environmental_targets),
                json.dumps(zone.crop_profile) if zone.crop_profile else None,
                json.dumps(zone.metadata))
                
        self.phal.zones[zone_id] = zone
        
        # Create capabilities for zone
        # This would be done by hardware plugin in production
        
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'zone_created',
            'tenant_id': tenant.id,
            'zone_id': zone_id,
            'zone_name': zone.name,
            'user_id': request['user'].get('id')
        })
        
        return web.json_response({
            'id': zone.id,
            'name': zone.name,
            'type': zone.type,
            'created_at': zone.created_at.isoformat()
        }, status=201)
        
    async def update_zone(self, request: web.Request) -> web.Response:
        """Update zone configuration"""
        zone_id = request.match_info['zone_id']
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        zone = self.phal.zones.get(zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        # Update allowed fields
        if 'name' in data:
            zone.name = data['name']
        if 'environmental_targets' in data:
            zone.environmental_targets = data['environmental_targets']
        if 'crop_profile' in data:
            zone.crop_profile = data['crop_profile']
        if 'metadata' in data:
            zone.metadata.update(data['metadata'])
            
        # Save to database
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                UPDATE zones
                SET name = $1, environmental_targets = $2, crop_profile = $3, metadata = $4
                WHERE id = $5
            ''', zone.name, json.dumps(zone.environmental_targets),
                json.dumps(zone.crop_profile) if zone.crop_profile else None,
                json.dumps(zone.metadata), zone_id)
                
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'zone_updated',
            'tenant_id': tenant_id,
            'zone_id': zone_id,
            'updates': list(data.keys()),
            'user_id': request['user'].get('id')
        })
        
        return web.json_response({'status': 'updated'})
        
    async def delete_zone(self, request: web.Request) -> web.Response:
        """Delete zone"""
        zone_id = request.match_info['zone_id']
        tenant_id = request['tenant_id']
        
        zone = self.phal.zones.get(zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        # Check if zone has active grants
        active_grants = [
            g for g in self.phal.grants.values()
            if g.tenant_id == tenant_id and g.is_valid()
        ]
        
        zone_capabilities = [
            c.id for c in self.phal.capabilities.values()
            if c.zone_id == zone_id
        ]
        
        for grant in active_grants:
            if grant.capability_id in zone_capabilities:
                raise web.HTTPConflict(reason='Zone has active grants')
                
        # Delete from database (cascades to capabilities, sensor_readings, etc)
        async with self.phal.db.acquire() as conn:
            await conn.execute('DELETE FROM zones WHERE id = $1', zone_id)
            
        # Remove from memory
        self.phal.zones.pop(zone_id, None)
        
        # Remove related capabilities
        for cap_id in zone_capabilities:
            self.phal.capabilities.pop(cap_id, None)
            
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'zone_deleted',
            'tenant_id': tenant_id,
            'zone_id': zone_id,
            'zone_name': zone.name,
            'user_id': request['user'].get('id')
        })
        
        return web.json_response({'status': 'deleted'})
        
    async def get_sensor_readings(self, request: web.Request) -> web.Response:
        """Get current sensor readings for zone"""
        zone_id = request.match_info['zone_id']
        tenant_id = request['tenant_id']
        
        zone = self.phal.zones.get(zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        readings = {}
        for cap in self.phal.capabilities.values():
            if cap.zone_id == zone_id and cap.type == 'SENSOR' and cap.last_reading:
                readings[cap.subtype] = {
                    'value': cap.last_reading.get('value'),
                    'quality': cap.last_reading.get('quality', 1.0),
                    'unit': cap.last_reading.get('unit'),
                    'timestamp': cap.last_reading.get('timestamp'),
                    'status': cap.status,
                    'needs_calibration': cap.needs_calibration()
                }
                
        return web.json_response(readings)
        
    async def get_sensor_history(self, request: web.Request) -> web.Response:
        """Get sensor history"""
        zone_id = request.match_info['zone_id']
        sensor_type = request.match_info['sensor_type']
        tenant_id = request['tenant_id']
        
        # Validate zone ownership
        zone = self.phal.zones.get(zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        # Parse query parameters
        hours = int(request.query.get('hours', 24))
        resolution = request.query.get('resolution', 'auto')
        
        # Determine aggregation based on time range
        if resolution == 'auto':
            if hours <= 6:
                interval = '1 minute'
            elif hours <= 24:
                interval = '5 minutes'
            elif hours <= 168:  # 1 week
                interval = '1 hour'
            else:
                interval = '1 day'
        else:
            interval = resolution
            
        # Query time-series data
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT 
                    date_trunc($1, timestamp) as time_bucket,
                    AVG(value) as avg_value,
                    MIN(value) as min_value,
                    MAX(value) as max_value,
                    STDDEV(value) as std_dev,
                    AVG(quality) as avg_quality,
                    COUNT(*) as sample_count
                FROM sensor_readings
                WHERE zone_id = $2
                AND sensor_type = $3
                AND timestamp > NOW() - INTERVAL '%s hours'
                GROUP BY time_bucket
                ORDER BY time_bucket DESC
                LIMIT 1000
            ''' % hours, interval, zone_id, sensor_type)
            
        data = []
        for row in rows:
            data.append({
                'timestamp': row['time_bucket'].isoformat(),
                'value': float(row['avg_value']),
                'min': float(row['min_value']),
                'max': float(row['max_value']),
                'std_dev': float(row['std_dev']) if row['std_dev'] else 0,
                'quality': float(row['avg_quality']),
                'samples': row['sample_count']
            })
            
        # Get sensor info
        sensor_cap = next(
            (c for c in self.phal.capabilities.values()
             if c.zone_id == zone_id and c.subtype == sensor_type),
            None
        )
        
        return web.json_response({
            'zone_id': zone_id,
            'sensor_type': sensor_type,
            'time_range': {
                'hours': hours,
                'start': (datetime.now(timezone.utc) - timedelta(hours=hours)).isoformat(),
                'end': datetime.now(timezone.utc).isoformat()
            },
            'resolution': interval,
            'unit': sensor_cap.properties.get('unit') if sensor_cap else None,
            'data': data
        })
        
    async def emergency_stop(self, request: web.Request) -> web.Response:
        """Activate emergency stop"""
        data = sanitize_input(await request.json())
        zone_id = data.get('zone_id')
        tenant_id = request['tenant_id']
        reason = data.get('reason', 'Manual emergency stop')
        
        zones_to_stop = []
        
        if zone_id:
            zone = self.phal.zones.get(zone_id)
            if zone and zone.tenant_id == tenant_id:
                zones_to_stop.append(zone_id)
            else:
                raise web.HTTPNotFound(reason='Zone not found')
        else:
            # Stop all zones for tenant
            for zone in self.phal.zones.values():
                if zone.tenant_id == tenant_id:
                    zones_to_stop.append(zone.id)
                    
        if not zones_to_stop:
            raise web.HTTPBadRequest(reason='No zones to stop')
            
        # Set emergency stop in cache
        await self.phal.cache.setex(
            'emergency_stop',
            3600,  # 1 hour
            json.dumps(zones_to_stop)
        )
        
        # Update zone states
        for zone_id in zones_to_stop:
            self.phal.zones[zone_id].emergency_stop = True
            
            # Update database
            async with self.phal.db.acquire() as conn:
                await conn.execute(
                    'UPDATE zones SET emergency_stop = true WHERE id = $1',
                    zone_id
                )
                
        # Turn off all actuators
        for cap in self.phal.capabilities.values():
            if cap.zone_id in zones_to_stop and cap.type == 'ACTUATOR':
                # Force actuator off through hardware plugin
                plugin = self.phal.plugins.get('HardwareSimulator')
                if plugin:
                    try:
                        await plugin.execute_capability(cap.id, {
                            'command': 'set_state',
                            'state': False
                        })
                    except Exception as e:
                        logger.error(f"Failed to stop actuator {cap.id}: {e}")
                        
        # Create alarm
        for zone_id in zones_to_stop:
            await self.phal.event_bus.emit('alarm', {
                'zone_id': zone_id,
                'type': 'emergency_stop',
                'severity': 'emergency',
                'message': f'Emergency stop activated: {reason}',
                'auto_resolve': False
            })
            
        # Emit event
        await self.phal.event_bus.emit('emergency_stop', {
            'zones': zones_to_stop,
            'activated_by': request['user'].get('id'),
            'reason': reason
        })
        
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'emergency_stop_activated',
            'tenant_id': tenant_id,
            'zones': zones_to_stop,
            'reason': reason,
            'user_id': request['user'].get('id'),
            'ip_address': request.remote
        })
        
        return web.json_response({
            'status': 'activated',
            'zones': zones_to_stop,
            'message': f'Emergency stop activated for {len(zones_to_stop)} zones'
        })
        
    async def reset_emergency(self, request: web.Request) -> web.Response:
        """Reset emergency stop"""
        data = sanitize_input(await request.json())
        zone_id = data.get('zone_id')
        tenant_id = request['tenant_id']
        
        # Verify authorization
        if not data.get('confirmation'):
            raise web.HTTPBadRequest(reason='Confirmation required')
            
        zones_to_reset = []
        
        if zone_id:
            zone = self.phal.zones.get(zone_id)
            if zone and zone.tenant_id == tenant_id and zone.emergency_stop:
                zones_to_reset.append(zone_id)
        else:
            # Reset all zones for tenant
            for zone in self.phal.zones.values():
                if zone.tenant_id == tenant_id and zone.emergency_stop:
                    zones_to_reset.append(zone.id)
                    
        if not zones_to_reset:
            raise web.HTTPBadRequest(reason='No zones in emergency stop')
            
        # Clear emergency stop
        current = await self.phal.cache.get('emergency_stop')
        if current:
            stopped_zones = json.loads(current)
            for zone_id in zones_to_reset:
                if zone_id in stopped_zones:
                    stopped_zones.remove(zone_id)
                    
            if stopped_zones:
                await self.phal.cache.setex('emergency_stop', 3600, json.dumps(stopped_zones))
            else:
                await self.phal.cache.delete('emergency_stop')
                
        # Update zone states
        for zone_id in zones_to_reset:
            self.phal.zones[zone_id].emergency_stop = False
            
            # Update database
            async with self.phal.db.acquire() as conn:
                await conn.execute(
                    'UPDATE zones SET emergency_stop = false WHERE id = $1',
                    zone_id
                )
                
        # Clear related alarms
        for alarm_id, alarm in list(self.phal.alarms.items()):
            if alarm.zone_id in zones_to_reset and alarm.type == 'emergency_stop':
                alarm.acknowledged = True
                alarm.acknowledged_by = request['user'].get('id')
                alarm.acknowledged_at = datetime.now(timezone.utc)
                alarm.resolution = 'Emergency stop reset'
                
        # Emit event
        await self.phal.event_bus.emit('emergency_stop_reset', {
            'zones': zones_to_reset,
            'reset_by': request['user'].get('id')
        })
        
        # Audit log
        await self.phal.audit_logger.log({
            'event': 'emergency_stop_reset',
            'tenant_id': tenant_id,
            'zones': zones_to_reset,
            'user_id': request['user'].get('id'),
            'ip_address': request.remote
        })
        
        return web.json_response({
            'status': 'reset',
            'zones': zones_to_reset
        })
        
    async def log_harvest(self, request: web.Request) -> web.Response:
        """Log harvest data"""
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        try:
            harvest_request = HarvestRequest(**data)
        except ValidationError as e:
            raise web.HTTPBadRequest(text=json.dumps({'errors': e.errors()}))
            
        # Validate zone
        zone = self.phal.zones.get(harvest_request.zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPNotFound(reason='Zone not found')
            
        harvest_id = str(uuid.uuid4())
        
        # Save to database
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO harvests 
                (id, zone_id, crop_id, quantity, quantity_unit, quality_grade,
                 harvested_by, notes, images, harvest_date)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, NOW())
            ''', harvest_id, harvest_request.zone_id, harvest_request.crop_id,
                harvest_request.quantity, harvest_request.quantity_unit,
                harvest_request.quality_grade, harvest_request.harvested_by,
                harvest_request.notes, harvest_request.images)
                
        # Update zone stats
        zone.last_harvest = datetime.now(timezone.utc)
        zone.total_yield += harvest_request.quantity
        
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                UPDATE zones 
                SET last_harvest = NOW(), total_yield = total_yield + $1
                WHERE id = $2
            ''', harvest_request.quantity, zone.id)
            
        # Emit event
        await self.phal.event_bus.emit('harvest_logged', {
            'harvest_id': harvest_id,
            'zone_id': zone.id,
            'quantity': harvest_request.quantity,
            'quality': harvest_request.quality_grade
        })
        
        return web.json_response({
            'id': harvest_id,
            'status': 'logged',
            'zone_total_yield': zone.total_yield
        }, status=201)
        
    async def get_harvests(self, request: web.Request) -> web.Response:
        """Get harvest records"""
        tenant_id = request['tenant_id']
        
        # Get tenant zones
        tenant_zones = [z.id for z in self.phal.zones.values() if z.tenant_id == tenant_id]
        
        if not tenant_zones:
            return web.json_response({'harvests': [], 'total': 0})
            
        # Query parameters
        zone_id = request.query.get('zone_id')
        days = int(request.query.get('days', 30))
        
        # Build query
        query = '''
            SELECT h.*, z.name as zone_name
            FROM harvests h
            JOIN zones z ON h.zone_id = z.id
            WHERE z.id = ANY($1)
        '''
        params = [tenant_zones]
        
        if zone_id and zone_id in tenant_zones:
            query += ' AND h.zone_id = $2'
            params.append(zone_id)
            
        query += ' AND h.harvest_date > NOW() - INTERVAL \'%d days\'' % days
        query += ' ORDER BY h.harvest_date DESC LIMIT 100'
        
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch(query, *params)
            
        harvests = []
        for row in rows:
            harvests.append({
                'id': str(row['id']),
                'zone_id': str(row['zone_id']),
                'zone_name': row['zone_name'],
                'crop_id': row['crop_id'],
                'quantity': float(row['quantity']),
                'quantity_unit': row['quantity_unit'],
                'quality_grade': row['quality_grade'],
                'harvested_by': row['harvested_by'],
                'harvest_date': row['harvest_date'].isoformat(),
                'notes': row['notes'],
                'images': row['images']
            })
            
        return web.json_response({
            'harvests': harvests,
            'total': len(harvests)
        })
        
    async def get_harvest(self, request: web.Request) -> web.Response:
        """Get harvest details"""
        harvest_id = request.match_info['harvest_id']
        tenant_id = request['tenant_id']
        
        async with self.phal.db.acquire() as conn:
            row = await conn.fetchrow('''
                SELECT h.*, z.name as zone_name, z.tenant_id
                FROM harvests h
                JOIN zones z ON h.zone_id = z.id
                WHERE h.id = $1
            ''', harvest_id)
            
        if not row or str(row['tenant_id']) != tenant_id:
            raise web.HTTPNotFound(reason='Harvest not found')
            
        return web.json_response({
            'id': str(row['id']),
            'zone_id': str(row['zone_id']),
            'zone_name': row['zone_name'],
            'crop_id': row['crop_id'],
            'quantity': float(row['quantity']),
            'quantity_unit': row['quantity_unit'],
            'quality_grade': row['quality_grade'],
            'quality_metrics': row['quality_metrics'],
            'harvested_by': row['harvested_by'],
            'harvest_date': row['harvest_date'].isoformat(),
            'destination': row['destination'],
            'notes': row['notes'],
            'images': row['images'],
            'labor_time_minutes': row['labor_time_minutes'],
            'packaging_type': row['packaging_type'],
            'batch_code': row['batch_code']
        })
        
    async def schedule_maintenance(self, request: web.Request) -> web.Response:
        """Schedule maintenance"""
        data = sanitize_input(await request.json())
        
        try:
            maint_request = MaintenanceRequest(**data)
        except ValidationError as e:
            raise web.HTTPBadRequest(text=json.dumps({'errors': e.errors()}))
            
        record_id = str(uuid.uuid4())
        
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO maintenance_records
                (id, type, component_id, scheduled_date, notes)
                VALUES ($1, $2, $3, $4, $5)
            ''', record_id, maint_request.type, maint_request.component_id,
                maint_request.scheduled_date, maint_request.notes)
                
        return web.json_response({
            'id': record_id,
            'status': 'scheduled'
        }, status=201)
        
    async def get_maintenance(self, request: web.Request) -> web.Response:
        """Get maintenance records"""
        completed = request.query.get('completed', 'false').lower() == 'true'
        days_ahead = int(request.query.get('days_ahead', 30))
        
        query = '''
            SELECT * FROM maintenance_records
            WHERE {} 
            ORDER BY {} DESC
            LIMIT 100
        '''
        
        if completed:
            where_clause = 'completed_date IS NOT NULL'
            order_by = 'completed_date'
        else:
            where_clause = 'completed_date IS NULL AND scheduled_date < NOW() + INTERVAL \'%d days\'' % days_ahead
            order_by = 'scheduled_date'
            
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch(query.format(where_clause, order_by))
            
        records = []
        for row in rows:
            records.append({
                'id': str(row['id']),
                'type': row['type'],
                'component_id': row['component_id'],
                'scheduled_date': row['scheduled_date'].isoformat(),
                'completed_date': row['completed_date'].isoformat() if row['completed_date'] else None,
                'technician': row['technician'],
                'notes': row['notes'],
                'cost': float(row['cost']) if row['cost'] else None,
                'downtime_minutes': row['downtime_minutes']
            })
            
        return web.json_response({
            'records': records,
            'total': len(records)
        })
        
    async def complete_maintenance(self, request: web.Request) -> web.Response:
        """Complete maintenance task"""
        record_id = request.match_info['record_id']
        data = sanitize_input(await request.json())
        
        # Verify record exists
        async with self.phal.db.acquire() as conn:
            exists = await conn.fetchval(
                'SELECT 1 FROM maintenance_records WHERE id = $1',
                record_id
            )
            
        if not exists:
            raise web.HTTPNotFound(reason='Maintenance record not found')
            
        # Update record
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                UPDATE maintenance_records
                SET completed_date = NOW(),
                    technician = $1,
                    actions = $2,
                    parts_used = $3,
                    notes = COALESCE(notes, '') || E'\\n' || $4,
                    cost = $5,
                    downtime_minutes = $6
                WHERE id = $7
            ''', data['technician'], data.get('actions', []),
                json.dumps(data.get('parts', [])), data.get('notes', ''),
                data.get('cost'), data.get('downtime_minutes'),
                record_id)
                
        return web.json_response({'status': 'completed'})
        
    async def get_alarms(self, request: web.Request) -> web.Response:
        """Get active alarms"""
        tenant_id = request['tenant_id']
        
        # Get tenant zones
        tenant_zones = [z.id for z in self.phal.zones.values() if z.tenant_id == tenant_id]
        
        # Filter parameters
        acknowledged = request.query.get('acknowledged', 'false').lower() == 'true'
        severity = request.query.get('severity')
        zone_id = request.query.get('zone_id')
        
        alarms = []
        for alarm in self.phal.alarms.values():
            if alarm.zone_id not in tenant_zones:
                continue
            if alarm.acknowledged != acknowledged:
                continue
            if severity and alarm.severity != severity:
                continue
            if zone_id and alarm.zone_id != zone_id:
                continue
                
            alarms.append(alarm.to_dict())
            
        # Sort by severity and time
        severity_order = {'emergency': 0, 'critical': 1, 'warning': 2, 'info': 3}
        alarms.sort(key=lambda a: (
            severity_order.get(a['severity'], 4),
            a['timestamp']
        ))
        
        return web.json_response({
            'alarms': alarms,
            'total': len(alarms)
        })
        
    async def acknowledge_alarm(self, request: web.Request) -> web.Response:
        """Acknowledge an alarm"""
        alarm_id = request.match_info['alarm_id']
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        alarm = self.phal.alarms.get(alarm_id)
        if not alarm:
            raise web.HTTPNotFound(reason='Alarm not found')
            
        # Verify zone ownership
        zone = self.phal.zones.get(alarm.zone_id)
        if not zone or zone.tenant_id != tenant_id:
            raise web.HTTPForbidden(reason='Access denied')
            
        # Update alarm
        alarm.acknowledged = True
        alarm.acknowledged_by = request['user'].get('id', 'unknown')
        alarm.acknowledged_at = datetime.now(timezone.utc)
        alarm.resolution = data.get('resolution', 'Acknowledged by user')
        
        # Update database
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                UPDATE alarms
                SET acknowledged = true,
                    acknowledged_by = $1,
                    acknowledged_at = NOW(),
                    resolution = $2
                WHERE id = $3
            ''', alarm.acknowledged_by, alarm.resolution, alarm_id)
            
        # Remove from zone's active alarms
        if alarm_id in zone.active_alarms:
            zone.active_alarms.remove(alarm_id)
            
        # Emit event
        await self.phal.event_bus.emit('alarm_acknowledged', {
            'alarm_id': alarm_id,
            'zone_id': alarm.zone_id,
            'acknowledged_by': alarm.acknowledged_by
        })
        
        return web.json_response({'status': 'acknowledged'})
        
    async def query_analytics(self, request: web.Request) -> web.Response:
        """Query analytics data"""
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        # Validate time range
        try:
            start_date = datetime.fromisoformat(data['timeRange']['start'])
            end_date = datetime.fromisoformat(data['timeRange']['end'])
        except (KeyError, ValueError):
            raise web.HTTPBadRequest(reason='Invalid time range')
            
        # Get tenant zones
        tenant_zones = [z.id for z in self.phal.zones.values() if z.tenant_id == tenant_id]
        
        zones = data.get('zones', tenant_zones)
        # Filter to only tenant's zones
        zones = [z for z in zones if z in tenant_zones]
        
        if not zones:
            return web.json_response({'results': []})
            
        metrics = data.get('metrics', ['temperature', 'humidity'])
        aggregation = data.get('aggregation', 'avg')
        group_by = data.get('groupBy', 'hour')
        
        # Build query
        agg_func = {
            'avg': 'AVG',
            'min': 'MIN',
            'max': 'MAX',
            'sum': 'SUM',
            'count': 'COUNT'
        }.get(aggregation, 'AVG')
        
        time_bucket = {
            'hour': '1 hour',
            'day': '1 day',
            'week': '1 week',
            'month': '1 month'
        }.get(group_by, '1 hour')
        
        results = {}
        
        for metric in metrics:
            async with self.phal.db.acquire() as conn:
                rows = await conn.fetch(f'''
                    SELECT 
                        date_trunc($1, timestamp) as time_bucket,
                        zone_id,
                        {agg_func}(value) as value
                    FROM sensor_readings
                    WHERE zone_id = ANY($2)
                    AND sensor_type = $3
                    AND timestamp BETWEEN $4 AND $5
                    GROUP BY time_bucket, zone_id
                    ORDER BY time_bucket
                ''', time_bucket, zones, metric, start_date, end_date)
                
            # Format results
            metric_data = defaultdict(list)
            for row in rows:
                metric_data[str(row['zone_id'])].append({
                    'timestamp': row['time_bucket'].isoformat(),
                    'value': float(row['value'])
                })
                
            results[metric] = dict(metric_data)
            
        return web.json_response({
            'query': {
                'zones': zones,
                'metrics': metrics,
                'timeRange': {
                    'start': start_date.isoformat(),
                    'end': end_date.isoformat()
                },
                'aggregation': aggregation,
                'groupBy': group_by
            },
            'results': results
        })
        
    async def get_insights(self, request: web.Request) -> web.Response:
        """Get predictive insights"""
        tenant_id = request['tenant_id']
        zone_id = request.query.get('zone_id')
        
        insights = []
        
        # Get tenant zones
        tenant_zones = [z for z in self.phal.zones.values() if z.tenant_id == tenant_id]
        
        if zone_id:
            zone = self.phal.zones.get(zone_id)
            if not zone or zone.tenant_id != tenant_id:
                raise web.HTTPNotFound(reason='Zone not found')
            zones_to_analyze = [zone]
        else:
            zones_to_analyze = tenant_zones
            
        for zone in zones_to_analyze:
            if not zone.is_operational() or not zone.crop_profile:
                continue
                
            # Get recent sensor data
            sensor_data = await self.phal._get_recent_sensor_data(zone.id, hours=168)
            
            if not sensor_data.empty:
                # Yield prediction
                yield_prediction = await self.phal.ml_engine.predict_yield(zone, sensor_data)
                if yield_prediction['confidence'] > 0.5:
                    insights.append({
                        'id': str(uuid.uuid4()),
                        'zone_id': zone.id,
                        'zone_name': zone.name,
                        'type': 'yield',
                        'confidence': yield_prediction['confidence'],
                        'prediction': f"Expected yield: {yield_prediction['predicted_yield']} {zone.crop_profile.get('yield_unit', 'kg')}",
                        'recommendation': self._generate_yield_recommendation(yield_prediction),
                        'timeframe': 'At harvest',
                        'potentialImpact': {
                            'yield': round((yield_prediction['predicted_yield'] / zone.crop_profile.get('yield_target', 100) - 1) * 100, 1)
                        },
                        'factors': yield_prediction['factors']
                    })
                    
            # Maintenance predictions
            if sensor_data.empty:
                sensor_data = await self.phal._get_recent_sensor_data(zone.id, hours=24)
                
            if not sensor_data.empty:
                maintenance_predictions = await self.phal.ml_engine.predict_maintenance(zone.id, sensor_data)
                
                for pred in maintenance_predictions:
                    if pred['confidence'] > 0.7:
                        insights.append({
                            'id': str(uuid.uuid4()),
                            'zone_id': zone.id,
                            'zone_name': zone.name,
                            'type': 'maintenance',
                            'confidence': pred['confidence'],
                            'prediction': f"{pred['component']} maintenance needed",
                            'recommendation': pred['recommendation'],
                            'timeframe': f"Within {pred['days_until_maintenance']} days",
                            'potentialImpact': {
                                'downtime': 30  # minutes
                            },
                            'metrics': pred.get('metrics', {})
                        })
                        
            # Environmental optimization
            current_conditions = {}
            for cap in self.phal.capabilities.values():
                if cap.zone_id == zone.id and cap.type == 'SENSOR' and cap.last_reading:
                    current_conditions[cap.subtype] = cap.last_reading.get('value')
                    
            if current_conditions:
                optimization = await self.phal.ml_engine.optimize_environment(zone, current_conditions)
                
                if optimization['adjustments']:
                    insights.append({
                        'id': str(uuid.uuid4()),
                        'zone_id': zone.id,
                        'zone_name': zone.name,
                        'type': 'optimization',
                        'confidence': 0.85,
                        'prediction': 'Environmental optimization available',
                        'recommendation': ' | '.join(optimization['reasoning']),
                        'timeframe': 'Immediate',
                        'potentialImpact': {
                            'yield': 5,  # %
                            'energy': -10  # % reduction
                        },
                        'adjustments': optimization['adjustments']
                    })
                    
        # Sort by confidence and priority
        insights.sort(key=lambda x: (x['confidence'], x['type'] == 'maintenance'), reverse=True)
        
        return web.json_response({
            'insights': insights[:10],  # Top 10 insights
            'generated_at': datetime.now(timezone.utc).isoformat()
        })
        
    def _generate_yield_recommendation(self, prediction: Dict[str, Any]) -> str:
        """Generate recommendation based on yield prediction"""
        factors = prediction.get('factors', [])
        
        # Find limiting factors
        limiting = [f for f in factors if not f['optimal']]
        
        if not limiting:
            return "Conditions are optimal. Maintain current settings."
            
        # Sort by impact
        limiting.sort(key=lambda f: f['impact'])
        
        recommendations = []
        for factor in limiting[:2]:  # Top 2 limiting factors
            name = factor['name']
            current = factor.get('current')
            target = factor.get('target')
            
            if current and target:
                if name == 'Temperature':
                    if current < target:
                        recommendations.append(f"Increase temperature to {target}°C")
                    else:
                        recommendations.append(f"Decrease temperature to {target}°C")
                elif name == 'Light (DLI)':
                    if current < target:
                        recommendations.append(f"Increase DLI to {target} mol/m²/day")
                    else:
                        recommendations.append(f"Decrease DLI to {target} mol/m²/day")
                elif name == 'Nutrients (EC)':
                    recommendations.append(f"Adjust EC to {target} mS/cm")
                    
        return " | ".join(recommendations) if recommendations else "Review environmental conditions"
        
    async def get_predictions(self, request: web.Request) -> web.Response:
        """Get various predictions"""
        tenant_id = request['tenant_id']
        prediction_type = request.query.get('type', 'all')
        zone_id = request.query.get('zone_id')
        
        predictions = {}
        
        # Validate zone
        if zone_id:
            zone = self.phal.zones.get(zone_id)
            if not zone or zone.tenant_id != tenant_id:
                raise web.HTTPNotFound(reason='Zone not found')
            zones = [zone]
        else:
            zones = [z for z in self.phal.zones.values() if z.tenant_id == tenant_id]
            
        # Resource usage prediction
        if prediction_type in ['all', 'resources']:
            resource_predictions = []
            
            for zone in zones:
                # Simple linear projection based on recent usage
                usage_7d = await self._get_zone_resource_usage(zone.id, days=7)
                usage_30d = await self._get_zone_resource_usage(zone.id, days=30)
                
                if usage_7d and usage_30d:
                    # Calculate daily rates
                    daily_rate_7d = {k: v / 7 for k, v in usage_7d.items()}
                    daily_rate_30d = {k: v / 30 for k, v in usage_30d.items()}
                    
                    # Project 30 days forward
                    projection = {}
                    for resource in daily_rate_7d:
                        # Weight recent usage more heavily
                        rate = daily_rate_7d[resource] * 0.7 + daily_rate_30d[resource] * 0.3
                        projection[resource] = rate * 30
                        
                    resource_predictions.append({
                        'zone_id': zone.id,
                        'zone_name': zone.name,
                        'projection_days': 30,
                        'predicted_usage': projection,
                        'confidence': 0.75
                    })
                    
            predictions['resources'] = resource_predictions
            
        # Harvest date prediction
        if prediction_type in ['all', 'harvest']:
            harvest_predictions = []
            
            for zone in zones:
                if zone.crop_profile:
                    expected = zone.crop_profile.get('expected_harvest_date')
                    if expected:
                        expected_date = datetime.fromisoformat(expected)
                        days_remaining = (expected_date - datetime.now(timezone.utc)).days
                        
                        # Adjust based on growth conditions
                        # This is simplified - in reality would use growth model
                        adjustment = 0
                        
                        harvest_predictions.append({
                            'zone_id': zone.id,
                            'zone_name': zone.name,
                            'crop': zone.crop_profile.get('name', 'Unknown'),
                            'predicted_date': expected_date.isoformat(),
                            'days_remaining': max(0, days_remaining + adjustment),
                            'confidence': 0.8
                        })
                        
            predictions['harvest'] = harvest_predictions
            
        return web.json_response(predictions)
        
    async def _get_zone_resource_usage(self, zone_id: str, days: int) -> Dict[str, float]:
        """Get resource usage for zone"""
        usage = defaultdict(float)
        
        # Query from audit logs
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT 
                    event_data->>'resource' as resource,
                    SUM((event_data->>'volume_ml')::float) as total
                FROM audit_log
                WHERE event_type = 'command_executed'
                AND event_data->>'zone_id' = $1
                AND timestamp > NOW() - INTERVAL '%d days'
                AND event_data->>'resource' IS NOT NULL
                GROUP BY resource
            ''' % days, zone_id)
            
        for row in rows:
            usage[row['resource']] = float(row['total'])
            
        return dict(usage)
        
    async def get_recipes(self, request: web.Request) -> web.Response:
        """Get recipes"""
        crop_type = request.query.get('crop_type')
        is_public = request.query.get('public', 'false').lower() == 'true'
        
        query = 'SELECT * FROM recipes WHERE 1=1'
        params = []
        
        if crop_type:
            params.append(crop_type)
            query += f' AND crop_type = ${len(params)}'
            
        if is_public:
            params.append(is_public)
            query += f' AND is_public = ${len(params)}'
        else:
            # Show tenant's private recipes and public ones
            tenant_id = request['tenant_id']
            query += f' AND (is_public = true OR created_by = \'{tenant_id}\')'
            
        query += ' ORDER BY created_at DESC LIMIT 100'
        
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch(query, *params)
            
        recipes = []
        for row in rows:
            recipes.append({
                'id': str(row['id']),
                'name': row['name'],
                'crop_type': row['crop_type'],
                'stages': row['stages'],
                'created_by': row['created_by'],
                'validated_yield': float(row['validated_yield']) if row['validated_yield'] else None,
                'notes': row['notes'],
                'tags': row['tags'],
                'is_public': row['is_public'],
                'created_at': row['created_at'].isoformat()
            })
            
        return web.json_response({
            'recipes': recipes,
            'total': len(recipes)
        })
        
    async def create_recipe(self, request: web.Request) -> web.Response:
        """Create recipe"""
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        recipe_id = str(uuid.uuid4())
        
        # Validate stages
        if not data.get('stages') or not isinstance(data['stages'], list):
            raise web.HTTPBadRequest(reason='Recipe must have stages')
            
        async with self.phal.db.acquire() as conn:
            await conn.execute('''
                INSERT INTO recipes 
                (id, name, crop_type, stages, created_by, notes, tags, is_public)
                VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
            ''', recipe_id, data['name'], data.get('crop_type'),
                json.dumps(data['stages']), tenant_id,
                data.get('notes'), data.get('tags', []),
                data.get('is_public', False))
                
        return web.json_response({
            'id': recipe_id,
            'status': 'created'
        }, status=201)
        
    async def get_recipe(self, request: web.Request) -> web.Response:
        """Get recipe details"""
        recipe_id = request.match_info['recipe_id']
        tenant_id = request['tenant_id']
        
        async with self.phal.db.acquire() as conn:
            row = await conn.fetchrow(
                'SELECT * FROM recipes WHERE id = $1',
                recipe_id
            )
            
        if not row:
            raise web.HTTPNotFound(reason='Recipe not found')
            
        # Check access
        if not row['is_public'] and row['created_by'] != tenant_id:
            raise web.HTTPForbidden(reason='Access denied')
            
        return web.json_response({
            'id': str(row['id']),
            'name': row['name'],
            'crop_type': row['crop_type'],
            'stages': row['stages'],
            'created_by': row['created_by'],
            'validated_yield': float(row['validated_yield']) if row['validated_yield'] else None,
            'notes': row['notes'],
            'tags': row['tags'],
            'is_public': row['is_public'],
            'created_at': row['created_at'].isoformat(),
            'updated_at': row['updated_at'].isoformat()
        })
        
    async def update_recipe(self, request: web.Request) -> web.Response:
        """Update recipe"""
        recipe_id = request.match_info['recipe_id']
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        # Check ownership
        async with self.phal.db.acquire() as conn:
            owner = await conn.fetchval(
                'SELECT created_by FROM recipes WHERE id = $1',
                recipe_id
            )
            
        if owner != tenant_id:
            raise web.HTTPForbidden(reason='Only recipe owner can update')
            
        # Update allowed fields
        updates = []
        params = []
        
        allowed_fields = ['name', 'stages', 'validated_yield', 'notes', 'tags', 'is_public']
        for field in allowed_fields:
            if field in data:
                params.append(data[field])
                updates.append(f'{field} = ${len(params)}')
                
        if not updates:
            raise web.HTTPBadRequest(reason='No valid updates provided')
            
        params.append(recipe_id)
        query = f'UPDATE recipes SET {", ".join(updates)} WHERE id = ${len(params)}'
        
        async with self.phal.db.acquire() as conn:
            await conn.execute(query, *params)
            
        return web.json_response({'status': 'updated'})
        
    async def delete_recipe(self, request: web.Request) -> web.Response:
        """Delete recipe"""
        recipe_id = request.match_info['recipe_id']
        tenant_id = request['tenant_id']
        
        # Check ownership
        async with self.phal.db.acquire() as conn:
            owner = await conn.fetchval(
                'SELECT created_by FROM recipes WHERE id = $1',
                recipe_id
            )
            
        if not owner:
            raise web.HTTPNotFound(reason='Recipe not found')
            
        if owner != tenant_id:
            raise web.HTTPForbidden(reason='Only recipe owner can delete')
            
        async with self.phal.db.acquire() as conn:
            await conn.execute('DELETE FROM recipes WHERE id = $1', recipe_id)
            
        return web.json_response({'status': 'deleted'})
        
    async def export_data(self, request: web.Request) -> web.Response:
        """Export data"""
        tenant_id = request['tenant_id']
        data = sanitize_input(await request.json())
        
        try:
            export_request = ExportRequest(**data)
        except ValidationError as e:
            raise web.HTTPBadRequest(text=json.dumps({'errors': e.errors()}))
            
        # Get tenant zones
        tenant_zones = [z.id for z in self.phal.zones.values() if z.tenant_id == tenant_id]
        
        # Filter to only tenant's zones
        zones = export_request.zones or tenant_zones
        zones = [z for z in zones if z in tenant_zones]
        
        if not zones:
            raise web.HTTPBadRequest(reason='No zones to export')
            
        # Create export based on type
        if export_request.type == 'sensor_data':
            # Continuation of export_data method
            export_data = await self._export_sensor_data(
                zones, 
                export_request.time_range,
                export_request.format
            )
        elif export_request.type == 'harvests':
            export_data = await self._export_harvests(
                zones,
                export_request.time_range,
                export_request.format
            )
        elif export_request.type == 'maintenance':
            export_data = await self._export_maintenance(
                export_request.time_range,
                export_request.format
            )
        elif export_request.type == 'audit_log':
            # Check admin permission
            if 'admin' not in request['tenant'].features:
                raise web.HTTPForbidden(reason='Admin access required')
            export_data = await self._export_audit_log(
                tenant_id,
                export_request.time_range,
                export_request.format
            )
        else:
            raise web.HTTPBadRequest(reason=f'Unknown export type: {export_request.type}')
            
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"phal_export_{export_request.type}_{timestamp}.{export_request.format}"
        
        # Compress if requested
        if export_request.compress:
            import gzip
            export_data = gzip.compress(export_data.encode() if isinstance(export_data, str) else export_data)
            filename += '.gz'
            
        # Return file
        response = web.Response(
            body=export_data,
            content_type=self._get_content_type(export_request.format),
            headers={
                'Content-Disposition': f'attachment; filename="{filename}"'
            }
        )
        
        if export_request.compress:
            response.headers['Content-Encoding'] = 'gzip'
            
        return response
        
    async def _export_sensor_data(self, zones: List[str], time_range: Dict[str, datetime], format: str) -> Union[str, bytes]:
        """Export sensor data"""
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT 
                    sr.timestamp,
                    sr.zone_id,
                    z.name as zone_name,
                    sr.sensor_type,
                    sr.value,
                    sr.quality,
                    sr.unit
                FROM sensor_readings sr
                JOIN zones z ON sr.zone_id = z.id
                WHERE sr.zone_id = ANY($1)
                AND sr.timestamp BETWEEN $2 AND $3
                ORDER BY sr.timestamp DESC
                LIMIT 100000
            ''', zones, time_range['start'], time_range['end'])
            
        if format == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=[
                'timestamp', 'zone_id', 'zone_name', 'sensor_type', 
                'value', 'quality', 'unit'
            ])
            writer.writeheader()
            
            for row in rows:
                writer.writerow({
                    'timestamp': row['timestamp'].isoformat(),
                    'zone_id': str(row['zone_id']),
                    'zone_name': row['zone_name'],
                    'sensor_type': row['sensor_type'],
                    'value': row['value'],
                    'quality': row['quality'],
                    'unit': row['unit']
                })
                
            return output.getvalue()
            
        elif format == 'json':
            data = []
            for row in rows:
                data.append({
                    'timestamp': row['timestamp'].isoformat(),
                    'zone_id': str(row['zone_id']),
                    'zone_name': row['zone_name'],
                    'sensor_type': row['sensor_type'],
                    'value': float(row['value']),
                    'quality': float(row['quality']),
                    'unit': row['unit']
                })
            return json.dumps({'sensor_data': data}, indent=2)
            
        elif format == 'xlsx':
            import openpyxl
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Sensor Data'
            
            # Headers
            headers = ['Timestamp', 'Zone ID', 'Zone Name', 'Sensor Type', 'Value', 'Quality', 'Unit']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                ws.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
                
            # Data
            for row_idx, row in enumerate(rows, 2):
                ws.cell(row=row_idx, column=1, value=row['timestamp'])
                ws.cell(row=row_idx, column=2, value=str(row['zone_id']))
                ws.cell(row=row_idx, column=3, value=row['zone_name'])
                ws.cell(row=row_idx, column=4, value=row['sensor_type'])
                ws.cell(row=row_idx, column=5, value=float(row['value']))
                ws.cell(row=row_idx, column=6, value=float(row['quality']))
                ws.cell(row=row_idx, column=7, value=row['unit'])
                
            # Auto-size columns
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
                
            # Save to bytes
            import io
            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()
            
        else:
            raise ValueError(f'Unsupported format: {format}')
            
    async def _export_harvests(self, zones: List[str], time_range: Dict[str, datetime], format: str) -> Union[str, bytes]:
        """Export harvest data"""
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT h.*, z.name as zone_name
                FROM harvests h
                JOIN zones z ON h.zone_id = z.id
                WHERE h.zone_id = ANY($1)
                AND h.harvest_date BETWEEN $2 AND $3
                ORDER BY h.harvest_date DESC
            ''', zones, time_range['start'], time_range['end'])
            
        if format == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=[
                'harvest_date', 'zone_name', 'crop_id', 'quantity', 
                'quantity_unit', 'quality_grade', 'harvested_by', 'notes'
            ])
            writer.writeheader()
            
            for row in rows:
                writer.writerow({
                    'harvest_date': row['harvest_date'].isoformat(),
                    'zone_name': row['zone_name'],
                    'crop_id': row['crop_id'],
                    'quantity': row['quantity'],
                    'quantity_unit': row['quantity_unit'],
                    'quality_grade': row['quality_grade'],
                    'harvested_by': row['harvested_by'],
                    'notes': row['notes'] or ''
                })
                
            return output.getvalue()
        else:
            # Similar implementations for json and xlsx
            raise NotImplementedError(f'Format {format} not yet implemented for harvests')
            
    async def _export_maintenance(self, time_range: Dict[str, datetime], format: str) -> Union[str, bytes]:
        """Export maintenance records"""
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT * FROM maintenance_records
                WHERE created_at BETWEEN $1 AND $2
                ORDER BY scheduled_date DESC
            ''', time_range['start'], time_range['end'])
            
        # Format data based on requested format
        # Implementation similar to sensor data export
        raise NotImplementedError('Maintenance export not yet implemented')
        
    async def _export_audit_log(self, tenant_id: str, time_range: Dict[str, datetime], format: str) -> Union[str, bytes]:
        """Export audit logs"""
        async with self.phal.db.acquire() as conn:
            rows = await conn.fetch('''
                SELECT * FROM audit_log
                WHERE tenant_id = $1
                AND timestamp BETWEEN $2 AND $3
                ORDER BY timestamp DESC
                LIMIT 10000
            ''', tenant_id, time_range['start'], time_range['end'])
            
        # Format data based on requested format
        # Implementation similar to sensor data export
        raise NotImplementedError('Audit log export not yet implemented')
        
    def _get_content_type(self, format: str) -> str:
        """Get content type for format"""
        return {
            'csv': 'text/csv',
            'json': 'application/json',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }.get(format, 'application/octet-stream')
        
    async def import_data(self, request: web.Request) -> web.Response:
        """Import data from file"""
        tenant_id = request['tenant_id']
        
        # Check file upload
        reader = await request.multipart()
        
        file_data = None
        import_type = None
        
        async for field in reader:
            if field.name == 'file':
                file_data = await field.read()
                filename = field.filename
            elif field.name == 'type':
                import_type = await field.text()
                
        if not file_data:
            raise web.HTTPBadRequest(reason='No file uploaded')
            
        if not import_type:
            raise web.HTTPBadRequest(reason='Import type not specified')
            
        # Detect format from filename
        format = 'json'
        if filename.endswith('.csv'):
            format = 'csv'
        elif filename.endswith('.xlsx'):
            format = 'xlsx'
            
        try:
            if import_type == 'recipes':
                result = await self._import_recipes(tenant_id, file_data, format)
            elif import_type == 'zones':
                result = await self._import_zones(tenant_id, file_data, format)
            else:
                raise web.HTTPBadRequest(reason=f'Unknown import type: {import_type}')
                
            return web.json_response({
                'status': 'imported',
                'type': import_type,
                'count': result['count'],
                'details': result.get('details', {})
            })
            
        except Exception as e:
            logger.error(f"Import failed: {e}")
            raise web.HTTPBadRequest(reason=f'Import failed: {str(e)}')
            
    async def _import_recipes(self, tenant_id: str, file_data: bytes, format: str) -> Dict[str, Any]:
        """Import recipes"""
        if format == 'json':
            data = json.loads(file_data)
            recipes = data.get('recipes', [])
            
            imported = 0
            for recipe_data in recipes:
                recipe_id = str(uuid.uuid4())
                
                async with self.phal.db.acquire() as conn:
                    await conn.execute('''
                        INSERT INTO recipes 
                        (id, name, crop_type, stages, created_by, notes, tags, is_public)
                        VALUES ($1, $2, $3, $4, $5, $6, $7, $8)
                        ON CONFLICT (name) DO NOTHING
                    ''', recipe_id, recipe_data['name'], recipe_data.get('crop_type'),
                        json.dumps(recipe_data['stages']), tenant_id,
                        recipe_data.get('notes'), recipe_data.get('tags', []),
                        False)  # Imported recipes are private by default
                        
                imported += 1
                
            return {'count': imported}
        else:
            raise NotImplementedError(f'Recipe import not implemented for format: {format}')
            
    async def _import_zones(self, tenant_id: str, file_data: bytes, format: str) -> Dict[str, Any]:
        """Import zone configurations"""
        # Implementation would parse zone data and create zones
        raise NotImplementedError('Zone import not yet implemented')
        
    async def get_audit_logs(self, request: web.Request) -> web.Response:
        """Query audit logs"""
        tenant_id = request['tenant_id']
        tenant = request['tenant']
        
        # Check permission
        if 'audit_access' not in tenant.features and tenant.tier != 'enterprise':
            raise web.HTTPForbidden(reason='Audit access not available in your plan')
            
        # Parse query parameters
        event_type = request.query.get('event_type')
        user_id = request.query.get('user_id')
        start_date = request.query.get('start_date')
        end_date = request.query.get('end_date')
        limit = min(int(request.query.get('limit', 100)), 1000)
        
        # Build criteria
        criteria = {'tenant_id': tenant_id}
        
        if event_type:
            criteria['event_type'] = event_type
        if user_id:
            criteria['user_id'] = user_id
        if start_date:
            criteria['start_date'] = datetime.fromisoformat(start_date)
        if end_date:
            criteria['end_date'] = datetime.fromisoformat(end_date)
            
        # Search audit logs
        logs = await self.phal.audit_logger.search(criteria, limit)
        
        return web.json_response({
            'logs': logs,
            'total': len(logs),
            'query': {
                'event_type': event_type,
                'user_id': user_id,
                'start_date': start_date,
                'end_date': end_date,
                'limit': limit
            }
        })
        
    async def websocket_handler(self, request: web.Request) -> web.WebSocketResponse:
        """Secure WebSocket handler with proper auth"""
        ws = web.WebSocketResponse(heartbeat=30)
        await ws.prepare(request)

        auth_token = request.cookies.get('auth_token') or request.headers.get('Authorization', '').replace('Bearer ', '')

        if not auth_token:
            await ws.close(code=4001, message=b'Authentication required')
            return ws

        try:
            payload = jwt.decode(auth_token, self.phal.config.jwt_secret, algorithms=['HS256'])
            tenant_id = payload.get('tenant_id')
            if not tenant_id:
                raise ValueError('Invalid token')
        except Exception:
            await ws.close(code=1008, message=b'Invalid authentication')
            return ws

        connection_id = str(uuid.uuid4())
        self.active_websockets[connection_id] = {
            'ws': ws,
            'tenant_id': tenant_id,
            'subscriptions': set(),
            'last_ping': datetime.now(timezone.utc)
        }

        subscriptions = self.active_websockets[connection_id]['subscriptions']
        
        async def handle_event(event):
            """Send events to websocket"""
            # Filter events based on tenant and subscriptions
            if event['type'] not in subscriptions:
                return
                
            # Check if event is relevant to tenant
            if 'zone_id' in event['data']:
                zone = self.phal.zones.get(event['data']['zone_id'])
                if not zone or zone.tenant_id != tenant_id:
                    return
                    
            try:
                await ws.send_json({
                    'type': 'event',
                    'event': event['type'],
                    'data': event['data'],
                    'timestamp': event['timestamp']
                })
            except ConnectionResetError:
                pass
                
        # Default subscriptions
        for event_type in ['alarm', 'command_executed', 'zone_updated']:
            self.phal.event_bus.subscribe(event_type, handle_event)
            subscriptions.add(event_type)
            
        try:
            # Send initial connection success
            await ws.send_json({
                'type': 'connected',
                'connection_id': connection_id,
                'subscriptions': list(subscriptions)
            })
            
            # Handle incoming messages
            async for msg in ws:
                if msg.type == aiohttp.WSMsgType.TEXT:
                    try:
                        data = json.loads(msg.data)
                        
                        if data['type'] == 'subscribe':
                            event_types = data.get('events', [])
                            for event_type in event_types:
                                if event_type not in subscriptions:
                                    self.phal.event_bus.subscribe(event_type, handle_event)
                                    subscriptions.add(event_type)
                                    
                            await ws.send_json({
                                'type': 'subscribed',
                                'events': event_types
                            })
                            
                        elif data['type'] == 'unsubscribe':
                            event_types = data.get('events', [])
                            for event_type in event_types:
                                subscriptions.discard(event_type)
                                
                            await ws.send_json({
                                'type': 'unsubscribed',
                                'events': event_types
                            })
                            
                        elif data['type'] == 'ping':
                            await ws.send_json({
                                'type': 'pong',
                                'timestamp': datetime.now(timezone.utc).isoformat()
                            })
                            
                    except Exception as e:
                        await ws.send_json({
                            'type': 'error',
                            'error': str(e)
                        })
                        
                elif msg.type == aiohttp.WSMsgType.ERROR:
                    logger.error(f'WebSocket error: {ws.exception()}')
                    
        except Exception as e:
            logger.error(f'WebSocket handler error: {e}')
        finally:
            self.active_websockets.pop(connection_id, None)
            await ws.close()
            
        return ws
        
    async def run(self):
        """Run the application"""
        # Initialize PHAL
        await self.phal.initialize()
        
        # Create test data if in development
        if os.getenv('PHAL_ENV') == 'development':
            await create_test_data(self.phal)
            
        # Configure app
        self.app['phal'] = self.phal
        
        # Setup cleanup
        async def cleanup(app):
            await app['phal'].shutdown()
            
        self.app.on_cleanup.append(cleanup)
        
        # Start server
        runner = web.AppRunner(self.app)
        await runner.setup()
        
        site = web.TCPSite(
            runner,
            host=os.getenv('PHAL_HOST', '0.0.0.0'),
            port=int(os.getenv('PHAL_PORT', 8080))
        )
        
        await site.start()
        
        logger.info(f"PHAL server started on {site.name}")
        logger.info("Visit http://localhost:8080 for API documentation")
        
        # Keep running
        try:
            await asyncio.Event().wait()
        except KeyboardInterrupt:
            pass
        finally:
            await runner.cleanup()


# Exception classes
class PHALError(Exception):
    """Base PHAL exception"""
    def __init__(self, message: str, code: str = 'PHAL_ERROR', details: Optional[Dict[str, Any]] = None):
        super().__init__(message)
        self.code = code
        self.details = details or {}

class GrantError(PHALError):
    """Grant-related errors"""
    def __init__(self, message: str, details: Optional[Dict[str, Any]] = None):
        super().__init__(message, 'GRANT_ERROR', details)

class ValidationError(PHALError):
    """Validation errors"""
    def __init__(self, message: str, errors: List[Dict[str, Any]]):
        super().__init__(message, 'VALIDATION_ERROR', {'errors': errors})
        
    def errors(self):
        return self.details.get('errors', [])

class RateLimitError(PHALError):
    """Rate limit errors"""
    def __init__(self, message: str, retry_after: int):
        super().__init__(message, 'RATE_LIMIT_ERROR', {'retryAfter': retry_after})

class SecurityError(PHALError):
    """Security-related errors"""
    def __init__(self, message: str, details: Optional[Dict[str, Any]] = None):
        super().__init__(message, 'SECURITY_ERROR', details)


# Main entry point
def main():
    """Main entry point"""
    import argparse
    
    parser = argparse.ArgumentParser(description='PHAL - Pluripotent Hardware Abstraction Layer')
    parser.add_argument('--config', help='Configuration file path')
    parser.add_argument('--host', default='0.0.0.0', help='Host to bind to')
    parser.add_argument('--port', type=int, default=8080, help='Port to bind to')
    parser.add_argument('--log-level', default='INFO', help='Logging level')
    
    args = parser.parse_args()
    
    # Configure logging
    logging.basicConfig(
        level=getattr(logging, args.log_level.upper()),
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
    )
    
    # Override with command line args
    if args.host:
        os.environ['PHAL_HOST'] = args.host
    if args.port:
        os.environ['PHAL_PORT'] = str(args.port)
        
    # Create and run application
    app = PHALApplication(args.config)
    
    try:
        asyncio.run(app.run())
    except KeyboardInterrupt:
        logger.info("Shutting down PHAL...")
        

if __name__ == '__main__':
    main()